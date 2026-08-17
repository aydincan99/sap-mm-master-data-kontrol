# -*- coding: utf-8 -*-
"""
Malzeme Master Veri Kalite Kontrolü
====================================
SAP MM malzeme ana verisini (MARA) ve alternatif ölçü birimlerini (MARM)
ALTI kural üzerinden denetler; bulguları Excel raporları, Power BI'a hazır
bir veri modeli ve tarayıcıda çalışan bir HTML panel olarak üretir.

Kurallar:
 K1) Akıllı kod eşleşmesi — Ürün kodu, bağlı olduğu üst kodla (RC) beyaz
     listedeki alanlarda birebir aynı olmalıdır. Üst kodun boş olması,
     68 ile başlayan malzemede üst kod bulunması ve üst birim düzeyi silme
     işaretinin ürünlere kaskad edilmemesi de bu kural kapsamındadır.
 K2) Tanım ↔ ek veri alanı — Malzeme kısa metninin sonundaki kod (LF, RU …)
     ile ilgili alan iki yönde tutarlı olmalıdır; T26 gibi kodlar tanımın
     sonunda hiç bulunamaz.
 K3) Yasaklı kelimeler — NEW, YENİ/YENI ile 2026/2027 tanımın her yerinde;
     6 ve 7 yalnızca tanımın sonunda yasaklıdır.
 K4) Çokluk ↔ MARM sayacı — Tanımın sonundaki çokluk (…x24), MARM'daki
     çevrim sayacı değerlerinden biriyle eşleşmelidir. Muaf ürün hiyerarşisi
     listesindeki malzemeler kapsam dışıdır.
 K5) Üst kod benzersizliği — İki üst kod, ayırt edici alan kombinasyonunun
     tamamında birebir aynı olamaz.
 K6) Kod sonu ↔ Menşei — Ürün kodunun son iki hanesi, Menşei alanındaki
     sayısal kodla aynı olmalıdır (10=TR, 21=MY VAS …); boş menşei de
     bulgudur.

Ayrıca BİLGİ (hata değil) olarak raporlananlar:
  Alt tarifler — aynı tanımı paylaşan birden çok malzeme kodu. Sağlık
  hesabına dahil EDİLMEZ; yalnızca bilgi amaçlı listelenir.

Kullanım:
    pip install pandas openpyxl xlsxwriter
    python malzeme_kalite_kontrol.py

Dosya adlarını ve ayarları aşağıdaki AYARLAR bölümünden değiştirebilirsin.
"""

import re
import unicodedata
import re as _re
import numpy as np
import pandas as pd

# Colab ortamı tespiti
try:
    from google.colab import files as colab_files  # type: ignore
    COLAB = True
except ImportError:
    COLAB = False

# =====================================================================
# AYARLAR — kendi ortamına göre burayı düzenle
# =====================================================================

URUN_DOSYASI = "urun_kodlari.xlsx"   # ürün kodları export'u
UST_DOSYASI  = "ust_kodlar.xlsx"     # üst kodlar export'u
CIKTI_EXCEL  = "kalite_kontrol_sonuc.xlsx"
CIKTI_DASHBOARD = "veri_kalite_dashboard.html"  # otomatik üretilen görsel panel
CIKTI_BULGU_XLSX = "powerbi_bulgular.xlsx"      # biçimlendirilmiş bulgu tablosu
CIKTI_DUZELTILMIS = "powerbi_bulgular_duzeltilmis.xlsx"  # düzeltilmiş veri seti
CIKTI_KATEGORI = "kategori_analizi.xlsx"        # kategori bazında sayım/analiz
# Ürün ve üst kodlar TEK dosyada geliyorsa bu dosya kullanılır: malzeme
# kodu RC ile başlıyorsa üst kod, rakamla başlıyorsa ürün kodu sayılır.
BIRLESIK_DOSYA = "birlesik_liste.xlsx"
UST_KOD_ONEKI = "RC"

MARM_DOSYASI = "marm.xlsx"   # alternatif ölçü birimleri (MARM) — çokluk kontrolü
ISTISNA_DOSYASI = "istisnalar.xlsx"   # onaylı istisnalar (varsa bulgulardan düşülür)
AYARLAR_DOSYASI = "ayarlar.xlsx"      # kod listeleri (varsa buradaki değerler geçerli olur)
GECMIS_DOSYASI  = "calistirma_gecmisi.xlsx"  # her koşumun özet satırı buraya eklenir

# Çıktı dili otomatik algılanır: dosya başlıkları İngilizce ise "EN",
# değilse "TR". (İstenirse "TR"/"EN" yazarak sabitlenebilir.)
DIL = "AUTO"

# İngilizce SAP export başlıkları -> içerideki Türkçe adlar
EN2TR_KOLON = {
    "Material": "Malzeme", "Created on": "Yaratma tarihi",
    "Created On": "Yaratma tarihi", "Time created": "Oluşturma saati",
    "Created at": "Oluşturma saati", "Last Changed On": "Son değişiklik tarihi",
    "Last change": "Son değişiklik tarihi", "Time changed": "Son değişiklik saati",
    "Material type": "Malzeme türü", "Material Type": "Malzeme türü",
    "Material Group": "Mal grubu", "Material group": "Mal grubu",
    "Base Unit of Measure": "Temel ölçü birimi", "Base Unit": "Temel ölçü birimi",
    "Base material": "Temel malzeme", "Base Material": "Temel malzeme",
    "Reporting Sub Brand": "Raporlama Alt Markası",
    "Reporting Sub-Brand": "Raporlama Alt Markası",
    "Origin": "Menşei", "Reference Product Size": "Referans Ürün Boyutu",
    "S&OP Category": "S&OP Kategorisi", "Market": "Pazar",
    "Reporting Brand": "Raporlama Markası", "Brand2": "Marka2",
    "Brand 2": "Marka2", "Additional Field": "Ek Alan",
    "Add. Field": "Ek Alan", "Brand3": "Marka3", "Brand 3": "Marka3",
    "Product Size": "Ürün Boyutu", "Packaging Type": "Ambalaj Tipi",
    "Package Type": "Ambalaj Tipi", "SKU Group": "SKU Grup",
    "Brand1": "Marka1", "Brand 1": "Marka1", "Variant": "Varyant",
    "QR Barcode": "Kare Barkod", "Square Barcode": "Kare Barkod",
    "Material Description": "Malzeme tanımı",
    "Material Descriptions": "Malzeme tanımı",
    "Material short text": "Malzeme tanımı",
}


def dil_algila_ve_uyarla(df: pd.DataFrame) -> pd.DataFrame:
    """Başlıklar İngilizce ise Türkçe iç adlara çevirir ve DIL'i EN yapar."""
    global DIL
    en_hit = sum(1 for c in df.columns if str(c).strip() in EN2TR_KOLON)
    if en_hit >= 3:
        df = df.rename(columns={c: EN2TR_KOLON[str(c).strip()]
                                for c in df.columns
                                if str(c).strip() in EN2TR_KOLON})
        if DIL == "AUTO":
            DIL = "EN"
            print("Bilgi: İngilizce başlıklar algılandı — çıktılar İngilizce üretilecek.")
    elif DIL == "AUTO":
        DIL = "TR"
    return df

# KAPSAM FİLTRESİ — analiz yalnızca şu türlerde yapılır:
# ürün 6/7, üst kod 15. 1-5 dahil diğer türler HATA SAYILMAZ, atlanır.
GECERLI_URUN_TURLERI = {"6", "7"}
GECERLI_UST_TURLERI = {"15"}

# Kural 1'de KARŞILAŞTIRILACAK alanlar (beyaz liste) — yalnızca bunlar
# hata üretir; tarih/saat, mal grubu, ölçü birimi gibi diğer sütunlar
# hiçbir zaman bulgu oluşturmaz. Menşei/Pazar/Ek Alan/Ambalaj Tipi
# K2 kuralının konusudur, K1'de karşılaştırılmaz.
KIYAS_ALANLARI = [
    "Raporlama Alt Markası", "Referans Ürün Boyutu", "S&OP Kategorisi",
    "Raporlama Markası", "Marka2", "Marka3", "Ürün Boyutu",
    "SKU Grup", "Marka1", "Varyant", "Kare Barkod",
    "Ek Alan",  # ürün ile üst kod arasında birebir eşleşmeli (örn. LF)
]
ISTISNA_ALANLAR = ["Menşei", "Pazar", "Ambalaj Tipi"]

# SAP bayrak (işaret) alanları: değer "X" ise işaretli, boş ise işaretsiz.
# Karşılaştırma X/boş mantığıyla yapılır; "x", "X " gibi yazımlar eşitlenir.
BAYRAK_ALANLARI = ["Kare Barkod"]

# Bu alanlarda BOŞ değer hata sayılmaz: yalnızca iki taraf da doluyken
# ve farklıyken bulgu üretilir (örn. SKU Grup boş bırakılmış olabilir).
BOS_HATA_SAYILMAZ = ["SKU Grup"]

# Karşılaştırmaya hiç girmeyecek teknik alanlar (kimlik/tarih/saat vb.):
TEKNIK_ALANLAR = [
    "Malzeme", "Yaratma tarihi", "Oluşturma saati",
    "Son değişiklik tarihi", "Son değişiklik saati",
    "Temel malzeme", "Malzeme türü",
]

# Kural 1-2 hangi malzeme türlerine uygulanacak?
# Türlerin içinde "6" veya "7" geçenleri yakalar (örn. Z006, ZFG7, 6, 7...).
# Tam eşleşme istersen: MALZEME_TURU_FILTRESI = ["6", "7"] yapıp
# tam_eslesme_turu = True olarak değiştir.
MALZEME_TURU_FILTRESI = ["6", "7"]
TAM_ESLESME_TURU = False

# Kural 2: Her alanın geçerli kodları ve (varsa) uzun metin karşılıkları.
# Tanımın SONUNDA bu kodlardan biri geçiyorsa ilgili alan dolu olmalı.
# Kod hem Menşei hem Pazar'da varsa (EG, RU, TR gibi) birinde dolu olması yeterli.
KURAL2_KATEGORILER = {
    # Menşei alanı SAYISAL kod tutar (10 = TR, 30 = RU ...). Malzeme
    # tanımının sonunda bu sayılar aranmaz; sözlük yalnızca alan değerini
    # okumak/doğrulamak için tutulur.
    "Menşei": {
        "10": "TR", "11": "TR VAS", "20": "MY", "21": "MY VAS",
        "30": "RU", "31": "RU VAS", "40": "EG", "41": "EG VAS",
        "50": "ID", "99": "TG",
    },
    # Pazar alanı ÜLKE kodu tutar. Tanımın sonundaki ülke kısaltmaları
    # (RU, KZ, DE ...) önce bu alanda aranır.
    "Pazar": {
        "AR": "ARAPÇA", "BY": "BELARUS", "CS": "CIS", "CZ": "ÇEKYA",
        "DE": "ALMANYA", "EG": "MISIR", "EN": "İNGİLTERE", "ET": "ETİYOPYA",
        "EU": "AVRUPA", "EX": "EXPORT", "FI": "FİNLANDİYA", "FR": "FRANSA",
        "HU": "MACARİSTAN", "IQ": "IRAK", "IR": "İRAN", "KG": "KIRGIZİSTAN",
        "KZ": "KAZAKİSTAN", "LY": "LİBYA", "MA": "FAS", "MD": "MOLDOVA",
        "PH": "FİLİPİNLER", "PL": "FİLİSTİN", "PT": "PORTEKİZ",
        "RO": "ROMANYA", "RU": "RUSYA", "SA": "SUUDİ ARABİSTAN",
        "TN": "TUNUS", "TR": "TÜRKİYE", "UA": "UKRAYNA",
        "UK": "UNITED KINGDOM", "US": "AMERİKA", "YE": "YEMEN",
    },
    "Ek Alan": {
        # Mevcut kodlar
        "BF": "BENZOFENON FREE", "BMS": "BMS", "DVR": "DEVİR MALZEMELERİ",
        "FSC": "FSC", "IHL": "İHALE", "LF": "LILIAL FREE",
        "LQ": "LILIAL FREE QR", "PBF": "", "QR": "QR", "SF": "",
        "ST": "SVT", "TLF": "TESTER LILIAL FREE", "TS": "TESTER", "ZF": "",
        # Sonradan eklenen kodlar
        "E27": "", "EG": "", "ELR": "", "EPL": "", "FAS": "", "GL": "",
        "GLD": "", "HIJ": "", "K26": "", "K27": "", "KML": "", "ML": "",
        "MSB": "", "NM": "", "NMN": "", "PGS": "", "RU": "", "SA": "",
        "SCT": "", "SHR": "", "SLD": "", "SRP": "", "T26": "", "T27": "",
        "TR": "", "UA": "", "US": "", "UZB": "", "YA": "",
        # Son güncelleme ile gelen kodlar
        "BK": "", "E26": "",
    },
    "Ambalaj Tipi": {
        "SV": "SLEVE", "SP": "SRP", "SH": "SHIRINKLI",
        "PF": "PERFORRAJLI", "NB": "KUTUSUZ", "DP": "DOYPACK",
    },
}

# Tanımın SONUNDA kodu aranacak kategoriler. Menşei burada YOKTUR:
# o alan sayısal kod tutar, ürün adının sonundaki ülke kısaltmaları
# (RU, KZ, DE ...) doğrudan Pazar alanında aranır.
K2_TANIM_SONU_KATEGORILER = ["Pazar", "Ek Alan", "Ambalaj Tipi"]

# ---------------------------------------------------------------------
# KURAL 4 — Çokluk (xN) ile alternatif ölçü birimi sayacı karşılaştırması
# ---------------------------------------------------------------------
# Tanımın sonundaki "x24 / X4 / *6" ifadesi, MARM tablosundaki ADET
# birimine ait Sayaç (UMREZ) değeriyle aynı olmalıdır. Ton/KL gibi
# ağırlık-hacim birimlerine bakılmaz.
# Çokluk doğrulaması: malzemenin MARM'daki TÜM alternatif ölçü birimi
# satırlarındaki Sayaç değerleri toplanır. Tanımın sonundaki çokluk bu
# değerlerden HERHANGİ BİRİYLE eşleşiyorsa doğru kabul edilir.
# Aşağıdaki liste boş bırakıldığı için hiçbir birim dışarıda tutulmaz;
# istenirse birim adı eklenerek o birim değerlendirmeden çıkarılabilir.
MARM_HARIC_BIRIMLER = []
COKLUK_DESENI = r"[X*](\d{1,4})$"     # son kelimenin sonundaki çokluk

# ---------------------------------------------------------------------
# KURAL 5 — Üst kodların (RC) ayırt edici kombinasyonu benzersiz olmalı
# ---------------------------------------------------------------------
# MARA'daki marka / alt marka / varyant / ürün boyutu / S&OP kategorisi
# değerleri ile MARM'daki koli içi adet, iki farklı üst kodda birebir aynı
# olamaz; aynıysa bu kodlardan biri gereksiz demektir.
K5_PARMAK_IZI_ALANLARI = [
    "Marka1", "Raporlama Markası", "Raporlama Alt Markası",
    "Varyant", "Ürün Boyutu", "S&OP Kategorisi",
]

# "ÜB dzy.silme iştr." alanı dolu (X) olan malzeme SİLİNMİŞ sayılır ve
# hiçbir kuralda bulgu üretmez.
K5_SILME_ANAHTARLARI = ["SİLMEK", "SILMEK", "SİLME", "SILME", "DELET"]

# Bu ön eklerle başlayan malzeme kodlarının ÜST KODU OLMAMALIDIR.
# (K1 kapsamında denetlenir; üst kodu varsa bulgu üretilir.)
USTKODSUZ_ONEKLER = ["68"]

# ---------------------------------------------------------------------
# KURAL 6 — Malzeme kodunun son iki hanesi ↔ Menşei alanı
# ---------------------------------------------------------------------
# Ürün kodunun son iki hanesi, Menşei alanındaki sayısal kodla aynı olmalıdır.
# Örn. 60002810 -> son iki hane "10" (TR); Menşei alanında "21" (MY VAS)
# yazıyorsa bu bir uyuşmazlıktır. Menşei alanı BOŞ olan malzemeler de bulgu
# üretir. Geçerli menşei kodları KURAL2_KATEGORILER["Menşei"] sözlüğünden
# okunur (10=TR, 11=TR VAS, 20=MY, 21=MY VAS, 30=RU, 31=RU VAS, 40=EG,
# 41=EG VAS, 50=ID, 99=TG) ve ayarlar.xlsx ile genişletilebilir.
K6_AKTIF = True

# Kodun son iki hanesi geçerli bir menşei kodu DEĞİLSE (örn. "45"): varsayılan
# olarak bulgu üretilmez, yalnızca sayısı bilgi olarak yazdırılır. True
# yapılırsa bu malzemeler de ayrı bir bulgu olarak listelenir.
K6_GECERSIZ_SON_KOD_BILDIR = False

# K1 — "Üst kod listede bulunamadı" bulgusu:
# RC ön ekli (UST_KOD_ONEKI) üst kodlar geçerli birer referanstır. Yüklenen
# üst kod listesi kısmi bir export olabileceğinden, RC formatındaki bir üst
# kod listede yer almasa dahi HATA SAYILMAZ ve bulgu üretilmez. (Yalnızca
# RC dışı / bozuk kodlar "bulunamadı" bulgusu verir.) Bu davranışı kapatmak
# için False yapın.
UST_KOD_YOKSA_RC_MUAF = True

# K4 — MARA "Ürün hiyerarşisi" MUAFİYETİ:
# Aşağıdaki ürün hiyerarşisi kodlarına sahip malzemeler K4 (çokluk ↔ MARM)
# denetiminden MUAF tutulur; bu malzemeler için çokluk kontrolü yapılmaz.
# Karşılaştırma tam eşleşmedir (ürünün hiyerarşi değeri bu kümede ise muaf).
K4_MUAF_HIYERARSILER = {
    "110016", "110016001",
    "110016001001", "110016001001001", "110016001001002", "110016001001003",
    "110016001001004", "110016001001005", "110016001001006", "110016001001007",
    "110016001002", "110016001002001", "110016001002002", "110016001002003",
    "110016001002004", "110016001002005", "110016001002006", "110016001002007",
    "110016001003", "110016001003001", "110016001003002", "110016001003003",
    "110016001003004", "110016001003005", "110016001003006", "110016001003007",
    "110016001004", "110016001004001", "110016001004002", "110016001004003",
    "110016001004004", "110016001004005", "110016001004006", "110016001004007",
    "110016001005", "110016001005001", "110016001005002", "110016001005003",
    "110016001005004", "110016001005005", "110016001005006", "110016001005007",
    "110016001006", "110016001006001", "110016001006002", "110016001006003",
    "110016001006004", "110016001006005", "110016001006006", "110016001006007",
    "110016001007", "110016001007001",
    "110016001008", "110016001008001", "110016001008002", "110016001008003",
    "110016001008004", "110016001008005", "110016001008006", "110016001008007",
    "110016001009", "110016001009001", "110016001009002", "110016001009003",
    "110016001009004", "110016001009005", "110016001009006", "110016001009007",
    "110016001010", "110016001010001", "110016001010002", "110016001010003",
    "110016001010007",
    "110016001011", "110016001011001", "110016001011002", "110016001011003",
    "110016001011004", "110016001011005", "110016001011006", "110016001011007",
    "110016001012", "110016001012001", "110016001012002", "110016001012003",
    "110016001012004", "110016001012005", "110016001012006", "110016001012007",
    "110016002", "110016002001", "110016002001001", "110016002001002",
    "110016002001003", "110016002001004", "110016002001005", "110016002001006",
    "110016002001007",
    "110016999", "110017", "110019", "110019001", "110020", "130007",
}

# Ürün hiyerarşisi (MARA · PRDHA) sütununun olası adları.
HIYERARSI_SUTUN_ADAYLARI = [
    "Ürün hiyerarşisi", "Ürün Hiyerarşisi", "Urun hiyerarsisi",
    "Product hierarchy", "Product Hierarchy", "PRDHA", "Prod.hier.",
]


def hiyerarsi_normalize(v) -> str:
    """Hiyerarşi kodunu karşılaştırmaya hazırlar: boşlukları at, Excel'in
    sayı olarak okuyup eklediği '.0' son ekini temizle."""
    s = str("" if v is None else v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def hiyerarsi_sutunu_bul(df) -> str:
    """Ürün hiyerarşisi sütununu bulur (tam ad veya 'HİYERARŞİ'/'PRDHA' içeren)."""
    for aday in HIYERARSI_SUTUN_ADAYLARI:
        if aday in df.columns:
            return aday
    for c in df.columns:
        u = tr_upper(str(c))
        if "HİYERARŞİ" in u or "HIYERARSI" in u or "PRDHA" in u \
                or "PRODUCT HIERARCHY" in u:
            return c
    return ""

# Malzeme kısa metninin SONUNDA bulunması yasak kodlar (K2 kapsamında).
# Bu kodlar Ek Alan sözlüğünde tanımlı olsa bile tanım sonunda geçemez;
# normal K2 tanım-sonu/ters kontrol taramasından muaf tutulur (çift bulgu olmasın).
TANIM_SONU_YASAK = ["T26"]

# Bu kodlar aynı zamanda ölçü birimidir. Tanımda bir SAYIDAN hemen sonra
# geliyorlarsa (örn. "500 ML", "1 GL", "250ML") ölçü birimi kabul edilir ve
# kod olarak değerlendirilmez — aksi hâlde her ürün yanlış bulgu üretirdi.
OLCU_BIRIMI_KODLARI = ["ML", "GL", "L", "LT", "CL", "KG", "G", "GR", "MG", "CC"]

# Bazı kodlar tanımın sonuna BİTİŞİK yazılır (örn. "1Lx9LF" -> LF).
# Aşağıdaki kodlar, tanımın son kelimesinin sonuna yapışık olsa da tanınır.
# Uzun kodlar önce yazılmalı (TLF, LF'den önce denenmeli).
K2_BITISIK_KODLAR = ["TLF", "LF", "LQ", "QR"]

# Ters yön kontrolü: alan doluysa ama tanımın sonunda kodu yoksa da bildir
# ve düzeltilmiş tanım önerisi üret (toplu güncelleme listesi için).
# Yalnızca TERS_KONTROL_ALANLARI listesindeki alanlara uygulanır; çünkü
# örn. Ambalaj Tipi veya Menşei dolu olan her ürünün adında bu kodun
# yazması gerekmiyor — ama Ek Alan'da LF seçiliyse adda da olmalı.
TERS_KONTROL = True
TERS_KONTROL_ALANLARI = ["Ek Alan"]

# Kural 3: yasaklı kelimeler (tek başına kelime olarak aranır)
YASAKLI_KELIMELER = ["NEW", "YENİ", "YENI", "6", "7", "2026", "2027"]
# Bu sayılar tanımın ORTASINDA geçerse hata değildir (örn. "ARKO CREAM 6 150X4X18 ML"
# ürün varyantıdır); yalnızca tanımın SONUNDA yer alırsa yakalanır.
YASAKLI_SADECE_SONDA = ["6", "7"]

# İlk satırı başlık:değer olarak ekrana bas (sütun eşleşmesini doğrulamak için)
VERI_ONIZLEME = True

# Tanım sütunu için olası isimler (export'ta hangisi varsa onu bulur):
TANIM_SUTUN_ADAYLARI = [
    "Malzeme tanımı", "Malzeme Tanımı", "Malzeme kısa metni",
    "Malzeme kısa metinleri", "malzeme kısa metinleri",
    "Kısa metin", "Tanım", "Malzeme Kısa Metni", "Material Description",
]

# =====================================================================
# YARDIMCI FONKSİYONLAR
# =====================================================================

def tr_upper(s: str) -> str:
    """Türkçe karakterlere duyarlı büyük harfe çevirme (i -> İ, ı -> I)."""
    if not isinstance(s, str):
        return ""
    return s.replace("i", "İ").replace("ı", "I").upper()


def normalize(s) -> str:
    """Karşılaştırma için: boşlukları sadeleştir, büyük harfe çevir."""
    if pd.isna(s):
        return ""
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s)
    return tr_upper(s)


KANONIK_KOLONLAR = [
    "Malzeme", "Yaratma tarihi", "Oluşturma saati", "Malzeme türü",
    "Mal grubu", "Temel ölçü birimi", "Temel malzeme",
    "Raporlama Alt Markası", "Menşei", "Referans Ürün Boyutu",
    "S&OP Kategorisi", "Pazar", "Raporlama Markası", "Marka2", "Ek Alan",
    "Marka3", "Ürün Boyutu", "Ambalaj Tipi", "SKU Grup", "Marka1",
    "Varyant", "Kare Barkod",
    "Son değişiklik tarihi", "Son değişiklik saati",
    "Malzeme tanımı", "Malzeme kısa metni", "Malzeme kısa metinleri",
    "Kısa metin", "Tanım",
]


def kolon_adlarini_onar(df: pd.DataFrame) -> pd.DataFrame:
    """SAP export'unda Türkçe karakterler '?' olmuşsa (örn. 'Men?ei'),
    kanonik ada geri çevirir. '?' herhangi bir harfle eşleşir."""
    yeniden = {}
    for kolon in df.columns:
        if "?" not in kolon:
            continue
        desen = re.escape(kolon).replace(r"\?", ".")
        adaylar = [k for k in KANONIK_KOLONLAR if re.fullmatch(desen, k)]
        if len(adaylar) == 1:
            yeniden[kolon] = adaylar[0]
    if yeniden:
        for eski, yeni in yeniden.items():
            print(f"Bilgi: bozuk sütun adı onarıldı: '{eski}' -> '{yeni}'")
        df = df.rename(columns=yeniden)
    return df


def normalize_seri(s: pd.Series) -> pd.Series:
    """normalize() fonksiyonunun tüm sütuna tek seferde uygulanan hızlı hali."""
    return (s.fillna("").astype(str).str.strip()
            .str.replace(r"\s+", " ", regex=True)
            .str.replace("i", "İ", regex=False)
            .str.replace("ı", "I", regex=False)
            .str.upper())


def dosya_oku(yol: str) -> pd.DataFrame:
    if yol.lower().endswith((".xlsx", ".xls")):
        df = pd.read_excel(yol, dtype=str)
    else:
        # SAP CSV export'ları farklı kodlamalarla gelebilir; sırayla dene.
        df = None
        son_hata = None
        for enc in ("utf-8-sig", "utf-8", "cp1254", "latin-1", "utf-16"):
            try:
                df = pd.read_csv(yol, dtype=str, sep=None,
                                 engine="python", encoding=enc)
                print(f"Bilgi: '{yol}' dosyası '{enc}' kodlamasıyla okundu.")
                break
            except (UnicodeDecodeError, UnicodeError) as e:
                son_hata = e
                continue
        if df is None:
            raise RuntimeError(
                f"'{yol}' hiçbir kodlamayla okunamadı. Dosyayı Excel'de açıp "
                f"XLSX olarak kaydetmeyi deneyin. Son hata: {son_hata}")
        # Tek sütun geldiyse ayraç yanlış algılanmış olabilir; noktalı virgül dene
        if df.shape[1] == 1:
            for enc in ("utf-8-sig", "cp1254", "latin-1"):
                try:
                    tmp = pd.read_csv(yol, dtype=str, sep=";", encoding=enc)
                    if tmp.shape[1] > 1:
                        df = tmp
                        break
                except (UnicodeDecodeError, UnicodeError):
                    continue
    df.columns = [str(c).strip() for c in df.columns]
    df = kolon_adlarini_onar(df)
    df = dil_algila_ve_uyarla(df)
    return df


def tanim_sutunu_bul(df: pd.DataFrame, dosya_adi: str) -> str | None:
    # Önce tam ad eşleşmesi
    for aday in TANIM_SUTUN_ADAYLARI:
        if aday in df.columns:
            return aday
    # Sonra esnek arama: içinde tanım/metin/text/description geçen sütun
    for c in df.columns:
        c_norm = tr_upper(str(c))
        if "STANDART" in c_norm:
            continue
        if any(a in c_norm for a in ("TANIM", "METİN", "METIN",
                                      "DESCRIPTION", "TEXT", "MAKTX")):
            print(f"Bilgi: '{dosya_adi}' için tanım sütunu olarak '{c}' kullanılıyor.")
            return c
    print(f"UYARI: '{dosya_adi}' içinde malzeme tanımı sütunu bulunamadı.")
    print("       Dosyadaki mevcut sütunlar şunlar:")
    for c in df.columns:
        print(f"         - {c}")
    print("       SAP export'una 'Malzeme tanımı' (MAKTX) sütununu ekleyin")
    print("       veya TANIM_SUTUN_ADAYLARI listesine doğru adı yazın.")
    print("       Kural 2, 3 ve 4 bu dosya için atlanacak.\n")
    return None


def tur_filtrele(df: pd.DataFrame) -> pd.DataFrame:
    if "Malzeme türü" not in df.columns:
        return df
    tur = df["Malzeme türü"].fillna("").astype(str).str.strip()
    if TAM_ESLESME_TURU:
        maske = tur.isin(MALZEME_TURU_FILTRESI)
    else:
        maske = tur.apply(lambda t: any(f in t for f in MALZEME_TURU_FILTRESI))
    return df[maske].copy()


# =====================================================================
# KURAL 1 — Ürün kodu <-> Üst kod hatalı eşleşme kontrolü
# =====================================================================

def _tanim_sonu_yasak_bulgular(df: pd.DataFrame, tanim_kolonu: str,
                               kaynak: str) -> pd.DataFrame:
    """Malzeme kısa metni yasak bir kodla bitiyorsa K2 bulgusu üretir.
    (Tanım ↔ ek veri alanı kuralının bir parçasıdır.)"""
    if not tanim_kolonu or df.empty or not TANIM_SONU_YASAK:
        return pd.DataFrame()
    t = normalize_seri(df[tanim_kolonu])
    parcalar = []
    for kod in TANIM_SONU_YASAK:
        k = tr_upper(kod)
        # Hem ayrık ("… 100 ML T26") hem bitişik ("…100MLT26") yazım
        desen = rf"(?:^|[\s\-_/.]|\d){re.escape(k)}$"
        m = t.str.contains(desen, regex=True)
        if not m.any():
            continue
        d = df[m]
        ham = d[tanim_kolonu].astype(str)
        # Yasak kod tanımdan çıkarılmış hâli (öneri)
        onerilen = ham.str.replace(rf"[\s\-_/.]*{re.escape(k)}\s*$", "",
                                   regex=True, case=False).str.strip()
        parcalar.append(pd.DataFrame({
            "Kaynak": kaynak,
            "Malzeme": d["Malzeme"],
            "Tanım": ham,
            "Anahtar Kelime": k,
            "Sorun": (f"Malzeme kısa metni '{k}' ile bitiyor — bu kod "
                      f"tanımın sonunda kullanılamaz"),
            "Kontrol Edilen Alanlar": "Tanım sonu",
            "Önerilen Düzeltme": (f"Tanımın sonundaki '{k}' ifadesini "
                                  f"kaldırın"),
            "Önerilen Tanım": onerilen,
        }))
    if not parcalar:
        return pd.DataFrame()
    return pd.concat(parcalar, ignore_index=True)


def kural1(urun: pd.DataFrame, ust: pd.DataFrame,
           haric_tanim_kolonlari=()) -> pd.DataFrame:
    """Ürünler üst kodlarına tek merge ile bağlanır; her kıyas alanı
    vektörel olarak karşılaştırılır (satır döngüsü yok)."""
    if "Temel malzeme" not in urun.columns or "Malzeme" not in ust.columns:
        print("UYARI: Kural 1 için 'Temel malzeme' / 'Malzeme' sütunları gerekli.")
        return pd.DataFrame()

    urun_f = tur_filtrele(urun).copy()
    if urun_f.empty:
        return pd.DataFrame()
    urun_f["_temel"] = normalize_seri(urun_f["Temel malzeme"])

    ust2 = ust.copy()
    ust2["_key"] = normalize_seri(ust2["Malzeme"])
    ust2 = ust2.drop_duplicates("_key")

    ortak = [c for c in urun.columns if c in ust.columns]
    kiyas_alanlari = []
    for c in KIYAS_ALANLARI:
        if c not in ortak or c in TANIM_SUTUN_ADAYLARI \
                or c in haric_tanim_kolonlari:
            continue
        if normalize_seri(ust2[c]).eq("").all():
            print(f"  Bilgi: '{c}' üst kodlarda hiç dolu olmadığı için "
                  f"K1 karşılaştırması dışı bırakıldı.")
            continue
        kiyas_alanlari.append(c)

    parcalar = []
    kolonlar = ["Malzeme", "Üst Kod", "Alan", "Ürün Değeri",
                "Üst Kod Değeri", "Sorun"]

    # 68 ile başlayan malzemelerin üst kodu OLMAMALIDIR
    onek_maske = urun_f["Malzeme"].astype(str).str.strip().str.startswith(
        tuple(USTKODSUZ_ONEKLER))
    ustkodsuz = onek_maske & (urun_f["_temel"] != "")
    if ustkodsuz.any():
        d = urun_f[ustkodsuz]
        onek_metni = " / ".join(USTKODSUZ_ONEKLER)
        parcalar.append(pd.DataFrame({
            "Malzeme": d["Malzeme"], "Üst Kod": d["Temel malzeme"],
            "Alan": "Temel malzeme", "Ürün Değeri": d["Temel malzeme"],
            "Üst Kod Değeri": "(boş olmalı)",
            "Sorun": (f"{onek_metni} ile başlayan malzemenin üst kodu "
                      f"olmamalı")}))
    # 68'li kodlar normal karşılaştırmaya girmez (üst kodu beklenmez)
    urun_f = urun_f[~onek_maske].copy()

    bos = urun_f["_temel"] == ""
    if bos.any():
        d = urun_f[bos]
        parcalar.append(pd.DataFrame({
            "Malzeme": d["Malzeme"], "Üst Kod": "", "Alan": "Temel malzeme",
            "Ürün Değeri": "", "Üst Kod Değeri": "",
            "Sorun": "Üst kod (Temel malzeme) boş"}))

    m = urun_f[~bos].merge(ust2, left_on="_temel", right_on="_key",
                           how="left", suffixes=("", "__ust"))
    bulunamadi = m["_key"].isna()
    if bulunamadi.any():
        d = m[bulunamadi]
        if UST_KOD_YOKSA_RC_MUAF:
            # RC formatındaki üst kodlar geçerli referans sayılır; yüklenen
            # listede yer almasalar bile "bulunamadı" bulgusu üretilmez.
            rc_maske = d["_temel"].astype(str).str.startswith(
                tr_upper(UST_KOD_ONEKI))
            d = d[~rc_maske]
        if not d.empty:
            parcalar.append(pd.DataFrame({
                "Malzeme": d["Malzeme"], "Üst Kod": d["Temel malzeme"],
                "Alan": "Temel malzeme", "Ürün Değeri": d["Temel malzeme"],
                "Üst Kod Değeri": "", "Sorun": "Üst kod listede bulunamadı"}))

    # RC muaf kodlar es'e (eşleşenler) de girmez: üst kayıt yok olduğundan
    # alan karşılaştırmaları NaN üretmesin diye bulunamadı olarak dışta kalır.
    es = m[~bulunamadi]
    for alan in kiyas_alanlari:
        u = normalize_seri(es[alan])
        p = normalize_seri(es[alan + "__ust"])
        if alan in BAYRAK_ALANLARI:
            # Yalnızca işaretli/işaretsiz durumu karşılaştırılır
            u = u.where(u == "", "X")
            p = p.where(p == "", "X")
        fark = u != p
        if alan in BOS_HATA_SAYILMAZ:
            fark = fark & (u != "") & (p != "")
        if fark.any():
            d = es[fark]
            if alan in BAYRAK_ALANLARI:
                yaz = lambda s: np.where(normalize_seri(s) == "",
                                         "(işaretsiz)", "X (işaretli)")
                urun_deger, ust_deger = yaz(d[alan]), yaz(d[alan + "__ust"])
                sorun = "Barkod işareti üst kod ile uyuşmuyor"
            else:
                urun_deger, ust_deger = d[alan], d[alan + "__ust"]
                sorun = "Alan değeri üst kod ile uyuşmuyor"
            parcalar.append(pd.DataFrame({
                "Malzeme": d["Malzeme"], "Üst Kod": d["Temel malzeme"],
                "Alan": alan, "Ürün Değeri": urun_deger,
                "Üst Kod Değeri": ust_deger,
                "Sorun": sorun}))

    # --- Üst birim düzeyi silme işareti (ÜB dzy.silme iştr.) ---
    # Üst kod silinmek üzere işaretlenmişse ona bağlı TÜM ürün kodları da
    # işaretli olmalıdır. Tersi (ürün işaretli, üst kod işaretsiz) hata DEĞİL.
    silme_kolonu = ""
    for c in ortak:
        if any(a in tr_upper(str(c)) for a in K5_SILME_ANAHTARLARI):
            silme_kolonu = c
            break
    if silme_kolonu and not es.empty:
        u_isaret = normalize_seri(es[silme_kolonu])
        p_isaret = normalize_seri(es[silme_kolonu + "__ust"])
        # Üst kod silinmek üzere işaretliyse ona bağlı TÜM ürünler silinmelidir.
        # Ürün henüz işaretsizse eksik işaret, işaretliyse silme işlemi
        # takibi olarak her iki durumda da bulgu üretilir.
        kaskad = p_isaret != ""
        if kaskad.any():
            d = es[kaskad]
            d_urun = u_isaret[kaskad]
            parcalar.append(pd.DataFrame({
                "Malzeme": d["Malzeme"], "Üst Kod": d["Temel malzeme"],
                "Alan": silme_kolonu,
                "Ürün Değeri": np.where(d_urun == "", "(işaretsiz)",
                                        "X (işaretli)"),
                "Üst Kod Değeri": "X (işaretli)",
                "Sorun": np.where(
                    d_urun == "",
                    ("Üst kod silinmek üzere işaretli — bu ürün de silinmeli, "
                     "ancak silme işareti konulmamış"),
                    ("Üst kod silinmek üzere işaretli — bu ürün de silinecek "
                     "(işareti konulmuş, silme işlemi tamamlanmalı)"))}))

    if not parcalar:
        return pd.DataFrame(columns=kolonlar)
    return (pd.concat(parcalar, ignore_index=True)[kolonlar]
            .sort_values(["Malzeme", "Alan"]).reset_index(drop=True))


# =====================================================================
# KURAL 2 — Tanım sonu <-> ek veri alanı tutarlılığı (örn. LF)
# =====================================================================

def _bitisik_kod_bul(parca: str, tum_kodlar) -> str:
    """'200MLx24RU' gibi bitişik yazılmış son ek kodunu ayıklar.
    Kod, rakamdan hemen sonra geliyorsa (24RU, 9LF) güvenle tanınır;
    ayrıca K2_BITISIK_KODLAR listesindekiler her durumda tanınır."""
    olcu = {tr_upper(x) for x in OLCU_BIRIMI_KODLARI}
    for k in sorted(tum_kodlar, key=len, reverse=True):
        if k in olcu:
            continue          # "500ML" ölçü birimidir, Ek Alan kodu değil
        if len(parca) > len(k) and parca.endswith(k):
            onceki = parca[-len(k) - 1]
            if onceki.isdigit() or k in [tr_upper(x) for x in K2_BITISIK_KODLAR]:
                return k
    return ""


def _kural2_kod_haritalari():
    """Kod -> hangi alanlarda geçerli; ve tek/çift kelimelik kod kümeleri."""
    kod_alanlari = {}
    _yasak = {tr_upper(x) for x in TANIM_SONU_YASAK}
    for alan in K2_TANIM_SONU_KATEGORILER:
        for kod in KURAL2_KATEGORILER.get(alan, {}):
            k = tr_upper(kod)
            if k in _yasak:
                continue          # tanım sonunda yasak; K1 bunu bildirir
            kod_alanlari.setdefault(k, []).append(alan)
    tek = {k for k in kod_alanlari if " " not in k}
    cift = {k for k in kod_alanlari if " " in k}
    return kod_alanlari, tek, cift


def _tanim_sonu_kodlari(tanim_norm: str, tek: set, cift: set) -> list:
    """Tanımın SONUNDAN geriye doğru art arda gelen kodları toplar.
    Örn. 'ARKO SOAP TR LF' -> ['LF', 'TR']"""
    parcalar = re.split(r"[\s\-_/.]+", tanim_norm)
    parcalar = [p for p in parcalar if p]
    bulunan = []
    tum_kod_kumesi = tek | cift
    i = len(parcalar)
    while i > 1:  # en az bir parça asıl isim olarak kalsın
        if i >= 2 and " ".join(parcalar[i-2:i]) in cift:
            bulunan.append(" ".join(parcalar[i-2:i]))
            i -= 2
        elif parcalar[i-1] in tek:
            # "…500 ML" -> ML burada ölçü birimidir, kod değil
            if (parcalar[i-1] in {tr_upper(x) for x in OLCU_BIRIMI_KODLARI}
                    and i >= 2 and parcalar[i-2]
                    and parcalar[i-2][-1].isdigit()):
                break
            bulunan.append(parcalar[i-1])
            i -= 1
        else:
            # Bitişik yazılmış kod (…200MLx24RU) — bulunursa alınır ve durulur
            bitisik = _bitisik_kod_bul(parcalar[i-1], tum_kod_kumesi)
            if bitisik:
                bulunan.append(bitisik)
            break
    return bulunan


def _harf_katla(s: str) -> str:
    """İ/I farkını yok sayarak karşılaştırma için."""
    return s.replace("İ", "I")


def _alan_dolu_mu(deger_norm: str, kod: str, alan: str) -> bool:
    """Alan değeri bu kodu karşılıyor mu? (kod veya uzun metin olarak)"""
    if not deger_norm:
        return False
    d = _harf_katla(deger_norm)
    k = _harf_katla(kod)
    if d == k or k in d.split():
        return True
    metin = _harf_katla(tr_upper(KURAL2_KATEGORILER[alan].get(kod, "") or ""))
    return bool(metin) and d == metin


def kural2(df: pd.DataFrame, tanim_kolonu: str, kaynak: str) -> pd.DataFrame:
    if not tanim_kolonu:
        return pd.DataFrame()
    yasak_bulgular = _tanim_sonu_yasak_bulgular(df, tanim_kolonu, kaynak)
    kod_alanlari, tek, cift = _kural2_kod_haritalari()
    # Kapsam filtresi ana akışta uygulandığı için burada tür süzmesi yapılmaz;
    # aksi hâlde üst kod (tür 15) kayıtları K2 denetiminin dışında kalırdı.
    df_f = df
    if df_f.empty:
        return pd.DataFrame()

    # Ön filtre: tanımı bir kodla bitenler + (ters kontrol açıksa) ilgili
    # alanı dolu olanlar dışındaki satırlar hiç dolaşılmaz.
    t_norm = normalize_seri(df_f[tanim_kolonu])
    tum_kodlar = sorted(tek | cift, key=len, reverse=True)
    desen = (r"(?:^|[\s\-_/.])(?:"
             + "|".join(re.escape(k) for k in tum_kodlar) + r")$")
    aday = t_norm.str.contains(desen, regex=True)
    # Bitişik yazılmış son ekler: rakamdan sonra gelen her kod + LF/TLF/LQ/QR
    desen2 = (r"\d(?:" + "|".join(re.escape(k) for k in tum_kodlar) + r")$")
    aday |= t_norm.str.contains(desen2, regex=True)
    if K2_BITISIK_KODLAR:
        desen3 = (r"(?:" + "|".join(re.escape(tr_upper(k))
                                    for k in K2_BITISIK_KODLAR) + r")$")
        aday |= t_norm.str.contains(desen3, regex=True)
    if TERS_KONTROL:
        for alan in TERS_KONTROL_ALANLARI:
            if alan in df_f.columns:
                aday |= normalize_seri(df_f[alan]) != ""
    df_f = df_f[aday]

    bulgular = []
    for _, satir in df_f.iterrows():
        tanim = normalize(satir.get(tanim_kolonu))
        if not tanim:
            continue
        alan_degerleri = {a: normalize(satir.get(a))
                          for a in KURAL2_KATEGORILER if a in df_f.columns}

        # Yön 1: tanımın sonunda kod var -> ilgili alan dolu olmalı
        for kod in _tanim_sonu_kodlari(tanim, tek, cift):
            ilgili_alanlar = [a for a in kod_alanlari.get(kod, [])
                              if a in alan_degerleri]
            if not ilgili_alanlar:
                continue
            karsilandi = any(_alan_dolu_mu(alan_degerleri[a], kod, a)
                             for a in ilgili_alanlar)
            if not karsilandi:
                bulgular.append({
                    "Kaynak": kaynak,
                    "Malzeme": satir.get("Malzeme"),
                    "Tanım": satir.get(tanim_kolonu),
                    "Anahtar Kelime": kod,
                    "Sorun": ("Tanımın sonunda '" + kod + "' var ama "
                              + " / ".join(ilgili_alanlar)
                              + " alanı doldurulmamış veya uyuşmuyor"),
                    "Kontrol Edilen Alanlar": "; ".join(
                        f"{a}={alan_degerleri[a] or '-'}"
                        for a in ilgili_alanlar),
                    "Önerilen Düzeltme": (" / ".join(ilgili_alanlar)
                                          + " alanına '" + kod + "' girilmeli"),
                    "Önerilen Tanım": "",
                })

        # Yön 2 (opsiyonel): alan dolu ama tanımın sonunda kodu yok
        if TERS_KONTROL:
            son_kodlar = set(_tanim_sonu_kodlari(tanim, tek, cift))
            for alan, deger in alan_degerleri.items():
                if alan not in TERS_KONTROL_ALANLARI or not deger:
                    continue
                for kod in KURAL2_KATEGORILER[alan]:
                    k = tr_upper(kod)
                    if k in {tr_upper(x) for x in TANIM_SONU_YASAK}:
                        continue      # tanım sonunda yasak; K1 bildirir
                    if _alan_dolu_mu(deger, k, alan) and k not in son_kodlar:
                        bulgular.append({
                            "Kaynak": kaynak,
                            "Malzeme": satir.get("Malzeme"),
                            "Tanım": satir.get(tanim_kolonu),
                            "Anahtar Kelime": k,
                            "Sorun": (alan + " alanında '" + k
                                      + "' seçili ama tanımın sonunda yok"),
                            "Kontrol Edilen Alanlar": f"{alan}={deger}",
                            "Önerilen Düzeltme": "Tanımın sonuna '" + k + "' eklenmeli",
                            "Önerilen Tanım": (str(satir.get(tanim_kolonu)).strip()
                                               + " " + k),
                        })
                        break
    sonuc = pd.DataFrame(bulgular)
    if not yasak_bulgular.empty:
        sonuc = (pd.concat([yasak_bulgular, sonuc], ignore_index=True)
                 if not sonuc.empty else yasak_bulgular)
    return sonuc


# =====================================================================
# KURAL 3 — Yasaklı kelimeler
# =====================================================================

def kural3(df: pd.DataFrame, tanim_kolonu: str, kaynak: str) -> pd.DataFrame:
    """Yasaklı kelimeler tüm sütunda vektörel regex ile aranır;
    yalnızca eşleşen az sayıda satır ayrıntılandırılır."""
    if not tanim_kolonu:
        return pd.DataFrame()
    t = normalize_seri(df[tanim_kolonu])
    maskeler = {}
    for kelime in YASAKLI_KELIMELER:
        k = tr_upper(kelime)
        if kelime in YASAKLI_SADECE_SONDA:
            # 6/7 yalnızca tanımın SONUNDA bağımsız kelime ise yakalanır
            desen = rf"(?<!\S){re.escape(k)}$"
        elif k.isdigit():
            # Yıllar (2026, 2027) bağımsız kelime olarak her yerde yakalanır
            desen = rf"(?<!\S){re.escape(k)}(?!\S)"
        else:
            desen = rf"(?<![\wÇĞİÖŞÜ]){re.escape(k)}(?![\wÇĞİÖŞÜ])"
        maskeler[kelime] = t.str.contains(desen, regex=True)
    maske_df = pd.DataFrame(maskeler, index=df.index)
    herhangi = maske_df.any(axis=1) & (t != "")
    if not herhangi.any():
        return pd.DataFrame()
    alt = maske_df[herhangi]
    bulunanlar = alt.apply(
        lambda r: ", ".join(k for k, v in r.items() if v), axis=1)
    d = df[herhangi]
    return pd.DataFrame({
        "Kaynak": kaynak, "Malzeme": d["Malzeme"],
        "Tanım": d[tanim_kolonu], "Yasaklı Kelimeler": bulunanlar,
        "Sorun": "Tanımda yasaklı kelime var"}).reset_index(drop=True)


# =====================================================================
# KURAL 4 — Aynı tanım, farklı malzeme kodu
# =====================================================================

def alt_tarifler(df: pd.DataFrame, tanim_kolonu: str, kaynak: str) -> pd.DataFrame:
    if not tanim_kolonu:
        return pd.DataFrame()
    tmp = df.copy()
    tmp["_norm"] = tmp[tanim_kolonu].map(normalize)
    tmp = tmp[tmp["_norm"] != ""]
    grup = tmp.groupby("_norm")["Malzeme"].nunique()
    tekrarli = grup[grup > 1].index
    sonuc = tmp[tmp["_norm"].isin(tekrarli)].copy()
    if sonuc.empty:
        return pd.DataFrame()
    sonuc = sonuc.sort_values(["_norm", "Malzeme"])
    # Aynı tanımı paylaşan tüm kodların listesi (satır başına eklenir)
    kod_listesi = (sonuc.groupby("_norm")["Malzeme"]
                   .apply(lambda s: ", ".join(sorted(s.astype(str).unique()))))
    out = sonuc[["Malzeme", tanim_kolonu]].rename(columns={tanim_kolonu: "Tanım"})
    out.insert(0, "Kaynak", kaynak)
    out["Sorun"] = "Üst kodun alt tarifleri — aynı tanım birden fazla kodda (bilgi)"
    out["Tekrar Sayısı"] = sonuc["_norm"].map(grup)
    out["Paylaşan Kodlar"] = sonuc["_norm"].map(kod_listesi)
    return out


# =====================================================================
# DIŞ AYAR DOSYASI (ayarlar.xlsx) — Python bilmeden liste güncelleme
# =====================================================================

def ayarlar_yukle():
    """ayarlar.xlsx varsa kod listelerini oradan alır.
    Beklenen sütunlar: Grup | Değer | Metin (opsiyonel)
    Gruplar: YASAKLI, MENŞEİ, PAZAR, EK ALAN, AMBALAJ TİPİ, ÜRÜN TÜRÜ, ÜST TÜRÜ"""
    import os
    global YASAKLI_KELIMELER, KURAL2_KATEGORILER
    global GECERLI_URUN_TURLERI, GECERLI_UST_TURLERI, TANIM_SONU_YASAK
    if not os.path.exists(AYARLAR_DOSYASI):
        return
    try:
        a = pd.read_excel(AYARLAR_DOSYASI, dtype=str).fillna("")
    except Exception as e:
        print(f"UYARI: {AYARLAR_DOSYASI} okunamadı ({e}); gömülü ayarlar kullanılacak.")
        return
    if "Grup" not in a.columns or "Değer" not in a.columns:
        print(f"UYARI: {AYARLAR_DOSYASI} 'Grup' ve 'Değer' sütunlarını içermeli.")
        return
    a["_g"] = a["Grup"].map(tr_upper).str.strip()
    a["_d"] = a["Değer"].astype(str).str.strip()
    a = a[a["_d"] != ""]

    def grup(ad):
        alt = a[a["_g"] == tr_upper(ad)]
        return alt

    # --- BİRLEŞTİRME MANTIĞI ---
    # Programda tanımlı kurallar her zaman geçerli kalır; Excel'de fazladan
    # görülen her satır YENİ KURAL olarak eklenir. Hiçbir kural silinmez.
    yeni_kayit = []

    g = grup("YASAKLI")
    if not g.empty:
        mevcut = {tr_upper(k) for k in YASAKLI_KELIMELER}
        for kelime in g["_d"]:
            if tr_upper(kelime) not in mevcut:
                YASAKLI_KELIMELER.append(kelime)
                mevcut.add(tr_upper(kelime))
                yeni_kayit.append(("YASAKLI", kelime, ""))

    for alan in ("Menşei", "Pazar", "Ek Alan", "Ambalaj Tipi"):
        g = grup(alan)
        if g.empty:
            continue
        hedef = KURAL2_KATEGORILER.setdefault(alan, {})
        mevcut = {tr_upper(k) for k in hedef}
        metinler = (list(g["Metin"]) if "Metin" in g.columns
                    else [""] * len(g))
        for kod, metin in zip(g["_d"], metinler):
            k = tr_upper(kod)
            if k not in mevcut:
                hedef[k] = tr_upper(str(metin or ""))
                mevcut.add(k)
                yeni_kayit.append((alan, k, str(metin or "")))
            elif metin and not hedef.get(k):
                hedef[k] = tr_upper(str(metin))   # açıklama tamamlandı

    for grup_adi, kume in (("ÜRÜN TÜRÜ", GECERLI_URUN_TURLERI),
                           ("ÜST TÜRÜ", GECERLI_UST_TURLERI)):
        g = grup(grup_adi)
        for d in g["_d"] if not g.empty else []:
            tur = d.lstrip("0") or "0"
            if tur not in kume:
                kume.add(tur)
                yeni_kayit.append((grup_adi, tur, ""))

    g = grup("TANIM SONU YASAK")
    if not g.empty:
        mevcut = {tr_upper(k) for k in TANIM_SONU_YASAK}
        for kod in g["_d"]:
            if tr_upper(kod) not in mevcut:
                TANIM_SONU_YASAK.append(tr_upper(kod))
                mevcut.add(tr_upper(kod))
                yeni_kayit.append(("TANIM SONU YASAK", tr_upper(kod), ""))

    print(f"Bilgi: {AYARLAR_DOSYASI} yüklendi — gömülü kurallar korundu.")
    if yeni_kayit:
        print(f"  {len(yeni_kayit)} YENİ KURAL eklendi:")
        for gr, deger, metin in yeni_kayit[:25]:
            print(f"    + {gr}: {deger}" + (f"  ({metin})" if metin else ""))
        if len(yeni_kayit) > 25:
            print(f"    … ve {len(yeni_kayit) - 25} kayıt daha")
    else:
        print("  Yeni kural yok — dosyadaki tüm satırlar zaten tanımlı.")


# =====================================================================
# ONAYLI İSTİSNALAR (istisnalar.xlsx) — bilinen/kabul edilen bulgular
# =====================================================================

def istisna_yukle() -> pd.DataFrame:
    """istisnalar.xlsx varsa okur. Beklenen sütunlar:
    Malzeme | Kural (K1..K5 veya HEPSİ) | Onaylayan | Not"""
    import os
    if not os.path.exists(ISTISNA_DOSYASI):
        return pd.DataFrame()
    try:
        i = pd.read_excel(ISTISNA_DOSYASI, dtype=str).fillna("")
    except Exception as e:
        print(f"UYARI: {ISTISNA_DOSYASI} okunamadı ({e}).")
        return pd.DataFrame()
    if "Malzeme" not in i.columns:
        print(f"UYARI: {ISTISNA_DOSYASI} 'Malzeme' sütunu içermeli.")
        return pd.DataFrame()
    i["_m"] = i["Malzeme"].astype(str).str.strip()
    i["_k"] = (i["Kural"].map(tr_upper).str.strip().str[:5]
               if "Kural" in i.columns else "HEPSİ")
    i.loc[i["_k"] == "", "_k"] = "HEPSİ"
    print(f"Bilgi: {ISTISNA_DOSYASI} yüklendi — {len(i)} onaylı istisna uygulanacak.")
    return i


def istisna_uygula(df: pd.DataFrame, kod: str, ist: pd.DataFrame):
    """Bulgu tablosundan onaylı istisnaları düşer; düşülenleri ayrıca döndürür."""
    if df.empty or ist.empty:
        return df, pd.DataFrame()
    hedef = set(ist.loc[ist["_k"].isin([kod, "HEPSİ"]), "_m"])
    if not hedef:
        return df, pd.DataFrame()
    m = df["Malzeme"].astype(str).str.strip().isin(hedef)
    atlanan = df[m].copy()
    if not atlanan.empty:
        atlanan.insert(0, "İstisna Kuralı", kod)
    return df[~m].reset_index(drop=True), atlanan


# =====================================================================
# ÇALIŞTIRMA GEÇMİŞİ — denetim izi
# =====================================================================

def gecmis_kaydet(ozet: pd.DataFrame, durum: pd.DataFrame, istisna_n: int):
    import os, datetime
    simdi = datetime.datetime.now()
    toplam = len(durum)
    dogru = int((durum["Genel Durum"] == "DOĞRU").sum())
    satir = {"Tarih": simdi.strftime("%d.%m.%Y"),
             "Saat": simdi.strftime("%H:%M"),
             "Toplam Malzeme": toplam, "Doğru": dogru,
             "Hatalı": toplam - dogru,
             "Veri Sağlığı %": round(100 * dogru / toplam, 1) if toplam else 0,
             "Onaylı İstisna": istisna_n}
    for _, s in ozet.iterrows():
        satir[str(s["Kural"])[:2] + " Bulgu"] = int(s["Bulgu Sayısı"])
    yeni = pd.DataFrame([satir])
    if os.path.exists(GECMIS_DOSYASI):
        try:
            eski = pd.read_excel(GECMIS_DOSYASI)
            yeni = pd.concat([eski, yeni], ignore_index=True)
        except Exception:
            pass
    yeni.to_excel(GECMIS_DOSYASI, index=False)
    print(f"Geçmiş kaydı  : {GECMIS_DOSYASI} ({len(yeni)}. koşum)")


# =====================================================================
# KURAL 5 — Malzeme türü kontrolü
# =====================================================================

def kapsam_filtrele(df: pd.DataFrame, gecerli, ad: str):
    """Yalnızca geçerli türdeki kayıtları analize alır. 1-5 dahil kapsam
    dışı türler HATA DEĞİLDİR; sessizce atlanır, sadece bilgi verilir."""
    if "Malzeme türü" not in df.columns:
        return df, 0
    t = (df["Malzeme türü"].fillna("").astype(str)
         .str.strip().str.lstrip("0"))
    m = t.isin(gecerli)
    dis = int((~m).sum())
    if dis:
        turler = sorted(set(t[~m].replace("", "(boş)")))
        print(f"Bilgi: {ad} dosyasında {dis} kayıt kapsam dışı türde "
              f"({', '.join(turler)}) — hata sayılmadan analiz dışı bırakıldı.")
    return df[m].reset_index(drop=True), dis


# =====================================================================
# DİL DESTEĞİ (Excel çıktıları için TR varsayılan, EN çeviri)
# =====================================================================

BASLIK_EN = {
    "Malzeme Kısa Metni": "Material Description",
    "Alt Tarif Sayısı": "Sub-description Count",
    "Malzeme": "Material", "Üst Kod": "Parent Code", "Kural": "Rule",
    "Hata Nedeni": "Error Reason", "İlgili Alan": "Related Field",
    "Mevcut Değer": "Current Value", "Olması Gereken": "Expected Value",
    "Yapılacak Düzeltme": "Correction To Apply", "Kaynak": "Source",
    "Tanım": "Description", "Değer": "Value", "Toplam": "Total",
    "Hatalı": "Faulty", "Hata %": "Error %", "Bulgu": "Findings",
    "Malzeme Sayısı": "Material Count", "Malzeme Türü": "Material Type",
    "Beklenen": "Expected", "Sorun": "Issue",
}
CUMLE_EN = [
    ("Alan değeri üst kod ile uyuşmuyor", "Field value does not match parent code"),
    ("Barkod işareti üst kod ile uyuşmuyor", "Barcode flag does not match parent code"),
    ("(işaretsiz)", "(not flagged)"), ("X (işaretli)", "X (flagged)"),
    ("Üst kod (Temel malzeme) boş", "Parent code (base material) is empty"),
    ("Üst kod listede bulunamadı", "Parent code not found in list"),
    ("Malzemeye geçerli bir üst kod (Temel malzeme) bağlayın",
     "Assign a valid parent code (base material)"),
    ("Temel malzeme alanındaki kodu kontrol edin — üst kod listesinde yok",
     "Check the base material code — not found in parent list"),
    ("' alanını üst kodun değeriyle eşitleyin: ", "' field must equal parent value: "),
    ("Tanımın sonunda '", "Description ends with '"),
    ("' var ama ", "' but "),
    (" alanı doldurulmamış veya uyuşmuyor", " field is empty or mismatched"),
    (" alanında '", " field has '"),
    ("' seçili ama tanımın sonunda yok",
     "' selected but missing at the end of the description"),
    (" alanına '", " field: enter '"),
    ("' girilmeli", "'"),
    ("Tanımın sonuna '", "Append '"),
    ("' eklenmeli", "' to the end of the description"),
    ("Tanımda yasaklı kelime var", "Forbidden word in description"),
    ("Tanımda yasaklı kelime", "Forbidden word in description"),
    ("Tanımdan şu ifadeleri çıkarın: ", "Remove from description: "),
    ("Yasaklı kelimesiz tanım", "Description without forbidden words"),
    ("Üst kodun alt tarifleri — aynı tanım birden fazla kodda (bilgi)",
     "Sub-descriptions of the parent — same description on multiple codes (info)"),
    ("Üst kodun alt tarifleri (bilgi — hata değil)",
     "Sub-descriptions of the parent (info — not an error)"),
    ("Aynı tanımı paylaşan ", "Codes sharing this description ("),
    (" kod: ", "): "),
    ("Aynı tanım birden fazla kodda", "Same description on multiple codes"),
    ("Aynı tanım birden fazla malzeme kodunda kullanılmış",
     "Same description used by multiple material codes"),
    ("Her kod için benzersiz tanım", "Unique description per code"),
    ("Bu tanımı paylaşan ", "Separate or merge the "),
    (" kodu ayrıştırın veya kayıtları birleştirin",
     " codes sharing this description"),
    ("Beklenmeyen malzeme türü", "Unexpected material type"),
    ("Malzeme türünü kontrol edin — beklenen: ",
     "Check the material type — expected: "),
    ("Ürün Kodu", "Product Code"), ("Üst Kod", "Parent Code"),
    ("(boş)", "(empty)"), ("DOĞRU", "OK"), ("HATALI", "ERROR"),
]


def L(tr_metin: str) -> str:
    """Sabit etiketler için dil seçimi (AUTO algılamadan sonra çözülür)."""
    if DIL != "EN":
        return tr_metin
    return BASLIK_EN.get(tr_metin, tr_metin)


def ceviri_metin(s):
    if DIL != "EN" or not isinstance(s, str):
        return s
    for tr, en in CUMLE_EN:
        s = s.replace(tr, en)
    return s


def ceviri_df(df: pd.DataFrame) -> pd.DataFrame:
    """DIL=EN ise başlıkları ve üretilmiş metinleri çevirir."""
    if DIL != "EN" or df.empty:
        return df
    df = df.rename(columns=BASLIK_EN).copy()
    for kolon in df.columns:
        df[kolon] = df[kolon].map(ceviri_metin)
    return df


def birlesik_ayristir(df: pd.DataFrame):
    """Tek dosyadaki karışık listeyi ürün ve üst kod olarak ayırır.
    RC ile başlayan kodlar üst kod, rakamla başlayanlar ürün kodudur."""
    if df.empty or "Malzeme" not in df.columns:
        return pd.DataFrame(), pd.DataFrame(), 0
    kod = df["Malzeme"].astype(str).str.strip().str.upper()
    ust_m = kod.str.startswith(tr_upper(UST_KOD_ONEKI))
    urun_m = kod.str.match(r"^\d")
    disi = int((~ust_m & ~urun_m).sum())
    return (df[urun_m].reset_index(drop=True),
            df[ust_m].reset_index(drop=True), disi)


def silinmis_malzemeler(urun: pd.DataFrame, ust: pd.DataFrame) -> set:
    """'ÜB dzy.silme iştr.' alanı dolu olan malzemeler silinmiş sayılır ve
    hiçbir kuralda bulgu üretmez."""
    kodlar = set()
    for df in (urun, ust):
        if df is None or df.empty:
            continue
        kolon = silme_isareti_kolonu(df)
        if not kolon:
            continue
        isaret = normalize_seri(df[kolon])
        kodlar |= set(df.loc[isaret != "", "Malzeme"].astype(str).str.strip())
    return kodlar


def silinmis_dus(df: pd.DataFrame, silinmis: set):
    """Bulgu tablosundan silinmiş malzemeleri çıkarır.
    İSTİSNA: üst kodun silinmesinden doğan 'bu ürün de silinmeli' bulguları
    korunur — ürün işaretli olsa bile silme işleminin takibi gerekir."""
    if df is None or df.empty or not silinmis:
        return df, 0
    m = df["Malzeme"].astype(str).str.strip().isin(silinmis)
    for kolon in ("Alan", "İlgili Alan"):
        if kolon in df.columns:
            kaskad = df[kolon].astype(str).map(
                lambda v: any(a in tr_upper(v) for a in K5_SILME_ANAHTARLARI))
            m = m & ~kaskad
            break
    return df[~m].reset_index(drop=True), int(m.sum())


# =====================================================================
# KURAL 4 — Çokluk (xN) ↔ MARM adet birimi sayacı
# =====================================================================

def marm_oku(yol: str = None) -> pd.DataFrame:
    """MARM (alternatif ölçü birimleri) dosyasını okur ve sütunlarını
    standart adlara indirger: Malzeme | Birim | Sayaç | Payda."""
    import os
    yol = yol or MARM_DOSYASI
    if not os.path.exists(yol):
        return pd.DataFrame()
    try:
        m = dosya_oku(yol)
    except Exception as e:
        print(f"UYARI: {yol} okunamadı ({e}).")
        return pd.DataFrame()

    def bul(*anahtarlar, haric=()):
        for c in m.columns:
            u = tr_upper(str(c))
            if any(h in u for h in haric):
                continue
            if any(a in u for a in anahtarlar):
                return c
        return None

    kolon_malzeme = bul("MALZEME", "MATERIAL", "MATNR")
    kolon_birim = bul("ALTERNATİF", "ALTERNATIF", "AÖB", "AOB", "MEINH",
                      "ALT. UOM", "ALTERNATIVE UNIT", "ÖLÇÜ BİRİMİ",
                      "OLCU BIRIMI", "UOM", "BİRİM", "BIRIM",
                      haric=("TEMEL", "BASE"))
    kolon_sayac = bul("SAYAÇ", "SAYAC", "UMREZ", "NUMERATOR", "PAY ",
                      haric=("PAYDA",))
    kolon_payda = bul("PAYDA", "UMREN", "DENOMINATOR")

    if not (kolon_malzeme and kolon_birim and kolon_sayac):
        print("UYARI: MARM dosyasında Malzeme / Alternatif ölçü birimi / "
              "Sayaç sütunları bulunamadı. Kural 4 atlanacak.")
        print(f"       Görülen sütunlar: {list(m.columns)[:12]}")
        return pd.DataFrame()

    d = pd.DataFrame({
        "Malzeme": m[kolon_malzeme].astype(str).str.strip(),
        "Birim": normalize_seri(m[kolon_birim]),
        "Sayaç": pd.to_numeric(m[kolon_sayac], errors="coerce"),
    })
    d["Payda"] = (pd.to_numeric(m[kolon_payda], errors="coerce")
                  if kolon_payda else 1)
    d = d[d["Malzeme"] != ""]
    print(f"Bilgi: MARM yüklendi — {len(d)} satır "
          f"({d['Birim'].nunique()} farklı ölçü birimi).")
    return d


def tanim_cokluk(tanim_norm: str, tek: set, cift: set):
    """Tanımın sonundaki çokluk sayısını döndürür (yoksa None).
    Önce Pazar/Ek Alan/Ambalaj kodları soyulur: '...200MLX24RU LF' -> 24"""
    if not tanim_norm:
        return None
    parcalar = [p for p in re.split(r"[\s\-_/.]+", tanim_norm) if p]
    # Sondaki kod eklerini soy
    i = len(parcalar)
    while i > 1:
        if i >= 2 and " ".join(parcalar[i-2:i]) in cift:
            i -= 2
            continue
        if parcalar[i-1] in tek:
            i -= 1
            continue
        break
    if i < 1:
        return None
    son = parcalar[i-1]
    # Bitişik kod varsa onu da soy: "200MLX24RU" -> "200MLX24"
    bitisik = _bitisik_kod_bul(son, tek | cift)
    if bitisik:
        son = son[: -len(bitisik)]
    eslesme = re.search(COKLUK_DESENI, son)
    if not eslesme:
        return None
    try:
        return int(eslesme.group(1))
    except ValueError:
        return None


def kural4(urun: pd.DataFrame, ust: pd.DataFrame, marm: pd.DataFrame,
           urun_tanim: str, tek: set, cift: set) -> pd.DataFrame:
    """Tanım sonundaki çokluk (xN), MARM'daki adet birimi sayacıyla
    karşılaştırılır. Ürünün kendi kaydı yoksa üst koduna bakılır."""
    if marm.empty or not urun_tanim or urun.empty:
        return pd.DataFrame()

    haric = {tr_upper(b) for b in MARM_HARIC_BIRIMLER}
    kullanilir = marm[~marm["Birim"].isin(haric)].copy()
    if kullanilir.empty:
        print("UYARI: MARM'da çokluk doğrulamasına uygun birim satırı yok; "
              "Kural 4 atlanacak.")
        return pd.DataFrame()

    # Malzeme -> MARM'da o malzeme için geçen SAYAÇ değerleri kümesi.
    # Tanımın sonundaki çokluk bu değerlerden biriyle eşleşmelidir.
    kullanilir["Sayaç"] = pd.to_numeric(kullanilir["Sayaç"], errors="coerce")
    kullanilir = kullanilir.dropna(subset=["Sayaç"])

    cokluk_kumesi, cokluk_birim = {}, {}
    for _, s in kullanilir.iterrows():
        kod = str(s["Malzeme"]).strip()
        try:
            sayac = float(s["Sayaç"])
        except (TypeError, ValueError):
            continue
        if sayac <= 0 or abs(sayac - round(sayac)) > 1e-6:
            continue
        deger = int(round(sayac))
        cokluk_kumesi.setdefault(kod, set()).add(deger)
        cokluk_birim.setdefault(kod, {}).setdefault(deger, str(s["Birim"]))

    # MARA ürün hiyerarşisi muafiyeti: bu kodlara sahip ürünler K4 dışıdır.
    hiyerarsi_kolonu = hiyerarsi_sutunu_bul(urun) if K4_MUAF_HIYERARSILER else ""
    muaf_sayaci = 0

    bulgular = []
    dogrulanamayan = 0
    coklukla = 0
    for _, s in urun.iterrows():
        malzeme = str(s["Malzeme"]).strip()
        if hiyerarsi_kolonu and hiyerarsi_normalize(
                s.get(hiyerarsi_kolonu)) in K4_MUAF_HIYERARSILER:
            muaf_sayaci += 1
            continue        # muaf hiyerarşi — çokluk kontrolü yapılmaz
        tanim = str(s.get(urun_tanim, "") or "")
        n = tanim_cokluk(normalize(tanim), tek, cift)
        if n is None:
            continue
        coklukla += 1
        # Malzemenin kendi MARM çevrim faktörleri esas alınır. Hiç
        # paketleme çevrimi tanımlı değilse doğrulama yapılamaz.
        gecerli = cokluk_kumesi.get(malzeme)
        if not gecerli:
            dogrulanamayan += 1        # MARM'da bu malzeme için kayıt yok
            continue
        if n not in gecerli:
            beklenen = sorted(gecerli)
            birimler = cokluk_birim.get(malzeme, {})
            bulgular.append({
                "Kaynak": "Ürün Kodu", "Malzeme": malzeme,
                "Tanım": tanim, "Tanımdaki Çokluk": n,
                "MARM Çokluk": ", ".join(f"x{d}" for d in beklenen),
                "MARM Birimi": ", ".join(
                    f"{birimler.get(d, '')}={d}" for d in beklenen),
                "MARM Kaydı": malzeme,
                "Sorun": ("Tanımdaki çokluk, MARM'daki hiçbir sayaç "
                          "değeriyle eşleşmiyor")})

    print(f"  Çokluk taşıyan ürün: {coklukla:,} · MARM'da paketleme çevrimi "
          f"bulunmayan (doğrulanamayan): {dogrulanamayan:,}")
    if hiyerarsi_kolonu:
        print(f"  K4 muaf (ürün hiyerarşisi): {muaf_sayaci:,} ürün "
              f"('{hiyerarsi_kolonu}' sütunu)")
    elif K4_MUAF_HIYERARSILER:
        print("  Bilgi: ürün hiyerarşisi sütunu bulunamadı — K4 muafiyeti "
              "uygulanamadı (export'a 'Ürün hiyerarşisi' sütununu ekleyin).")
    return pd.DataFrame(bulgular)


# =====================================================================
# KURAL 5 — Üst kod tekilliği ve silme işareti
# =====================================================================

def silme_isareti_kolonu(df: pd.DataFrame) -> str:
    """'Üst birim düzeyindeki malzemeyi silmek için işaretle' sütununu bulur."""
    for c in df.columns:
        u = tr_upper(str(c))
        if any(a in u for a in K5_SILME_ANAHTARLARI):
            return c
    return ""


def kural5(ust: pd.DataFrame, marm: pd.DataFrame, urun: pd.DataFrame,
           ust_tanim: str, tek: set, cift: set) -> pd.DataFrame:
    """Üst kodların (RC) ayırt edici kombinasyonu benzersiz olmalıdır.

    Karşılaştırılan bileşenler (K5_PARMAK_IZI_ALANLARI + koli içi adet):
      MARA: Marka1 · Raporlama Markası · Raporlama Alt Markası · Varyant ·
            Ürün Boyutu · S&OP Kategorisi  (varsayılan altı alan; yalnızca
            dosyada MEVCUT olanlar kullanılır)
      MARM: koli içi adet — malzeme kısa metnindeki çokluk (xN) MARM'daki
            sayaç değerleri arasında da bulunuyorsa DOĞRULANMIŞ sayılır ve
            imzaya bu değer girer. Doğrulanamayan üst kod K5'e alınmaz.

    Bu alanların tamamı ile doğrulanmış koli içi adet birlikte imzayı
    oluşturur. Aynı imzayı paylaşan iki veya daha fazla üst kod bulgu üretir; her
    satırda o üst koda bağlı ürün (alt tarif) sayısı da belirtilir.
    """
    if ust.empty:
        return pd.DataFrame()

    alanlar = [c for c in K5_PARMAK_IZI_ALANLARI if c in ust.columns]
    if not alanlar:
        print("UYARI: Kural 5 için ayırt edici alanlar bulunamadı "
              f"(aranan: {', '.join(K5_PARMAK_IZI_ALANLARI)}).")
        return pd.DataFrame()

    # --- MARM sayaç kümeleri ---
    sayac_kume = {}
    if not marm.empty:
        gecici = marm.copy()
        gecici["Sayaç"] = pd.to_numeric(gecici["Sayaç"], errors="coerce")
        gecici = gecici.dropna(subset=["Sayaç"])
        for kod, grup in gecici.groupby(
                gecici["Malzeme"].astype(str).str.strip()):
            sayac_kume[kod] = {int(v) for v in grup["Sayaç"] if v > 0}

    # --- Üst koda bağlı ürün (alt tarif) sayısı ---
    alt_tarif_sayisi = {}
    if not urun.empty and "Temel malzeme" in urun.columns:
        bag = urun[["Malzeme", "Temel malzeme"]].copy()
        bag["_ust"] = bag["Temel malzeme"].astype(str).str.strip()
        bag = bag[bag["_ust"] != ""]
        alt_tarif_sayisi = (bag.groupby("_ust")["Malzeme"]
                            .nunique().to_dict())

    d = ust.copy()
    d["_kod"] = d["Malzeme"].astype(str).str.strip()

    # --- Doğrulanmış koli içi adet ---
    koli, dogrulanamayan = [], 0
    for _, s in d.iterrows():
        kod = s["_kod"]
        n = (tanim_cokluk(normalize(str(s.get(ust_tanim, "") or "")),
                          tek, cift) if ust_tanim else None)
        kume = sayac_kume.get(kod, set())
        koli.append(str(n) if (n is not None and n in kume) else "")
    d["_koli"] = koli
    dogrulanamayan = int((d["_koli"] == "").sum())

    # Koli içi adedi doğrulanamayan üst kodlar K5 kapsamına alınmaz
    d = d[d["_koli"] != ""]
    if dogrulanamayan:
        print(f"  Koli içi adedi doğrulanamayan üst kod: {dogrulanamayan:,} "
              f"(tanımdaki çokluk MARM sayacında yok) — K5 dışı")
    if d.empty:
        return pd.DataFrame()

    sutunlar = [normalize_seri(d[c]) for c in alanlar] + [d["_koli"]]
    d["_imza"] = ["|".join(str(v) for v in satir) for satir in zip(*sutunlar)]

    bulgular = []
    for imza, grup in d.groupby("_imza"):
        kodlar = sorted(set(grup["_kod"]))
        if len(kodlar) < 2:
            continue
        parcalar = imza.split("|")
        okunur = "; ".join(
            f"{a}={v}" for a, v in zip(alanlar, parcalar) if v)
        okunur += f"; koli içi adet={parcalar[-1]}"
        for kod in kodlar:
            digerleri = [k for k in kodlar if k != kod]
            alt_n = int(alt_tarif_sayisi.get(kod, 0))
            bulgular.append({
                "Kaynak": "Üst Kod", "Malzeme": kod,
                "Eşleşen Kod Sayısı": len(kodlar),
                "Eşleşen Kodlar": ", ".join(digerleri),
                "Alt Tarif Sayısı": alt_n,
                "Eşleşenlerin Alt Tarifleri": ", ".join(
                    f"{k}={int(alt_tarif_sayisi.get(k, 0))}" for k in digerleri),
                "Koli İçi Adet": parcalar[-1],
                "Ayırt Edici Değerler": okunur,
                "Sorun": (f"Üst kodun ayırt edici {len(alanlar) + 1} bileşeni "
                          f"({', '.join(alanlar)} + koli içi adet) başka bir "
                          f"üst kodla birebir aynı")})
    return pd.DataFrame(bulgular)


# =====================================================================
# KURAL 6 — Malzeme kodunun son iki hanesi ↔ Menşei alanı
# =====================================================================

def mensei_koda_cevir(deger) -> str:
    """Menşei alanındaki değeri iki haneli sayısal koda indirger.

    '10' / '10.0' / ' 10 ' -> '10';  'TR' veya 'TR VAS' gibi metin
    karşılıkları sözlükten koda çevrilir. Çözülemezse ham değer döner
    (bu durumda karşılaştırma zaten uyuşmazlık verir)."""
    ham = str("" if deger is None else deger).strip()
    # Excel'den gelen boş hücreler NaN olur; str(NaN) -> "nan" olduğu için
    # açıkça boşa çevrilmelidir (aksi hâlde "boş" yerine "uyuşmuyor" denir).
    try:
        if pd.isna(deger):
            return ""
    except (TypeError, ValueError):
        pass
    if ham.lower() in ("nan", "none", "nat", "<na>"):
        return ""
    if ham.endswith(".0"):
        ham = ham[:-2]
    if not ham:
        return ""
    kodlar = KURAL2_KATEGORILER.get("Menşei", {})
    if ham in kodlar:
        return ham
    # Tek haneli girilmiş olabilir: "9" -> "09"
    if ham.isdigit() and len(ham) == 1 and ham.zfill(2) in kodlar:
        return ham.zfill(2)
    hedef = _harf_katla(tr_upper(ham))
    for kod, metin in kodlar.items():
        if metin and _harf_katla(tr_upper(str(metin))) == hedef:
            return kod
    return ham


def kural6(urun: pd.DataFrame, tanim_kolonu: str = None) -> pd.DataFrame:
    """Ürün kodunun son iki hanesi Menşei alanındaki kodla aynı olmalıdır.

    Örn. 60002810 -> son iki hane "10" (TR). Menşei "21" (MY VAS) ise
    uyuşmazlık bulgusu üretilir. Menşei boş olan malzemeler de bildirilir.
    Yalnızca ÜRÜN kodlarına (rakamla başlayan, son iki hanesi sayı olan)
    uygulanır; üst kodlar (RC…) kapsam dışıdır."""
    if not K6_AKTIF or urun.empty or "Malzeme" not in urun.columns:
        return pd.DataFrame()
    if "Menşei" not in urun.columns:
        print("UYARI: 'Menşei' sütunu bulunamadı — Kural 6 atlanacak.")
        return pd.DataFrame()

    kodlar = KURAL2_KATEGORILER.get("Menşei", {})
    gecerli_kodlar = {str(k).strip() for k in kodlar}

    d = urun.copy()
    d["_kod"] = d["Malzeme"].astype(str).str.strip()
    # Yalnızca sayısal ürün kodları (RC… ve bozuk kodlar kapsam dışı)
    d = d[d["_kod"].str.fullmatch(r"\d{3,}")]
    if d.empty:
        return pd.DataFrame()
    d["_son2"] = d["_kod"].str[-2:]
    d["_mensei"] = d["Menşei"].map(mensei_koda_cevir)

    # Son iki hanesi geçerli bir menşei kodu olmayanlar
    tanimsiz = ~d["_son2"].isin(gecerli_kodlar)
    tanimsiz_n = int(tanimsiz.sum())
    if tanimsiz_n and not K6_GECERSIZ_SON_KOD_BILDIR:
        d = d[~tanimsiz]

    parcalar = []
    kolonlar = ["Kaynak", "Malzeme", "Tanım", "Kod Son 2 Hane",
                "Menşei Değeri", "Beklenen Menşei", "Sorun",
                "Önerilen Düzeltme"]

    def _tanim(alt):
        return (alt[tanim_kolonu].astype(str)
                if tanim_kolonu and tanim_kolonu in alt.columns else "")

    def _metin(kod):
        aciklama = str(kodlar.get(kod, "") or "")
        return f"{kod} ({aciklama})" if aciklama else str(kod)

    # 1) Menşei boş
    bos = d["_mensei"] == ""
    if bos.any():
        alt = d[bos]
        parcalar.append(pd.DataFrame({
            "Kaynak": "Ürün Kodu", "Malzeme": alt["_kod"], "Tanım": _tanim(alt),
            "Kod Son 2 Hane": alt["_son2"], "Menşei Değeri": "(boş)",
            "Beklenen Menşei": alt["_son2"].map(_metin),
            "Sorun": "Menşei alanı boş — malzeme kodunun son iki hanesi dolu",
            "Önerilen Düzeltme": ("Menşei alanına kodun son iki hanesini girin: "
                                  + alt["_son2"].map(_metin))}))

    # 2) Menşei dolu ama kodun son iki hanesiyle uyuşmuyor
    uyusmaz = (~bos) & (d["_mensei"] != d["_son2"])
    if uyusmaz.any():
        alt = d[uyusmaz]
        parcalar.append(pd.DataFrame({
            "Kaynak": "Ürün Kodu", "Malzeme": alt["_kod"], "Tanım": _tanim(alt),
            "Kod Son 2 Hane": alt["_son2"],
            "Menşei Değeri": alt["_mensei"].map(_metin),
            "Beklenen Menşei": alt["_son2"].map(_metin),
            "Sorun": ("Menşei, malzeme kodunun son iki hanesiyle uyuşmuyor"),
            "Önerilen Düzeltme": ("Menşei alanını kodun son iki hanesiyle "
                                  "eşitleyin: " + alt["_son2"].map(_metin)
                                  + " — ya da malzeme kodunu gözden geçirin")}))

    # 3) (opsiyonel) Kodun son iki hanesi tanımlı bir menşei değil
    if tanimsiz_n and K6_GECERSIZ_SON_KOD_BILDIR:
        alt = d[tanimsiz]
        parcalar.append(pd.DataFrame({
            "Kaynak": "Ürün Kodu", "Malzeme": alt["_kod"], "Tanım": _tanim(alt),
            "Kod Son 2 Hane": alt["_son2"],
            "Menşei Değeri": alt["_mensei"].replace("", "(boş)"),
            "Beklenen Menşei": "(tanımlı menşei kodu)",
            "Sorun": ("Malzeme kodunun son iki hanesi tanımlı bir menşei "
                      "kodu değil"),
            "Önerilen Düzeltme": ("Malzeme kodunu veya menşei kod listesini "
                                  "kontrol edin")}))

    if tanimsiz_n and not K6_GECERSIZ_SON_KOD_BILDIR:
        print(f"  Bilgi: {tanimsiz_n:,} üründe kodun son iki hanesi tanımlı "
              f"bir menşei kodu değil — K6 dışı bırakıldı.")

    if not parcalar:
        return pd.DataFrame(columns=kolonlar)
    return (pd.concat(parcalar, ignore_index=True)[kolonlar]
            .sort_values("Malzeme").reset_index(drop=True))



# =====================================================================
# BASİT DÜZELTME TABLOSU (kullanıcı dostu tek bakışta yapı)
# =====================================================================

def basit_duzeltme_tablosu(k1, k2, k3, k4, k5, alt, urun, tanim_map=None,
                           birlestir=True, k4_satirlari=False,
                           k6=None) -> pd.DataFrame:
    """Her satırı kendi başına eksiksiz bir düzeltme talimatı olan sade tablo:
    Malzeme | Üst Kod | Kural | Hata Nedeni | İlgili Alan |
    Mevcut Değer | Olması Gereken | Yapılacak Düzeltme"""
    ust_kod_haritasi = {}
    if "Malzeme" in urun.columns and "Temel malzeme" in urun.columns:
        ust_kod_haritasi = dict(zip(urun["Malzeme"].astype(str),
                                    urun["Temel malzeme"].fillna("").astype(str)))
    satirlar = []

    if not k1.empty:
        for _, s in k1.iterrows():
            if s["Sorun"] == "Üst kod (Temel malzeme) boş":
                duzeltme = "Malzemeye geçerli bir üst kod (Temel malzeme) bağlayın"
                olmasi = "Geçerli bir RC üst kodu"
            elif s["Sorun"] == "Üst kod listede bulunamadı":
                duzeltme = "Temel malzeme alanındaki kodu kontrol edin — üst kod listesinde yok"
                olmasi = "Listedeki geçerli bir üst kod"
            else:
                duzeltme = (f"'{s['Alan']}' alanını üst kodun değeriyle eşitleyin: "
                            f"{s['Üst Kod Değeri']}")
                olmasi = s["Üst Kod Değeri"]
            satirlar.append({
                "Malzeme": s["Malzeme"], "Üst Kod": s["Üst Kod"], "Kural": "K1",
                "Hata Nedeni": s["Sorun"], "İlgili Alan": s["Alan"],
                "Mevcut Değer": s["Ürün Değeri"], "Olması Gereken": olmasi,
                "Yapılacak Düzeltme": duzeltme})

    if not k2.empty:
        # Aynı malzemede birden çok kod varsa (örn. "…24RU LF" -> RU ve LF)
        # bulgular TEK satırda, alt alta iki ayrı hata olarak gösterilir.
        gruplar = ([(m, g) for m, g in k2.groupby("Malzeme", sort=False)]
                   if birlestir else
                   [(s["Malzeme"], k2.iloc[[i]]) for i, (_, s) in enumerate(k2.iterrows())])
        for malzeme, grup in gruplar:
            nedenler, alan_listesi, olmasi_listesi, duzeltmeler = [], [], [], []
            for _, s in grup.iterrows():
                alanlar = "; ".join(p.split("=")[0].strip() for p in
                                    str(s.get("Kontrol Edilen Alanlar", "")).split(";") if p)
                onerilen_tanim = str(s.get("Önerilen Tanım", "") or "").strip()
                olmasi = onerilen_tanim if onerilen_tanim else str(s["Anahtar Kelime"])
                nedenler.append(str(s["Sorun"]))
                alan_listesi.append(alanlar or "Tanım")
                olmasi_listesi.append(olmasi)
                duzeltmeler.append(str(s.get("Önerilen Düzeltme", "") or ""))
            cok = len(nedenler) > 1

            def isaret(liste, _cok=cok):
                if not _cok:
                    return liste[0]
                return "\n".join(f"{i+1}) {v}" for i, v in enumerate(liste))
            satirlar.append({
                "Malzeme": malzeme,
                "Üst Kod": ust_kod_haritasi.get(str(malzeme), ""),
                "Kural": "K2", "Hata Nedeni": isaret(nedenler),
                "İlgili Alan": isaret(alan_listesi),
                "Mevcut Değer": grup.iloc[0]["Tanım"],
                "Olması Gereken": isaret(olmasi_listesi),
                "Yapılacak Düzeltme": isaret(duzeltmeler)})

    if not k3.empty:
        for _, s in k3.iterrows():
            satirlar.append({
                "Malzeme": s["Malzeme"],
                "Üst Kod": ust_kod_haritasi.get(str(s["Malzeme"]), ""),
                "Kural": "K3", "Hata Nedeni": "Tanımda yasaklı kelime",
                "İlgili Alan": "Tanım", "Mevcut Değer": s["Tanım"],
                "Olması Gereken": "Yasaklı kelimesiz tanım",
                "Yapılacak Düzeltme": (f"Tanımdan şu ifadeleri çıkarın: "
                                       f"{s['Yasaklı Kelimeler']}")})

    if k4_satirlari and not k4.empty:
        for _, s in k4.iterrows():
            n = int(s.get("Tekrar Sayısı", 0) or 0)
            satirlar.append({
                "Malzeme": s["Malzeme"],
                "Üst Kod": ust_kod_haritasi.get(str(s["Malzeme"]), ""),
                "Kural": "K4",
                "Hata Nedeni": "Üst kodun alt tarifleri (bilgi — hata değil)",
                "İlgili Alan": "Tanım", "Mevcut Değer": s.get("Tanım", ""),
                "Olması Gereken": "—",
                "Yapılacak Düzeltme": f"Aynı tanımı paylaşan {n} kod"})

    if not k4.empty:
        for _, s in k4.iterrows():
            satirlar.append({
                "Malzeme": s["Malzeme"],
                "Üst Kod": ust_kod_haritasi.get(str(s["Malzeme"]), ""),
                "Kural": "K4", "Hata Nedeni": s["Sorun"],
                "İlgili Alan": "Çokluk (xN) / MARM sayaç",
                "Mevcut Değer": f"tanım: x{s['Tanımdaki Çokluk']}",
                "Olması Gereken": f"{s['MARM Çokluk']} ({s['MARM Birimi']})",
                "Yapılacak Düzeltme": (
                    f"MARM'da tanımlı sayaç değerleri: {s['MARM Çokluk']} — "
                    f"tanımı bunlardan biriyle eşitleyin veya MARM kaydını "
                    f"düzeltin")})

    if not k5.empty:
        for _, s in k5.iterrows():
            satirlar.append({
                "Malzeme": s["Malzeme"],
                "Üst Kod": s["Eşleşen Kodlar"],
                "Kural": "K5", "Hata Nedeni": s["Sorun"],
                "İlgili Alan": ("Ayırt edici alanlar (Marka1 / Raporlama "
                                "Markası / Raporlama Alt Markası / Varyant / "
                                "Ürün Boyutu / S&OP) + koli içi adet"),
                "Mevcut Değer": s["Ayırt Edici Değerler"],
                "Olması Gereken": "Her üst kod için benzersiz kombinasyon",
                "Yapılacak Düzeltme": (
                    f"Aynı kombinasyona sahip {s['Eşleşen Kod Sayısı']} üst kod "
                    f"— bu kodun {s['Alt Tarif Sayısı']} alt tarifi var; "
                    f"eşleşenler: {s['Eşleşenlerin Alt Tarifleri']}. Birini "
                    f"kaldırın veya ayırt edici alanları farklılaştırın")})

    if k6 is not None and not k6.empty:
        for _, s in k6.iterrows():
            satirlar.append({
                "Malzeme": s["Malzeme"],
                "Üst Kod": ust_kod_haritasi.get(str(s["Malzeme"]), ""),
                "Kural": "K6", "Hata Nedeni": s["Sorun"],
                "İlgili Alan": "Menşei / Malzeme kodu son 2 hane",
                "Mevcut Değer": (f"menşei: {s['Menşei Değeri']} · "
                                 f"kod sonu: {s['Kod Son 2 Hane']}"),
                "Olması Gereken": s["Beklenen Menşei"],
                "Yapılacak Düzeltme": s["Önerilen Düzeltme"]})

    df = pd.DataFrame(satirlar)
    if not df.empty:
        tanim_map = tanim_map or {}
        df.insert(0, "Malzeme Kısa Metni",
                  df["Malzeme"].astype(str).map(tanim_map).fillna(""))
        # K4 (alt tarifler) ayrı satır olarak listelenmez; aynı tanımı paylaşan
        # kod sayısı "Olması Gereken"den hemen sonra bilgi sütunu olarak eklenir.
        alt_map = {}
        if not alt.empty and "Tekrar Sayısı" in alt.columns:
            alt_map = dict(zip(alt["Malzeme"].astype(str),
                               alt["Tekrar Sayısı"].astype(int)))
        konum = list(df.columns).index("Olması Gereken") + 1
        df.insert(konum, "Alt Tarif Sayısı",
                  df["Malzeme"].astype(str).map(alt_map).fillna(1).astype(int))
        df = df.sort_values(["Malzeme", "Kural"]).reset_index(drop=True)
    return df


# =====================================================================
# KATEGORİ ANALİZİ EXCEL'İ
# =====================================================================

# Öncelikli 4 kategori en başta; kalanlar sağda devam eder
KATEGORI_SIRASI = [
    "Malzeme türü", "Mal grubu", "Raporlama Markası", "Raporlama Alt Markası",
    "Menşei", "Pazar", "Ek Alan", "Ambalaj Tipi",
    "S&OP Kategorisi", "Ürün Boyutu", "Temel ölçü birimi", "SKU Grup",
]
KATEGORI_UST_SINIR = 20   # bir kategoride bundan çok değer varsa ilk 20 + "Diğer"


def kategori_analiz_excel(urun, ust, durum, yol=None, wb=None):
    """Her kategori için: Değer | Toplam | Hatalı | Hata %  tabloları.
    wb verilirse o çalışma kitabına 'Kategori_Analizi' sayfası eklenir;
    verilmezse yol'a ayrı dosya yazar (eski davranış)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    tum = pd.concat([urun, ust], ignore_index=True)
    hatali_set = set(durum.loc[durum["Genel Durum"] == "HATALI", "Malzeme"].astype(str))

    ayri_dosya = wb is None
    if ayri_dosya:
        wb = Workbook()
        ws = wb.active
    else:
        ws = wb.create_sheet()
    ws.title = "Kategori_Analizi"

    lacivert = "1B2A3A"
    baslik_f = Font(bold=True, color="FFFFFF", size=10)
    baslik_d = PatternFill("solid", fgColor=lacivert)
    kategori_d = PatternFill("solid", fgColor="33628C")
    toplam_d = PatternFill("solid", fgColor="E8EFF5")
    zebra = PatternFill("solid", fgColor="F5F8FA")
    kirmizi = Font(color="C0392B")
    ince = Side(style="thin", color="DFE6EA")
    kenar = Border(bottom=ince)

    kategoriler = {}
    sutun = 1
    for kategori in KATEGORI_SIRASI:
        if kategori not in tum.columns:
            continue
        seri = tum[kategori].fillna("").astype(str).str.strip()
        seri = seri.replace("", "(boş)")
        gruplar = tum.assign(_deger=seri).groupby("_deger").agg(
            Toplam=("Malzeme", "nunique"),
            Hatalı=("Malzeme", lambda s: s.astype(str).isin(hatali_set).sum()),
        ).sort_values("Toplam", ascending=False)

        if len(gruplar) > KATEGORI_UST_SINIR:
            ilk = gruplar.head(KATEGORI_UST_SINIR)
            kalan = gruplar.iloc[KATEGORI_UST_SINIR:]
            diger = pd.DataFrame({"Toplam": [kalan["Toplam"].sum()],
                                  "Hatalı": [kalan["Hatalı"].sum()]},
                                 index=[f"Diğer ({len(kalan)} değer)"])
            gruplar = pd.concat([ilk, diger])

        # Kategori başlığı (4 sütun birleşik)
        ws.merge_cells(start_row=1, start_column=sutun,
                       end_row=1, end_column=sutun + 3)
        h = ws.cell(row=1, column=sutun, value=kategori.upper())
        h.font = Font(bold=True, color="FFFFFF", size=11)
        h.fill = kategori_d
        h.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Tablo başlığı
        for j, ad in enumerate((L("Değer"), L("Toplam"), L("Hatalı"), L("Hata %"))):
            c = ws.cell(row=2, column=sutun + j, value=ad)
            c.font = baslik_f
            c.fill = baslik_d

        # Satırlar
        r = 3
        for deger, s in gruplar.iterrows():
            oran = 100 * s["Hatalı"] / s["Toplam"] if s["Toplam"] else 0
            hucre = [deger, int(s["Toplam"]), int(s["Hatalı"]), round(oran, 1)]
            for j, v in enumerate(hucre):
                c = ws.cell(row=r, column=sutun + j, value=v)
                c.border = kenar
                if r % 2 == 1:
                    c.fill = zebra
                if j == 3 and oran >= 60:
                    c.font = kirmizi
            r += 1

        # Toplam satırı
        top_t = int(gruplar["Toplam"].sum())
        top_h = int(gruplar["Hatalı"].sum())
        top_o = round(100 * top_h / top_t, 1) if top_t else 0
        for j, v in enumerate(("TOPLAM", top_t, top_h, top_o)):
            c = ws.cell(row=r, column=sutun + j, value=v)
            c.font = Font(bold=True)
            c.fill = toplam_d

        # Sütun genişlikleri
        en_uzun = max([len(str(d)) for d in gruplar.index] + [10])
        ws.column_dimensions[get_column_letter(sutun)].width = min(en_uzun + 2, 30)
        for j in (1, 2, 3):
            ws.column_dimensions[get_column_letter(sutun + j)].width = 9

        kategoriler[kategori] = gruplar
        sutun += 5   # 4 sütun tablo + 1 boşluk

    ws.freeze_panes = "A3"
    if ayri_dosya:
        wb.save(yol)
    return kategoriler


# =====================================================================
# BİÇİMLİ BULGU EXCEL'İ
# =====================================================================

def bicimli_bulgu_excel(pbi: pd.DataFrame, yol: str):
    """Bulguları filtreli, renk kodlu, okunabilir bir Excel tablosuna yazar."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Findings" if DIL == "EN" else "Bulgular"

    if pbi.empty:
        ws["A1"] = "Bulgu yok"
        wb.save(yol)
        return

    df = pbi.sort_values(["Kural", "Malzeme"]).reset_index(drop=True)
    df = ceviri_df(df)
    kural_basligi = L("Kural")

    lacivert = "1B2A3A"
    kural_renk = {"K1": "E8EFF5", "K2": "FBF1DE", "K3": "FBEAEA", "K4": "EEE9F7"}
    zebra = "F5F8FA"
    ince = Side(style="thin", color="DFE6EA")
    kenar = Border(bottom=ince)

    # Başlık satırı
    for c, ad in enumerate(df.columns, 1):
        h = ws.cell(row=1, column=c, value=ad)
        h.font = Font(bold=True, color="FFFFFF", size=10)
        h.fill = PatternFill("solid", fgColor=lacivert)
        h.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 26

    # Veri satırları
    for r, (_, satir) in enumerate(df.iterrows(), start=2):
        kural_kodu = str(satir[kural_basligi])[:2]
        for c, ad in enumerate(df.columns, 1):
            deger = satir[ad]
            h = ws.cell(row=r, column=c,
                        value="" if pd.isna(deger) else deger)
            h.border = kenar
            h.alignment = Alignment(vertical="top", wrap_text=(ad in
                ("Malzeme Kısa Metni", "Material Description",
                 "İlgili Alan", "Related Field",
                 "Hata Nedeni", "Mevcut Değer", "Olması Gereken",
                 "Yapılacak Düzeltme", "Sorun", "Önerilen Düzeltme")))
            if ad == kural_basligi:
                h.fill = PatternFill("solid",
                                     fgColor=kural_renk.get(kural_kodu, "FFFFFF"))
                h.font = Font(bold=True, size=10)
            elif r % 2 == 0:
                h.fill = PatternFill("solid", fgColor=zebra)

    # Sütun genişlikleri (içeriğe göre, üst sınırlı)
    for c, ad in enumerate(df.columns, 1):
        icerik = df[ad].astype(str).str.len().quantile(0.9)
        ws.column_dimensions[get_column_letter(c)].width = \
            min(max(len(ad) + 2, float(icerik) + 2), 45)

    ws.freeze_panes = "A2"                       # başlık sabit
    ws.auto_filter.ref = ws.dimensions           # filtre okları

    # Özet sayfası
    oz = wb.create_sheet("Summary" if DIL == "EN" else "Özet")
    ozet_df = (df.groupby(kural_basligi).agg(
        Bulgu=(L("Malzeme"), "size"),
        Malzeme_Sayısı=(L("Malzeme"), "nunique")).reset_index())
    oz.append([L("Kural"), L("Bulgu"), L("Malzeme Sayısı")])
    for h in oz[1]:
        h.font = Font(bold=True, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor=lacivert)
    for _, s in ozet_df.iterrows():
        oz.append([s[kural_basligi], int(s["Bulgu"]), int(s["Malzeme_Sayısı"])])
    for col, w in zip("ABC", (34, 10, 16)):
        oz.column_dimensions[col].width = w

    wb.save(yol)


# =====================================================================
# DASHBOARD ÜRETİMİ
# =====================================================================

def dashboard_uret(ozet, durum, k2, pbi, basit, dosya, istisna_n=0,
                   kapsam_disi=0, gecmis=None):
    import datetime, json as _json
    toplam = len(durum)
    dogru = int((durum["Genel Durum"] == "DOĞRU").sum())
    hatali = toplam - dogru
    yuzde = (100 * dogru / toplam) if toplam else 0.0
    tarih = datetime.date.today().strftime("%d.%m.%Y")
    urun_n = int((durum["Kaynak"] == "Ürün Kodu").sum())
    ust_n = toplam - urun_n
    toplam_bulgu = int(ozet["Bulgu Sayısı"].sum())

    skor_cevre = 2 * 3.14159 * 62
    skor_yay = skor_cevre * (yuzde / 100)

    def _ip(kod):
        return (KURAL_ACIKLAMA.get(kod, "")
                .replace('"', "&quot;").replace("\n", " "))

    # ---------------------------------------------------------------
    # KURUMSAL KİMLİK — Evyap logosu ve satır içi SVG ikon seti
    # (yalnızca sunum katmanı; hiçbir hesaplamayı etkilemez)
    # ---------------------------------------------------------------
    import base64 as _b64, os as _os

    def _evyap_logo_goml():
        """evyap_logo.svg / evyap_logo.png dosyasını panele gömer.

        Logo bulunamazsa yalnızca yazı tabanlı bir marka kilidi kullanılır;
        logo hiçbir koşulda yeniden çizilmez, gerilmez veya renklendirilmez.
        """
        adaylar = []
        try:
            _kok = _os.path.dirname(_os.path.abspath(__file__))
        except Exception:
            _kok = _os.getcwd()
        for _d in (_os.getcwd(), _kok, _os.path.dirname(_os.path.abspath(dosya)) or "."):
            for _ad in ("evyap_logo.svg", "evyap_logo.png",
                        "evyap_logo.jpg", "evyap_logo.jpeg"):
                adaylar.append(_os.path.join(_d, _ad))
        for _y in adaylar:
            try:
                if not _os.path.isfile(_y):
                    continue
                with open(_y, "rb") as _f:
                    _ham = _f.read()
                _uzanti = _os.path.splitext(_y)[1].lower()
                _tip = {".svg": "image/svg+xml", ".png": "image/png",
                        ".jpg": "image/jpeg", ".jpeg": "image/jpeg"}[_uzanti]
                _b64s = _b64.b64encode(_ham).decode("ascii")
                return (f'<img src="data:{_tip};base64,{_b64s}" alt="Evyap" '
                        f'decoding="async">')
            except Exception:
                continue
        return ""

    _LOGO = _evyap_logo_goml()
    _LOGO_VAR = bool(_LOGO)
    _LOGO_YEDEK = '<span class="kilit-yedek">EVYAP</span>'

    def marka_kilidi(koyu=False, alt=True):
        """Evyap logosu + ürün adı + birim bilgisinden oluşan kompakt kilit."""
        ic = _LOGO if _LOGO_VAR else _LOGO_YEDEK
        # Koyu zeminde beyaz logo varyantı yoksa logo hafif bir marka plakasına oturur.
        sinif = "kilit-logo kilit-plaka" if (koyu and _LOGO_VAR) else "kilit-logo"
        metin = ('<span class="kilit-metin"><b>Malzeme Veri Kalitesi</b>'
                 '<small>Bilgi Teknolojileri · SAP MM</small></span>') if alt else \
                ('<span class="kilit-metin"><b>Malzeme Veri Kalitesi</b></span>')
        return (f'<div class="kilit"><span class="{sinif}">{ic}</span>'
                f'<span class="kilit-ayrac"></span>{metin}</div>')

    _IKON = {
        "ana": "<path d='M3 10.6 12 3.2l9 7.4'/><path d='M5.4 9.4V20.4h13.2V9.4'/>"
               "<path d='M9.8 20.4v-6h4.4v6'/>",
        "yukle": "<path d='M12 16V4'/><path d='m7.5 8.5 4.5-4.5 4.5 4.5'/>"
                 "<path d='M4 15.5v3A2.5 2.5 0 0 0 6.5 21h11a2.5 2.5 0 0 0 2.5-2.5v-3'/>",
        "panel": "<rect x='3' y='3' width='7.5' height='8.5' rx='1.5'/>"
                 "<rect x='13.5' y='3' width='7.5' height='5' rx='1.5'/>"
                 "<rect x='13.5' y='10.5' width='7.5' height='10.5' rx='1.5'/>"
                 "<rect x='3' y='14' width='7.5' height='7' rx='1.5'/>",
        "analiz": "<path d='M4 20V4'/><path d='M4 20h16'/><rect x='7.5' y='12' width='3' height='5'/>"
                  "<rect x='12.5' y='8.5' width='3' height='8.5'/>"
                  "<rect x='17.5' y='5.5' width='3' height='11.5'/>",
        "liste": "<rect x='3' y='4' width='18' height='16' rx='2'/><path d='M3 9h18'/>"
                 "<path d='M9 9v11'/><path d='M3 14.5h18'/>",
        "tarihce": "<circle cx='12' cy='12' r='8.5'/><path d='M12 7.2V12l3.2 2'/>",
        "ok": "<path d='m9 5 7 7-7 7'/>",
        "filtre": "<path d='M4 6h16'/><path d='M7 12h10'/><path d='M10 18h4'/>",
        "takvim": "<rect x='3.5' y='5' width='17' height='16' rx='2'/><path d='M3.5 10h17'/>"
                  "<path d='M8 3v4'/><path d='M16 3v4'/>",
        "saglik": "<path d='M3 12h4l2.5-6 4 12 2.5-6H21'/>",
        "bulgu": "<path d='M12 4.5 21 20H3Z'/><path d='M12 10v4.5'/><path d='M12 17.4h.01'/>",
        "malzeme": "<path d='m12 3 8.5 4.6v8.8L12 21l-8.5-4.6V7.6Z'/>"
                   "<path d='m3.5 7.6 8.5 4.7 8.5-4.7'/><path d='M12 12.3V21'/>",
        "tarama": "<circle cx='11' cy='11' r='7'/><path d='m20 20-3.6-3.6'/>",
        "onay": "<circle cx='12' cy='12' r='8.5'/><path d='m8.3 12.2 2.6 2.6 4.8-5.2'/>",
        "kalkan": "<path d='M12 3.2 19.5 6v6c0 4.2-3 7.4-7.5 8.8C7.5 19.4 4.5 16.2 4.5 12V6Z'/>"
                  "<path d='m9.2 12.2 2 2 3.6-4'/>",
        "dosya": "<path d='M13.5 3H7a2 2 0 0 0-2 2v14a2 2 0 0 0 2 2h10a2 2 0 0 0 2-2V8.5Z'/>"
                 "<path d='M13.5 3v5.5H19'/>",
        "katman": "<path d='m12 3 9 4.7-9 4.7-9-4.7Z'/><path d='m3 12.4 9 4.7 9-4.7'/>"
                  "<path d='m3 16.9 9 4.7 9-4.7'/>",
        "ayar": "<circle cx='12' cy='12' r='3'/><path d='M12 2.5v3'/><path d='M12 18.5v3'/>"
                "<path d='M21.5 12h-3'/><path d='M5.5 12h-3'/><path d='m18.7 5.3-2.1 2.1'/>"
                "<path d='m7.4 16.6-2.1 2.1'/><path d='m18.7 18.7-2.1-2.1'/>"
                "<path d='m7.4 7.4-2.1-2.1'/>",
        "kapat": "<path d='m6.5 6.5 11 11'/><path d='m17.5 6.5-11 11'/>",
        "indir": "<path d='M12 4v11'/><path d='m7.5 10.5 4.5 4.5 4.5-4.5'/>"
                 "<path d='M4.5 19.5h15'/>",
        "veritabani": "<ellipse cx='12' cy='6' rx='8' ry='3'/>"
                      "<path d='M4 6v12c0 1.7 3.6 3 8 3s8-1.3 8-3V6'/><path d='M4 12c0 1.7 3.6 3 8 3s8-1.3 8-3'/>",
        "bilgi": "<circle cx='12' cy='12' r='8.5'/><path d='M12 11.2v5'/><path d='M12 8h.01'/>",
        "kitap": "<path d='M4 5.5A2.5 2.5 0 0 1 6.5 3H20v15H6.5A2.5 2.5 0 0 0 4 20.5Z'/>"
                 "<path d='M4 20.5A2.5 2.5 0 0 1 6.5 18H20v3H6.5'/>",
    }

    def ik(ad, boyut=16, sinif="ikn"):
        """Satır içi SVG ikon — dış ikon kütüphanesi kullanılmaz."""
        return (f'<svg class="{sinif}" width="{boyut}" height="{boyut}" viewBox="0 0 24 24" '
                f'fill="none" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" '
                f'stroke-linejoin="round" aria-hidden="true" focusable="false">'
                f'{_IKON.get(ad, "")}</svg>')

    RENK = {"K1": "#3599B8", "K2": "#F2C80F", "K3": "#FD625E",
            "K4": "#A66999", "K5": "#118DFF", "K6": "#0E8A7E"}
    TEAL = "#01B8AA"; KOYU = "#374649"; MERCAN = "#FD625E"
    n_kural = len(ozet)
    b = list(ozet["Bulgu Sayısı"])
    m_say = list(ozet["Etkilenen Malzeme Sayısı"])
    kural_adlari = ["Akıllı kod eşleşmesi", "Tanım ↔ ek veri alanı",
                    "Yasaklı kelimeler", "Çokluk ↔ MARM sayacı",
                    "Üst kod benzersizliği", "Kod sonu ↔ Menşei"]

    en_b = max(b) or 1
    def _mini(i):
        acik = (f'<div class="minisut ipuc" data-kural="K{i+1}" '
                f'data-ipucu="{_ip(f"K{i+1}")}">')
        return (acik
                + f'<span class="msdeger">{b[i]}</span>'
                + f'<div class="msbar" style="height:{max(6, 100*b[i]/en_b):.0f}px;'
                + f'background:{RENK[f"K{i+1}"]}"></div>'
                + f'<span class="msad">K{i+1}</span></div>')

    kural_bar = "".join(_mini(i) for i in range(n_kural))

    # --- Veri kalitesi boyutları (DAMA/DQ çerçevesi): her boyut ayrı puanlanır ---
    boyutlar = boyut_puanlari(basit, durum, toplam)

    def _boyut_renk(p):
        return TEAL if p >= 95 else ("#E3A63B" if p >= 85 else MERCAN)

    boyut_html = "".join(
        f'<div class="boyut"><div class="bsatir">'
        f'<span class="bad">{ad_b}</span>'
        f'<span class="bpuan" style="color:{_boyut_renk(p)}">{p:.1f}</span></div>'
        f'<div class="bray"><div style="width:{p:.1f}%;background:{_boyut_renk(p)}"></div></div>'
        f'<div class="baciklama">{aciklama} · {n:,} malzeme etkilendi</div></div>'
        for ad_b, aciklama, p, n in boyutlar)

    # --- En sorunlu alanlar (aksiyon önceliği) ---
    alan_html = ""
    if not basit.empty:
        alan_say = (basit[basit["Kural"] != "K4"]["İlgili Alan"]
                    .value_counts().head(6))
        en_a = int(alan_say.max()) if len(alan_say) else 1
        alan_html = "".join(
            f'<div class="alanb"><div class="asatir"><span>{a}</span>'
            f'<span class="mono">{int(v):,}</span></div>'
            f'<div class="aray"><div style="width:{100*v/en_a:.0f}%"></div></div></div>'
            for a, v in alan_say.items())

    kural_lejant = "".join(
        f'<span class="lj"><i style="background:{RENK[f"K{i+1}"]}"></i>K{i+1} <b>{b[i]:,}</b></span>'
        for i in range(n_kural))
    cevre = 2 * 3.14159 * 44
    pay = cevre * (dogru / toplam) if toplam else 0

    # Marka x Kural yığılmış sütunlar
    marka_map = dict(zip(durum["Malzeme"].astype(str),
                         durum["Tanım"].astype(str).str.split().str[0]))
    _bos_kutu = ('<div class="bos" style="padding:26px 8px">Veri yüklendiğinde '
                 'bu grafik otomatik doldurulur.</div>')
    marka_html = (f'<div class="tile s7 ana-kart" id="c_marka"><div class="tbaslik">Bulgu Sayısı'
                  f'<small>Markaya ve kurala göre</small></div>{_bos_kutu}</div>')
    if not pbi.empty and any(marka_map.values()):
        p2 = pbi.copy()
        p2["_marka"] = p2["Malzeme"].astype(str).map(marka_map).fillna("")
        p2 = p2[p2["_marka"] != ""]
        if not p2.empty:
            p2["_k"] = p2["Kural"].astype(str).str[:2]
            piv = p2.groupby(["_marka", "_k"]).size().unstack(fill_value=0)
            piv["_t"] = piv.sum(axis=1)
            piv = piv.sort_values("_t", ascending=False).head(6)
            en_t = piv["_t"].max() or 1
            sutunlar = ""
            for marka, satir in piv.iterrows():
                parcalar = "".join(
                    f'<div style="height:{160*satir.get(f"K{i+1}",0)/en_t:.1f}px;'
                    f'background:{RENK[f"K{i+1}"]}"></div>'
                    for i in range(n_kural) if satir.get(f"K{i+1}", 0) > 0)
                sutunlar += (f'<div class="yigin"><span class="ydeger">{int(satir["_t"])}</span>'
                             f'<div class="ycubuk">{parcalar}</div>'
                             f'<span class="yad">{marka}</span></div>')
            lejant = "".join(
                f'<span class="lj"><i style="background:{RENK[f"K{i+1}"]}"></i>K{i+1}</span>'
                for i in range(n_kural))
            ort = piv["_t"].mean()
            ort_alt = 18 + 160 * ort / en_t
            marka_html = (f'<div class="tile s7 ana-kart" id="c_marka"><div class="tbaslik">Bulgu Sayısı'
                          f'<small>Markaya ve kurala göre · kesikli çizgi = ortalama</small></div>'
                          f'<div class="yiginlar ort-sarici">'
                          f'<div class="ort-cizgi" style="bottom:{ort_alt:.0f}px"></div>'
                          f'<div class="ort-etiket" style="bottom:{ort_alt+2:.0f}px">Ort: {ort:.0f}</div>'
                          f'{sutunlar}</div>'
                          f'<div class="lejantlar">{lejant}</div></div>')

    tur_html = (f'<div class="tile s5" id="c_tur"><div class="tbaslik">Bulgu Dağılımı'
                f'<small>Kural × malzeme türü</small></div>{_bos_kutu}</div>')
    if not pbi.empty and "Malzeme türü" in pbi.columns:
        turler = pbi["Malzeme türü"].astype(str).value_counts().head(2).index.tolist()
        bloklar = ""
        for i in range(n_kural):
            alt = pbi[pbi["Kural"].astype(str).str.startswith(f"K{i+1}")]
            n = len(alt)
            if n == 0:
                continue
            t1 = int((alt["Malzeme türü"].astype(str) == turler[0]).sum()) if turler else 0
            p1 = 100 * t1 / n
            bloklar += (f'<div class="turb"><div class="turust"><span>K{i+1}</span>'
                        f'<span class="mono">{n}</span></div>'
                        f'<div class="turray"><div style="width:{p1:.0f}%;background:var(--koyu)"></div>'
                        f'<div style="width:{100-p1:.0f}%;background:var(--koyu2)"></div></div></div>')
        lej = ""
        if len(turler) >= 2:
            lej = (f'<div class="lejantlar"><span class="lj"><i style="background:var(--koyu)"></i>'
                   f'Tür {turler[0]}</span><span class="lj"><i style="background:var(--koyu2)"></i>'
                   f'Tür {turler[1]} / diğer</span></div>')
        tur_html = (f'<div class="tile s5" id="c_tur"><div class="tbaslik">Bulgu Dağılımı'
                    f'<small>Kural × malzeme türü</small></div>{bloklar}{lej}</div>')

    etki_html = (f'<div class="tile s12" id="c_etki"><div class="tbaslik">Düzeltme Etkisi'
                 f'<small>İlgili kural çözülürse ulaşılacak veri sağlığı</small></div>{_bos_kutu}</div>')
    kural_kolonlari = [c for c in durum.columns
                       if c.startswith("K") and " " in c
                       and "bilgi" not in c][:n_kural]
    if toplam and kural_kolonlari:
        satirlar = ""
        for i, kc in enumerate(kural_kolonlari):
            digerleri = [d for d in kural_kolonlari if d != kc]
            sadece = int(((durum[kc] == "HATALI") &
                          (durum[digerleri] != "HATALI").all(axis=1)).sum())
            yeni = 100 * (dogru + sadece) / toplam
            satirlar += (f'<div class="etkib"><span class="ead">K{i+1}</span>'
                         f'<div class="eray"><div style="width:{yeni:.0f}%;'
                         f'background:{TEAL}"></div></div>'
                         f'<span class="mono">%{yeni:.1f}</span></div>')
        satirlar += (f'<div class="etkib"><span class="ead"><b>Tümü</b></span>'
                     f'<div class="eray"><div style="width:100%;background:var(--koyu)"></div></div>'
                     f'<span class="mono"><b>%100</b></span></div>')
        etki_html = (f'<div class="tile s12" id="c_etki"><div class="tbaslik">Düzeltme Etkisi'
                     f'<small>İlgili kural çözülürse ulaşılacak veri sağlığı · şu an %{yuzde:.1f}</small></div>'
                     f'{satirlar}</div>')

    rehber_kisa = [
        "Ürün, üst koduyla akıllı kod alanlarında birebir karşılaştırılır; boş/bulunamayan üst kodlar raporlanır",
        "Tanım sonundaki kod (LF, KZ…) ile alan iki yönde tutarlı olmalı; program düzeltilmiş tanımı önerir",
        "NEW, YENİ ve bağımsız 6/7/2026/2027 tanımda bulunamaz; KFR*6 gibi çarpanlar hata sayılmaz",
        "Tanım sonundaki çokluk (x24), MARM'daki ADET birimi sayacıyla aynı olmalı — ton/KL'ye bakılmaz",
        "İki üst kod aynı marka/alt marka/varyant/boyut/S&OP ve koli içi adete sahip olamaz",
        "Ürün kodunun son iki hanesi Menşei alanındaki kodla aynı olmalı (10=TR, 21=MY VAS…); boş menşei de bulgudur",
    ]
    rehber = "".join(
        f'<div class="tile s24 rehberk ipuc" data-ipucu="{_ip(f"K{i+1}")}">'
        f'<span class="hap" style="background:{RENK[f"K{i+1}"]}22;'
        f'color:{RENK[f"K{i+1}"]}">K{i+1}</span>'
        f'<b>{kural_adlari[i]}</b><p>{rehber_kisa[i]}</p></div>'
        for i in range(n_kural))

    # Düzeltme listesi (8 sütunlu basit tablo) — sayfa 3 verisi
    kaynak_map = dict(zip(durum["Malzeme"].astype(str), durum["Kaynak"]))
    kayitlar = []
    if not basit.empty:
        for _, s in basit.head(5000).iterrows():
            kayitlar.append({
                "t": str(s.get("Malzeme Kısa Metni", "") or ""),
                "kay": kaynak_map.get(str(s["Malzeme"]), "Ürün Kodu"),
                "alt": int(s.get("Alt Tarif Sayısı", 1) or 1),
                "m": str(s["Malzeme"]), "u": str(s.get("Üst Kod", "") or ""),
                "k": str(s["Kural"]), "n": str(s["Hata Nedeni"]),
                "a": str(s["İlgili Alan"]), "mv": str(s.get("Mevcut Değer", "") or ""),
                "og": str(s.get("Olması Gereken", "") or ""),
                "d": str(s.get("Yapılacak Düzeltme", "") or "")})
    veri_json = _json.dumps(kayitlar, ensure_ascii=False)

    # Kural özet tablosu: satır içi mini donut (bulgu payı)
    sat = ""
    for i in range(n_kural):
        pay_k = 100 * b[i] / toplam_bulgu if toplam_bulgu else 0
        cv = 2 * 3.14159 * 8
        sat += (f'<tr><td><span class="khap k{i+1}" style="background:{RENK[f"K{i+1}"]}">K{i+1}</span> '
                f'{kural_adlari[i]}</td>'
                f'<td class="say">{b[i]:,}</td><td class="say">{m_say[i]:,}</td>'
                f'<td class="say"><svg class="minidonut" width="20" height="20" viewBox="0 0 20 20">'
                f'<circle cx="10" cy="10" r="8" fill="none" class="iz" stroke-width="4"/>'
                f'<circle cx="10" cy="10" r="8" fill="none" stroke="{RENK[f"K{i+1}"]}" stroke-width="4" '
                f'stroke-dasharray="{cv*pay_k/100:.1f} {cv:.1f}" stroke-dashoffset="{cv/4:.1f}"/></svg>'
                f' %{pay_k:.0f}</td></tr>')
    ozet_tablo = ('<table class="ozet-tablo"><thead><tr><th>Kural</th><th style="text-align:right">Bulgu</th>'
                  '<th style="text-align:right">Malzeme</th><th style="text-align:right">Pay</th></tr></thead>'
                  f'<tbody>{sat}</tbody></table>')

    cfg = {
        "kiyas": KIYAS_ALANLARI, "bos": BOS_HATA_SAYILMAZ,
        "bayrak": BAYRAK_ALANLARI,
        "k2": KURAL2_KATEGORILER, "yasakli": YASAKLI_KELIMELER,
        "yasakliSonda": YASAKLI_SADECE_SONDA,
        "bitisik": K2_BITISIK_KODLAR,
        "olcuBirimi": [tr_upper(x) for x in OLCU_BIRIMI_KODLARI],
        "urunTur": sorted(GECERLI_URUN_TURLERI),
        "ustTur": sorted(GECERLI_UST_TURLERI),
        "tersAlan": TERS_KONTROL_ALANLARI, "ters": TERS_KONTROL,
        "sonKategori": K2_TANIM_SONU_KATEGORILER,
        "haricBirim": [tr_upper(b) for b in MARM_HARIC_BIRIMLER],
        "silmeAnahtar": K5_SILME_ANAHTARLARI,
        "ustkodsuzOnek": USTKODSUZ_ONEKLER,
        "ustKodOneki": tr_upper(UST_KOD_ONEKI),
        "ustYoksaRcMuaf": UST_KOD_YOKSA_RC_MUAF,
        "k4MuafHiyerarsiler": sorted(K4_MUAF_HIYERARSILER),
        "k6Aktif": K6_AKTIF,
        "k6GecersizBildir": K6_GECERSIZ_SON_KOD_BILDIR,
        "hiyerarsiAdaylari": HIYERARSI_SUTUN_ADAYLARI,
        "tanimSonuYasak": [tr_upper(x) for x in TANIM_SONU_YASAK],
        "k5Alanlar": K5_PARMAK_IZI_ALANLARI,
        "aciklama": KURAL_ACIKLAMA,
        "en2tr": EN2TR_KOLON,
        "kanonik": KANONIK_KOLONLAR,
        "renk": RENK,
    }
    cfg_json = _json.dumps(cfg, ensure_ascii=False)

    # --- Çalıştırma geçmişi (Python koşumları) ---
    gecmis_kayit = []
    if gecmis is not None and not gecmis.empty:
        for _, g in gecmis.iterrows():
            kayit = {"t": f"{g.get('Tarih', '')} {g.get('Saat', '')}".strip(),
                     "kaynak": "python",
                     "saglik": float(g.get("Veri Sağlığı %", 0) or 0),
                     "toplam": int(g.get("Toplam Malzeme", 0) or 0),
                     "dogru": int(g.get("Doğru", 0) or 0),
                     "hatali": int(g.get("Hatalı", 0) or 0),
                     "sayilar": {}, "alanlar": {}}
            for k in ("K1", "K2", "K3", "K4", "K5"):
                sut = f"{k} Bulgu"
                if sut in g.index:
                    kayit["sayilar"][k] = int(g.get(sut, 0) or 0)
            gecmis_kayit.append(kayit)
    # Bu koşumun alan dağılımı
    alan_dagilimi = {}
    if not basit.empty:
        alan_dagilimi = {str(a): int(n) for a, n in
                         basit["İlgili Alan"].value_counts().head(8).items()}
    if gecmis_kayit:
        gecmis_kayit[-1]["alanlar"] = alan_dagilimi
    gecmis_json = _json.dumps(gecmis_kayit, ensure_ascii=False)

    # K4 (alt tarifler) düzeltme listesinde satır olarak yer almaz; bilgi
    # sütunu olarak taşınır — bu yüzden filtre çipleri yalnızca K1-K3'tür.
    liste_kurallari = list(range(n_kural))
    cipler = '<button class="cip aktif" data-k="ALL">Tümü</button>' + "".join(
        f'<button class="cip ipuc" data-k="K{i+1}" '
        f'data-ipucu="{_ip(f"K{i+1}")}">K{i+1}</button>'
        for i in liste_kurallari)

    nav_kartlari = [
        ("p1", "panel", "Kontrol Paneli",
         "Genel özet: veri sağlığı, bulgu sayıları, durum halkası"),
        ("p2", "analiz", "Analiz Grafikleri",
         "Marka × kural dağılımı, tür kırılımı ve düzeltme etkisi"),
        ("p4", "tarihce", "Veri Sağlığı Tarihçesi",
         "Önceki çalıştırmalarla karşılaştırma — hangi alanlar iyileşti, "
         "hangileri kötüleşti"),
        ("p3", "liste", "Düzeltme Listesi",
         f"{len(kayitlar):,} bulgu — neyi, neyle, nasıl düzelteceğinizi satır satır gösterir"),
    ]
    giris_nav = "".join(
        f'<button class="navk" data-git="{pid}">'
        f'<span class="navno">{ik(ikon, 17)}</span>'
        f'<span class="navmetin">'
        f'<b>{ad}</b><small>{acik}</small></span>'
        f'{ik("ok", 16, "ikn ok")}</button>'
        for pid, ikon, ad, acik in nav_kartlari)

    stil = """
/* =====================================================================
   EVYAP · MALZEME VERİ KALİTESİ — KURUMSAL TASARIM SİSTEMİ
   Bilgi Teknolojileri · SAP MM
   ===================================================================== */
:root{
/* marka */
--marka:#0033A1;--marka2:#0A47C2;--marka-koyu:#00246E;
--marka-yumusak:#EAF0FB;--marka-cizgi:#C5D5F1;--marka-ust:#0B1B33;
/* durum */
--teal:#01B8AA;--teal-koyu:#0E8A7E;--teal-yumusak:#E3F7F4;
--mercan:#FD625E;--mercan-koyu:#C0392B;--mercan-yumusak:#FDECEB;
--amber:#E3A63B;--amber-yumusak:#FBF2DF;--koyu:#374649;--koyu2:#9DB2B6;
/* yüzey */
--bg:#F1F4F8;--tile:#FFFFFF;--tile-ust:#FFFFFF;--tile-duz:#F8FAFC;
--iz:#E4EAF1;--satir:#F6F8FB;--satir-hover:#EDF4FD;
--cizgi:#E1E8F0;--cizgi2:#EDF1F6;
/* metin */
--ink:#12233B;--ink2:#3E5169;--gri:#68798F;--silik:#8494A7;
/* gölge + ölçü */
--golge:0 1px 1px rgba(16,35,59,.04),0 2px 6px rgba(16,35,59,.05);
--golge2:0 1px 2px rgba(16,35,59,.05),0 8px 22px rgba(16,35,59,.08);
--r:12px;--r2:14px;--en:1680px;--ustyuk:146px;
--f:"Segoe UI Variable Text","Segoe UI Variable","Segoe UI",system-ui,-apple-system,"Helvetica Neue",Arial,sans-serif;
--fb:"Segoe UI Variable Display","Segoe UI Variable","Segoe UI",system-ui,sans-serif;
--mono:"Cascadia Mono",Consolas,"Courier New",monospace}

body.gece{
--marka:#5C8BE8;--marka2:#7BA4F0;--marka-koyu:#3D6FD1;
--marka-yumusak:#1B2740;--marka-cizgi:#31456A;--marka-ust:#0B1220;
--teal-yumusak:#123230;--mercan-yumusak:#32201F;--amber-yumusak:#2D2617;
--koyu:#6E8B92;--koyu2:#3E525A;
--bg:#12171F;--tile:#1A222C;--tile-ust:#202A36;--tile-duz:#161D26;
--iz:#2A3542;--satir:#1E2733;--satir-hover:#243244;
--cizgi:#2B3644;--cizgi2:#232D39;
--ink:#E8EEF5;--ink2:#B4C2D1;--gri:#94A4B5;--silik:#7E8EA0;
--golge:0 1px 2px rgba(0,0,0,.30);--golge2:0 6px 20px rgba(0,0,0,.38)}

*{margin:0;padding:0;box-sizing:border-box}
html{max-width:100%;overflow-x:hidden}
body{max-width:100%}
body{background:var(--bg);color:var(--ink);font-family:var(--f);font-size:13px;
line-height:1.5;-webkit-font-smoothing:antialiased;transition:background .25s,color .25s}
button,input,select{font-family:inherit}
.mono{font-family:var(--mono);font-variant-numeric:tabular-nums}
.say,.kpi b,.bpuan,.gsaglik,.gfark,.gbulgu,.gtarih,.ydeger,.msdeger,.g-mkart b{
font-variant-numeric:tabular-nums;font-feature-settings:"tnum" 1}
.ikn{flex:none;display:block}
:focus-visible{outline:2px solid var(--marka);outline-offset:2px;border-radius:6px}

/* ---------- SAYFA KABUĞU ---------- */
.sayfa{display:none;padding:0 0 52px}
.sayfa.acik{display:block;animation:belir .26s ease}
@keyframes belir{from{opacity:0;transform:translateY(6px)}to{opacity:1;transform:none}}
.kabuk{max-width:var(--en);margin:0 auto;padding:20px 24px 0}

/* =====================================================================
   MARKA KİLİDİ (logo + ürün adı)
   ===================================================================== */
.kilit{display:flex;align-items:center;gap:12px;min-width:0}
.kilit-logo{flex:none;display:flex;align-items:center;justify-content:center;height:32px}
.kilit-logo img,.kilit-logo svg{height:28px;width:auto;max-width:132px;display:block}
.kilit-plaka{background:#fff;border-radius:8px;padding:6px 10px;height:auto;
box-shadow:0 1px 3px rgba(0,0,0,.18)}
.kilit-ayrac{flex:none;width:1px;height:26px;background:var(--cizgi)}
.kilit-metin{min-width:0}
.kilit-metin b{display:block;font-family:var(--fb);font-size:15px;font-weight:600;
letter-spacing:-.2px;color:var(--ink);line-height:1.2;white-space:nowrap}
.kilit-metin small{display:block;font-size:11px;color:var(--gri);margin-top:2px;
letter-spacing:.2px;white-space:nowrap}
.kilit-yedek{font-family:var(--fb);font-size:19px;font-weight:700;letter-spacing:1.5px;
color:var(--marka);line-height:1;padding:2px 0}

/* =====================================================================
   ANA EKRAN (giriş) — kurumsal veri kalitesi portalı
   ===================================================================== */
#giris{display:none;min-height:100vh;padding:0;position:relative;overflow:hidden;
background:linear-gradient(160deg,#0E2144 0%,#0B1B33 46%,#081426 100%)}
#giris.acik{display:block}
.g-arka{position:absolute;inset:0;overflow:hidden;pointer-events:none}
.g-izgara{position:absolute;inset:0;
background-image:linear-gradient(rgba(255,255,255,.045) 1px,transparent 1px),
linear-gradient(90deg,rgba(255,255,255,.045) 1px,transparent 1px);
background-size:56px 56px;
-webkit-mask-image:radial-gradient(120% 90% at 78% 10%,#000 0%,transparent 72%);
mask-image:radial-gradient(120% 90% at 78% 10%,#000 0%,transparent 72%)}
.g-kayit{position:absolute;inset:0;opacity:.5}
.g-parlak{position:absolute;top:-220px;right:-160px;width:620px;height:620px;
background:radial-gradient(circle,rgba(1,184,170,.18),transparent 66%)}
.g-parlak.b{top:auto;bottom:-260px;left:-180px;right:auto;width:560px;height:560px;
background:radial-gradient(circle,rgba(10,71,194,.26),transparent 66%)}
.g-ic{position:relative;max-width:1240px;margin:0 auto;padding:20px 5vw 26px;
min-height:100vh;display:flex;flex-direction:column}

.g-ust{display:flex;justify-content:space-between;align-items:center;gap:14px;
padding-bottom:15px;border-bottom:1px solid rgba(255,255,255,.10);flex-wrap:wrap}
.g-ust .kilit-metin b{color:#fff;font-size:15.5px}
.g-ust .kilit-metin small{color:#8FA6C4;font-size:11px}
.g-ust .kilit-ayrac{background:rgba(255,255,255,.16)}
.g-ustsag{display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.g-rozet-mini{display:inline-flex;align-items:center;gap:6px;font-size:11px;font-weight:600;
letter-spacing:.4px;color:#B8CCE6;background:rgba(255,255,255,.07);
border:1px solid rgba(255,255,255,.14);border-radius:7px;padding:6px 11px;white-space:nowrap}
.g-rozet-mini .mono{font-family:var(--mono);font-size:11px;color:#DCE7F5}
.g-ustdugme{display:inline-flex;align-items:center;gap:8px;cursor:pointer;
border:1px solid rgba(255,255,255,.22);border-radius:8px;background:rgba(255,255,255,.06);
color:#EAF1FA;padding:8px 15px;font-size:12.5px;font-weight:600;
transition:background .15s,border-color .15s,transform .15s}
.g-ustdugme:hover{background:rgba(255,255,255,.13);border-color:rgba(255,255,255,.4);
transform:translateY(-1px)}

.g-kahraman{display:flex;gap:52px;align-items:center;justify-content:space-between;
flex-wrap:wrap;padding:38px 0 30px;flex:1}
.g-sol{flex:1;min-width:300px}
.g-etiket{display:inline-flex;align-items:center;gap:8px;font-size:11px;font-weight:700;
letter-spacing:2px;color:#8FD4CC;margin-bottom:16px}
.g-etiket::before{content:"";width:22px;height:2px;background:var(--teal);border-radius:2px}
.g-baslik{font-family:var(--fb);font-size:clamp(30px,4.2vw,48px);line-height:1.08;
color:#fff;font-weight:700;letter-spacing:-1px}
.g-vurgu{color:#5FD6C6}
.g-alt{margin-top:14px;max-width:540px;color:#A9BED6;font-size:14px;line-height:1.62}
.g-metrik{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:10px;
margin-top:26px;max-width:460px}
.g-mkart{background:rgba(255,255,255,.055);border:1px solid rgba(255,255,255,.12);
border-radius:10px;padding:12px 14px;display:flex;flex-direction:column;justify-content:flex-end}
.g-mkart span{display:block;font-size:11px;letter-spacing:.7px;color:#8FA6C4;
font-weight:600;text-transform:uppercase}
.g-mkart b{display:block;font-family:var(--fb);font-size:23px;color:#fff;font-weight:600;
margin-top:5px;letter-spacing:-.4px;line-height:1.1}

.g-sag{display:flex;flex-direction:column;align-items:stretch;gap:0;
background:rgba(255,255,255,.045);border:1px solid rgba(255,255,255,.11);
border-radius:16px;padding:20px 22px;min-width:268px}
.g-halka{display:flex;justify-content:center}
.g-skorbilgi{display:flex;flex-direction:column;gap:0;margin-top:14px;
border-top:1px solid rgba(255,255,255,.10);padding-top:4px}
.g-skorsatir{display:flex;align-items:center;gap:9px;color:#A9BED6;font-size:12.5px;
padding:7px 0}
.g-skorsatir b{color:#fff;font-weight:600;font-variant-numeric:tabular-nums;
min-width:64px;display:inline-block}
.g-skorsatir i{width:9px;height:9px;border-radius:2px;display:inline-block;flex:none}
.g-skorsatir i.ok{background:var(--teal)}
.g-skorsatir i.hata{background:var(--mercan)}
.g-durum{text-align:center;font-size:12px;color:#A9BED6;margin-top:10px;line-height:1.5}
.g-durum b{color:#fff}

.g-bolum-bas{display:flex;align-items:baseline;justify-content:space-between;gap:12px;
margin-bottom:12px;flex-wrap:wrap}
.g-bolum-bas h2{font-family:var(--fb);font-size:13px;font-weight:600;color:#DCE7F5;
letter-spacing:.6px;text-transform:uppercase}
.g-bolum-bas span{font-size:11.5px;color:#7C93AE}
.g-kartlar{display:grid;grid-template-columns:repeat(auto-fit,minmax(248px,1fr));gap:12px;
padding-bottom:22px}
.navk{display:flex;align-items:flex-start;gap:13px;width:100%;text-align:left;cursor:pointer;
background:rgba(255,255,255,.055);border:1px solid rgba(255,255,255,.12);border-radius:12px;
padding:16px 16px;color:#fff;font-family:var(--f);
transition:transform .16s,background .16s,border-color .16s}
.navk:hover{transform:translateY(-2px);background:rgba(255,255,255,.10);
border-color:rgba(1,184,170,.48)}
.navno{flex:none;width:34px;height:34px;border-radius:9px;background:rgba(255,255,255,.09);
border:1px solid rgba(255,255,255,.14);color:#7FE3D6;display:flex;align-items:center;
justify-content:center}
.navk:hover .navno{background:rgba(1,184,170,.18);border-color:rgba(1,184,170,.45)}
.navmetin{flex:1;min-width:0}
.navmetin .navsira{display:block;font-family:var(--mono);font-size:11px;letter-spacing:1.4px;
color:#7C93AE;margin-bottom:3px}
.navmetin b{display:block;font-size:14px;color:#fff;font-weight:600;letter-spacing:-.1px}
.navmetin small{display:block;color:#9BB2CC;font-size:11.5px;margin-top:5px;line-height:1.5}
.ok{flex:none;color:#6C86A4;margin-top:3px;transition:transform .16s,color .16s}
.navk:hover .ok{color:var(--teal);transform:translateX(2px)}
.g-dip{margin-top:auto;padding-top:14px;border-top:1px solid rgba(255,255,255,.07);
color:#6C86A4;font-size:11px;font-family:var(--mono);display:flex;gap:10px;flex-wrap:wrap}

/* =====================================================================
   İÇ SAYFA BAŞLIĞI — iki katmanlı kurumsal header
   ===================================================================== */
.ustbar{position:sticky;top:0;z-index:30;background:var(--tile);
border-bottom:1px solid var(--cizgi);box-shadow:0 1px 0 rgba(16,35,59,.03)}
.ustbar::before{content:"";display:block;height:3px;
background:linear-gradient(90deg,var(--marka) 0%,var(--marka2) 52%,var(--teal) 100%)}

/* ---- Sayfa başlık bloğu ---- */
.sayfa-bas{display:flex;align-items:flex-end;justify-content:space-between;gap:18px;
flex-wrap:wrap;margin-bottom:18px}
.sayfa-bas .goz{display:block;font-size:11px;font-weight:700;letter-spacing:1.3px;
color:var(--marka);text-transform:uppercase;margin-bottom:5px}
.sayfa-bas h1{font-family:var(--fb);font-size:21px;font-weight:600;letter-spacing:-.5px;
color:var(--ink);line-height:1.2}
.sayfa-bas p{font-size:12.5px;color:var(--gri);margin-top:5px;max-width:62ch;line-height:1.5}
.sayfa-bas-yan{display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.sayfa-bas-yan .ozet-kutu{display:flex;flex-direction:column;gap:2px;padding:8px 14px;
background:var(--tile);border:1px solid var(--cizgi);border-radius:10px;min-width:96px}
.sayfa-bas-yan .ozet-kutu span{font-size:11px;color:var(--gri);font-weight:600}
.sayfa-bas-yan .ozet-kutu b{font-family:var(--fb);font-size:17px;font-weight:600;color:var(--ink);
letter-spacing:-.4px;font-variant-numeric:tabular-nums}
.sayfa-bas-yan .ozet-kutu.iyi b{color:var(--teal-koyu)}
body.gece .sayfa-bas-yan .ozet-kutu.iyi b{color:#4FD8C9}
.ub-sar{max-width:var(--en);margin:0 auto;padding:0 24px}
.ub-ust{display:flex;align-items:center;gap:16px;padding:10px 0;flex-wrap:wrap}
.ub-cip{display:flex;align-items:center;gap:7px;flex-wrap:wrap;margin-left:4px}
.ub-durum{display:inline-flex;align-items:center;gap:7px;font-size:11.5px;font-weight:600;
color:var(--ink2);background:var(--tile-duz);border:1px solid var(--cizgi);
border-radius:7px;padding:5px 10px;white-space:nowrap}
.ub-durum .mono{font-family:var(--mono);color:var(--ink)}
.ub-durum .nokta{width:7px;height:7px;border-radius:50%;background:var(--teal);flex:none;
box-shadow:0 0 0 3px rgba(1,184,170,.16)}
.ub-arac{margin-left:auto;display:flex;align-items:center;gap:7px}
.ub-dugme{display:inline-flex;align-items:center;gap:7px;cursor:pointer;
border:1px solid var(--cizgi);border-radius:8px;background:var(--tile);color:var(--ink2);
padding:7px 12px;font-size:12px;font-weight:600;transition:all .15s}
.ub-dugme:hover{border-color:var(--marka-cizgi);color:var(--marka);background:var(--marka-yumusak)}
.tema::before,.yazdir::before,.arac-dugme::before{content:"";width:15px;height:15px;flex:none;
background:currentColor;
-webkit-mask-repeat:no-repeat;mask-repeat:no-repeat;-webkit-mask-position:center;
mask-position:center;-webkit-mask-size:contain;mask-size:contain}
.tema::before{-webkit-mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.8' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M20.5 14.3A8.5 8.5 0 0 1 9.7 3.5a8.5 8.5 0 1 0 10.8 10.8Z'/%3E%3C/svg%3E");
mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.8' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M20.5 14.3A8.5 8.5 0 0 1 9.7 3.5a8.5 8.5 0 1 0 10.8 10.8Z'/%3E%3C/svg%3E")}
.yazdir::before{-webkit-mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.8' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M6 9V3h12v6'/%3E%3Cpath d='M6 18H4a2 2 0 0 1-2-2v-4a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v4a2 2 0 0 1-2 2h-2'/%3E%3Crect x='6' y='14' width='12' height='7'/%3E%3C/svg%3E");
mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.8' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M6 9V3h12v6'/%3E%3Cpath d='M6 18H4a2 2 0 0 1-2-2v-4a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v4a2 2 0 0 1-2 2h-2'/%3E%3Crect x='6' y='14' width='12' height='7'/%3E%3C/svg%3E")}

/* sekme şeridi */
.sekmeler{display:flex;gap:2px;align-items:stretch;overflow-x:auto;
scrollbar-width:none;-ms-overflow-style:none;border-top:1px solid var(--cizgi2)}
.sekmeler::-webkit-scrollbar{display:none}
.sekmeler button{position:relative;display:inline-flex;align-items:center;gap:8px;
cursor:pointer;border:0;background:transparent;color:var(--gri);padding:11px 14px 10px;
font-size:12.5px;font-weight:600;white-space:nowrap;border-radius:8px 8px 0 0;
transition:color .15s,background .15s}
.sekmeler button:hover{color:var(--ink);background:var(--tile-duz)}
.sekmeler button::after{content:"";position:absolute;left:10px;right:10px;bottom:0;height:2.5px;
border-radius:2px 2px 0 0;background:transparent;transition:background .15s}
.sekmeler button.aktif{color:var(--marka)}
.sekmeler button.aktif::after{background:var(--marka)}
.sekmeler button .ikn{opacity:.85}

/* komut / arama şeridi */
.komut{position:relative;display:flex;align-items:center;padding:9px 0 11px;
border-top:1px solid var(--cizgi2)}
.komut::before{content:"";position:absolute;left:12px;top:50%;transform:translateY(-50%);
width:16px;height:16px;background:var(--silik);pointer-events:none;
-webkit-mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='2' stroke-linecap='round'%3E%3Ccircle cx='11' cy='11' r='7'/%3E%3Cpath d='m20 20-3.5-3.5'/%3E%3C/svg%3E");
mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='2' stroke-linecap='round'%3E%3Ccircle cx='11' cy='11' r='7'/%3E%3Cpath d='m20 20-3.5-3.5'/%3E%3C/svg%3E");
-webkit-mask-repeat:no-repeat;mask-repeat:no-repeat;-webkit-mask-size:contain;mask-size:contain}
.soru{width:100%;border:1px solid var(--cizgi);border-radius:9px;background:var(--tile-duz);
color:var(--ink);padding:9px 14px 9px 38px;font-size:12.5px;font-family:var(--f);
transition:border-color .15s,box-shadow .15s,background .15s}
.soru::placeholder{color:var(--silik)}
.soru:hover{border-color:var(--marka-cizgi)}
.soru:focus{outline:0;border-color:var(--marka);background:var(--tile);
box-shadow:0 0 0 3px rgba(0,51,161,.10)}
body.gece .soru:focus{box-shadow:0 0 0 3px rgba(92,139,232,.16)}

/* =====================================================================
   FİLTRE RAYI
   ===================================================================== */
.duzen{display:flex;gap:18px;align-items:flex-start}
.yan{flex:none;width:232px;background:var(--tile);border:1px solid var(--cizgi);
border-radius:var(--r2);padding:14px 14px 16px;box-shadow:var(--golge);
position:sticky;top:calc(var(--ustyuk) + 14px)}
.yan .marka-blok{display:flex;align-items:center;gap:8px;font-size:12px;font-weight:700;
letter-spacing:.8px;text-transform:uppercase;color:var(--ink2);padding-bottom:10px;
border-bottom:1px solid var(--cizgi2)}
.yan .marka-blok .ikn{color:var(--marka)}
.yan .marka-alt{display:none}
.yan .f-grup{border-top:1px solid var(--cizgi2);padding-top:11px;margin-top:11px}
.yan .f-grup:first-of-type{border-top:0;padding-top:12px;margin-top:0}
.yan label{display:block;font-size:11px;font-weight:600;color:var(--gri);letter-spacing:.4px;
margin-bottom:6px;text-transform:uppercase}
.yan select{width:100%;border:1px solid var(--cizgi);border-radius:8px;
background:var(--tile-duz);color:var(--ink);padding:9px 30px 9px 10px;font-size:12.5px;
cursor:pointer;appearance:none;-webkit-appearance:none;
background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%2368798F' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='m6 9 6 6 6-6'/%3E%3C/svg%3E");
background-repeat:no-repeat;background-position:right 9px center;background-size:14px;
transition:border-color .15s,box-shadow .15s}
.yan select:hover{border-color:var(--marka-cizgi)}
.yan select:focus{outline:0;border-color:var(--marka);box-shadow:0 0 0 3px rgba(0,51,161,.10)}
.yan .tarihk{display:flex;align-items:center;gap:8px;font-family:var(--mono);font-size:12px;
color:var(--ink);background:var(--tile-duz);border:1px solid var(--cizgi);border-radius:8px;
padding:8px 10px}
.yan .tarihk .ikn{color:var(--marka)}
.yan .bilgi-panel{background:var(--tile-duz);border:1px solid var(--cizgi2);border-radius:8px;
padding:9px 11px;font-size:11.5px;line-height:1.6;color:var(--gri)}
.yan .bilgi-panel b{color:var(--ink2);font-weight:600}
.icerik{flex:1;min-width:0}

/* =====================================================================
   KPI ŞERİDİ
   ===================================================================== */
.kpi-serit{display:grid;grid-template-columns:1.5fr repeat(4,1fr);gap:12px;margin-bottom:14px}
.kpi{position:relative;background:var(--tile);border:1px solid var(--cizgi);
border-radius:var(--r2);box-shadow:var(--golge);padding:14px 16px 15px;overflow:hidden;
display:flex;flex-direction:column;min-width:0}
.kpi::before{content:"";position:absolute;left:0;top:0;bottom:0;width:3px;
background:var(--cizgi);border-radius:0 3px 3px 0}
.kpi.ana{background:var(--tile-ust);border-color:var(--marka-cizgi);box-shadow:var(--golge2)}
.kpi.ana::before{background:var(--teal)}
.kpi.uyari::before{background:var(--mercan)}
.kpi-ust{display:flex;align-items:center;justify-content:space-between;gap:8px}
.kpi span{display:block;font-size:11.5px;font-weight:600;letter-spacing:.2px;color:var(--gri)}
.kpi .kpi-ikon{color:var(--silik);flex:none}
.kpi.ana .kpi-ikon{color:var(--teal)}
.kpi.uyari .kpi-ikon{color:var(--mercan)}
.kpi b{display:block;font-family:var(--fb);font-size:28px;font-weight:600;color:var(--ink);
line-height:1.14;margin-top:8px;letter-spacing:-1px}
.kpi.ana b{font-size:36px}
.kpi em{display:block;font-style:normal;font-size:11.5px;color:var(--gri);margin-top:6px;
line-height:1.45}
.kpi-ray{height:6px;background:var(--iz);border-radius:99px;overflow:hidden;margin-top:10px}
.kpi-ray div{height:100%;border-radius:99px;background:var(--teal);
transition:width .7s cubic-bezier(.4,0,.2,1)}

/* =====================================================================
   IZGARA + KART
   ===================================================================== */
.izg{display:grid;grid-template-columns:repeat(12,1fr);gap:14px}
.tile{background:var(--tile);padding:16px 18px 17px;border:1px solid var(--cizgi);
border-radius:var(--r2);box-shadow:var(--golge);overflow:visible;min-width:0}
.tile.ana-kart{background:var(--tile-ust);box-shadow:var(--golge2)}
.tile.duz{box-shadow:none;background:var(--tile-duz)}
.s2{grid-column:span 2}.s3{grid-column:span 3}.s4{grid-column:span 4}
.s5{grid-column:span 5}.s7{grid-column:span 7}.s12{grid-column:span 12}
.s24{grid-column:span 12}
@media(min-width:1000px){.s24{grid-column:span 4}}
.tbaslik{position:relative;display:block;font-family:var(--fb);font-size:14.5px;font-weight:600;
line-height:1.3;margin-bottom:13px;color:var(--ink);letter-spacing:-.15px;min-height:30px;
padding:1px 0 11px 40px;border-bottom:1px solid var(--cizgi2)}
/* ---- Kart başlığı ikon rozetleri (yalnızca CSS — yeniden çizimde de korunur) ---- */
.tile{--ikon:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.9' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M4 20V4'/%3E%3Cpath d='M4 20h16'/%3E%3Cpath d='m7 15 4-5 3.5 3L20 7'/%3E%3C/svg%3E");--ikon-renk:var(--marka);--ikon-zemin:var(--marka-yumusak)}
#c_boyut{--ikon:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.9' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M12 3.2 19.5 6v6c0 4.2-3 7.4-7.5 8.8C7.5 19.4 4.5 16.2 4.5 12V6Z'/%3E%3Cpath d='m9.2 12.2 2 2 3.6-4'/%3E%3C/svg%3E");--ikon-renk:var(--teal-koyu);--ikon-zemin:var(--teal-yumusak)}
.t-kural{--ikon:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.9' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M4 20V4'/%3E%3Cpath d='M4 20h16'/%3E%3Crect x='7.5' y='12' width='3' height='5'/%3E%3Crect x='12.5' y='8.5' width='3' height='8.5'/%3E%3Crect x='17.5' y='5.5' width='3' height='11.5'/%3E%3C/svg%3E");--ikon-renk:var(--marka);--ikon-zemin:var(--marka-yumusak)}
#c_donut{--ikon:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.9' stroke-linecap='round' stroke-linejoin='round'%3E%3Ccircle cx='12' cy='12' r='8.5'/%3E%3Cpath d='M12 3.5V12l6 6'/%3E%3C/svg%3E");--ikon-renk:var(--marka);--ikon-zemin:var(--marka-yumusak)}
#c_tablo{--ikon:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.9' stroke-linecap='round' stroke-linejoin='round'%3E%3Crect x='3' y='4' width='18' height='16' rx='2'/%3E%3Cpath d='M3 9h18'/%3E%3Cpath d='M9 9v11'/%3E%3Cpath d='M3 14.5h18'/%3E%3C/svg%3E");--ikon-renk:var(--marka);--ikon-zemin:var(--marka-yumusak)}
#c_alan{--ikon:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.9' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M12 4.5 21 20H3Z'/%3E%3Cpath d='M12 10v4.5'/%3E%3Cpath d='M12 17.4h.01'/%3E%3C/svg%3E");--ikon-renk:var(--mercan);--ikon-zemin:var(--mercan-yumusak)}
.t-kapsam{--ikon:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.9' stroke-linecap='round' stroke-linejoin='round'%3E%3Cellipse cx='12' cy='6' rx='8' ry='3'/%3E%3Cpath d='M4 6v12c0 1.7 3.6 3 8 3s8-1.3 8-3V6'/%3E%3Cpath d='M4 12c0 1.7 3.6 3 8 3s8-1.3 8-3'/%3E%3C/svg%3E");--ikon-renk:var(--marka);--ikon-zemin:var(--marka-yumusak)}
#c_marka{--ikon:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.9' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='m12 3 8.5 4.6v8.8L12 21l-8.5-4.6V7.6Z'/%3E%3Cpath d='m3.5 7.6 8.5 4.7 8.5-4.7'/%3E%3Cpath d='M12 12.3V21'/%3E%3C/svg%3E");--ikon-renk:var(--marka);--ikon-zemin:var(--marka-yumusak)}
#c_tur{--ikon:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.9' stroke-linecap='round' stroke-linejoin='round'%3E%3Ccircle cx='12' cy='12' r='8.5'/%3E%3Cpath d='M12 3.5V12h8.5'/%3E%3C/svg%3E");--ikon-renk:var(--marka);--ikon-zemin:var(--marka-yumusak)}
#c_etki{--ikon:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.9' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M3.5 16.5 9 11l4 3.5 7.5-7.5'/%3E%3Cpath d='M15.5 7h5v5'/%3E%3C/svg%3E");--ikon-renk:var(--teal-koyu);--ikon-zemin:var(--teal-yumusak)}
.t-liste{--ikon:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.9' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M8 6h13'/%3E%3Cpath d='M8 12h13'/%3E%3Cpath d='M8 18h13'/%3E%3Cpath d='M3.5 6h.01'/%3E%3Cpath d='M3.5 12h.01'/%3E%3Cpath d='M3.5 18h.01'/%3E%3C/svg%3E");--ikon-renk:var(--marka);--ikon-zemin:var(--marka-yumusak)}
.t-tarihce{--ikon:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.9' stroke-linecap='round' stroke-linejoin='round'%3E%3Ccircle cx='12' cy='12' r='8.5'/%3E%3Cpath d='M12 7.2V12l3.2 2'/%3E%3C/svg%3E");--ikon-renk:var(--teal-koyu);--ikon-zemin:var(--teal-yumusak)}
.tbaslik::before{content:"";position:absolute;left:0;top:-1px;width:30px;height:30px;
border-radius:9px;background:var(--ikon-zemin)}
.tbaslik::after{content:"";position:absolute;left:7px;top:6px;width:16px;height:16px;
background:var(--ikon-renk);-webkit-mask-repeat:no-repeat;mask-repeat:no-repeat;
-webkit-mask-size:contain;mask-size:contain;-webkit-mask-position:center;mask-position:center;
-webkit-mask-image:var(--ikon);mask-image:var(--ikon)}
.tbaslik small{display:block;font-weight:600;font-size:11px;letter-spacing:.3px;
color:var(--gri);margin-top:4px;text-transform:none}
.lejlist{display:flex;gap:14px;flex-wrap:wrap;align-items:center;margin-bottom:10px}
.lejlist .lj{font-size:11.5px;color:var(--gri);display:flex;align-items:center;gap:5px}
.lejlist .lj i{width:9px;height:9px;display:inline-block;border-radius:2px}
.lejlist .lj b{color:var(--ink);font-weight:600;font-variant-numeric:tabular-nums}
.lejantlar{display:flex;gap:14px;flex-wrap:wrap;margin-top:12px;padding-top:10px;
border-top:1px solid var(--cizgi2)}
.lj{font-size:11.5px;color:var(--gri);display:flex;align-items:center;gap:5px}
.lj i{width:9px;height:9px;display:inline-block;border-radius:2px}

/* boyutlar */
.boyut{margin:0;padding:10px 0;border-bottom:1px solid var(--cizgi2)}
.boyut:last-child{border-bottom:0;padding-bottom:2px}
.bsatir{display:flex;justify-content:space-between;align-items:baseline;gap:10px;margin-bottom:6px}
.bad{font-size:12.5px;font-weight:600;color:var(--ink)}
.bpuan{font-size:15px;font-weight:600;font-family:var(--mono);letter-spacing:-.3px}
.bray{height:8px;background:var(--iz);border-radius:99px;overflow:hidden}
.bray div{height:100%;border-radius:99px;transition:width .7s cubic-bezier(.4,0,.2,1)}
.baciklama{font-size:11px;color:var(--gri);margin-top:6px;line-height:1.45}

/* en sorunlu alanlar */
.alanb{margin:0;padding:9px 0;border-bottom:1px solid var(--cizgi2)}
.alanb:last-child{border-bottom:0}
.asatir{display:flex;justify-content:space-between;gap:10px;font-size:12px;margin-bottom:5px;
color:var(--ink)}
.asatir .mono{color:var(--ink);font-weight:600}
.aray{height:8px;background:var(--iz);border-radius:99px;overflow:hidden}
.aray div{height:100%;background:var(--marka);border-radius:99px;
transition:width .7s cubic-bezier(.4,0,.2,1)}
.alanb:first-child .aray div{background:var(--mercan)}

/* kapsam */
.kapsamk{display:flex;align-items:baseline;gap:8px;font-size:12px;color:var(--gri);
padding:9px 0;border-bottom:1px solid var(--cizgi2)}
.kapsamk:last-child{border-bottom:0}
.kapsamk b{color:var(--ink);font-size:15px;font-weight:600;font-family:var(--mono);
letter-spacing:-.4px;min-width:64px}

/* kural mini sütunlar */
.eksen{display:flex;justify-content:space-between;border-top:1px solid var(--cizgi2);
margin-top:8px;padding-top:6px}
.eksen span{font-size:11px;color:var(--gri);font-variant-numeric:tabular-nums}
.minisutlar{position:relative;display:flex;align-items:flex-end;gap:12px;height:150px;
padding-top:6px;background-image:repeating-linear-gradient(to top,var(--cizgi2) 0 1px,transparent 1px 25%);
background-position:0 -1px}
.minisut{flex:1;display:flex;flex-direction:column;align-items:center;justify-content:flex-end;
gap:5px;cursor:pointer;border-radius:8px 8px 0 0;transition:background .15s}
.minisut:hover{background:var(--tile-duz)}
.minisut.aktif .msbar{box-shadow:0 0 0 2px var(--tile),0 0 0 4px var(--ink)}
.minisut.pasif{cursor:default;opacity:.7}
.msbar{width:76%;max-width:38px;border-radius:4px 4px 0 0}
.msdeger{font-size:11.5px;color:var(--ink);font-weight:600}
.msad{font-size:11.5px;font-weight:600;color:var(--gri)}

/* donut */
.halka{display:flex;align-items:center;gap:16px;flex-wrap:wrap}
.halka .lejlist{flex-direction:column;align-items:flex-start;gap:8px;margin:0}

/* marka yığınları */
.yiginlar{display:flex;align-items:flex-end;gap:14px;height:206px;padding:6px 4px 0;
background-image:repeating-linear-gradient(to top,var(--cizgi2) 0 1px,transparent 1px 25%);
background-position:0 -1px}
.yigin{flex:1;display:flex;flex-direction:column;align-items:center;justify-content:flex-end;gap:5px}
.ycubuk{width:72%;max-width:46px;display:flex;flex-direction:column-reverse;
border-radius:4px 4px 0 0;overflow:hidden}
.ydeger{font-size:11.5px;color:var(--ink);font-weight:600}
.yad{font-size:11px;font-weight:600;color:var(--gri);text-align:center;
overflow:hidden;text-overflow:ellipsis;max-width:100%}
.ort-sarici{position:relative}
.ort-cizgi{position:absolute;left:0;right:0;border-top:2px dashed var(--silik);pointer-events:none}
.ort-etiket{position:absolute;right:2px;font-size:11px;color:var(--gri);background:var(--tile);
padding:0 4px;border-radius:3px}

/* tür kırılımı */
.turb{margin:0;padding:9px 0;border-bottom:1px solid var(--cizgi2)}
.turb:last-of-type{border-bottom:0}
.turust{display:flex;justify-content:space-between;font-size:12px;margin-bottom:5px;color:var(--ink)}
.turust span:first-child{font-weight:600}
.turray{display:flex;height:14px;background:var(--iz);border-radius:4px;overflow:hidden}

/* düzeltme etkisi */
.etkib{display:grid;grid-template-columns:46px 1fr 58px;align-items:center;gap:10px;
padding:8px 0;border-bottom:1px solid var(--cizgi2)}
.etkib:last-child{border-bottom:0}
.ead{font-size:12px;font-weight:600;color:var(--ink)}
.eray{background:var(--iz);height:12px;border-radius:99px;overflow:hidden}
.eray div{height:100%;border-radius:99px}
.etkib .mono{text-align:right;font-size:12px;color:var(--ink)}

/* kural rehberi */
.rehberk{cursor:help;display:flex;flex-direction:column;gap:0;
background:var(--tile-duz);box-shadow:none}
.rehberk .hap{align-self:flex-start;margin-bottom:9px}
.rehberk b{display:block;font-size:13px;margin-bottom:6px;color:var(--ink);font-weight:600}
.rehberk p{color:var(--gri);font-size:11.5px;line-height:1.55}
.hap{display:inline-flex;align-items:center;font-family:var(--mono);font-size:11.5px;
font-weight:700;padding:3px 10px;border-radius:5px;letter-spacing:.4px;
white-space:nowrap;word-break:keep-all;overflow-wrap:normal;hyphens:none}

/* özet tablo */
.ozet-tablo{width:100%;border-collapse:collapse}
.ozet-tablo th{font-size:11px;color:var(--gri);letter-spacing:.3px;font-weight:600;
text-align:left;padding:8px 8px;background:transparent;border-bottom:1px solid var(--cizgi);
position:static}
.ozet-tablo td{padding:9px 8px;border-bottom:1px solid var(--cizgi2);font-size:12.5px;
color:var(--ink)}
.ozet-tablo tr:last-child td{border-bottom:0}
.ozet-tablo td.say{text-align:right;font-family:var(--mono);font-variant-numeric:tabular-nums;
white-space:nowrap}
.minidonut{vertical-align:middle}
.minidonut .iz{stroke:var(--iz)}

/* =====================================================================
   DÜZELTME LİSTESİ (SAP çalışma alanı)
   ===================================================================== */
.arac{position:sticky;top:var(--ustyuk);z-index:12;display:flex;gap:8px;align-items:center;
margin:0 -18px 12px;padding:11px 18px;flex-wrap:wrap;background:var(--tile);
border-bottom:1px solid var(--cizgi2)}
.arac-sar{position:relative;flex:1;min-width:220px;display:flex;align-items:center}
.arac-sar::before{content:"";position:absolute;left:11px;width:15px;height:15px;
background:var(--silik);pointer-events:none;
-webkit-mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='2' stroke-linecap='round'%3E%3Ccircle cx='11' cy='11' r='7'/%3E%3Cpath d='m20 20-3.5-3.5'/%3E%3C/svg%3E");
mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='2' stroke-linecap='round'%3E%3Ccircle cx='11' cy='11' r='7'/%3E%3Cpath d='m20 20-3.5-3.5'/%3E%3C/svg%3E");
-webkit-mask-repeat:no-repeat;mask-repeat:no-repeat;-webkit-mask-size:contain;mask-size:contain}
.arama{width:100%;border:1px solid var(--cizgi);border-radius:9px;background:var(--tile-duz);
color:var(--ink);padding:9px 13px 9px 34px;font-size:12px;font-family:var(--mono);
transition:all .15s}
.arama::placeholder{font-family:var(--f);color:var(--silik)}
.arama:focus{outline:0;border-color:var(--marka);background:var(--tile);
box-shadow:0 0 0 3px rgba(0,51,161,.10)}
.cipler{display:flex;gap:6px;flex-wrap:wrap;align-items:center}
.cip{cursor:pointer;border:1px solid var(--cizgi);border-radius:8px;background:var(--tile);
color:var(--gri);padding:7px 13px;font-size:11.5px;font-weight:600;transition:all .15s}
.cip:hover{border-color:var(--marka-cizgi);color:var(--marka);background:var(--marka-yumusak)}
.cip.aktif{background:var(--marka);border-color:var(--marka);color:#fff}
body.gece .cip.aktif{color:#0B1220}
.arac-dugme{display:inline-flex;align-items:center;gap:7px;cursor:pointer;
border:1px solid var(--cizgi);border-radius:8px;background:var(--tile);color:var(--ink2);
padding:7px 13px;font-size:11.5px;font-weight:600;transition:all .15s}
.arac-dugme:hover{border-color:var(--marka-cizgi);color:var(--marka);background:var(--marka-yumusak)}
.arac-dugme.aktif{border-color:var(--marka);color:var(--marka);background:var(--marka-yumusak)}
.arac-dugme::before{-webkit-mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.9' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M4 20V4'/%3E%3Cpath d='M4 20h16'/%3E%3Cpath d='m7 15 4-5 3.5 3L20 7'/%3E%3C/svg%3E");
mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23000' stroke-width='1.9' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M4 20V4'/%3E%3Cpath d='M4 20h16'/%3E%3Cpath d='m7 15 4-5 3.5 3L20 7'/%3E%3C/svg%3E")}
.rozet-say{display:inline-flex;align-items:center;font-family:var(--mono);font-size:11.5px;
font-weight:600;color:var(--ink2);background:var(--tile-duz);border:1px solid var(--cizgi);
border-radius:99px;padding:5px 12px;white-space:nowrap;font-variant-numeric:tabular-nums}

.tabsar{position:relative;width:100%;max-width:100%;overflow-x:auto;border:1px solid var(--cizgi);
border-radius:10px;scrollbar-width:thin;scrollbar-color:var(--silik) transparent;
background:
linear-gradient(90deg,var(--tile) 30%,rgba(255,255,255,0)) left / 30px 100% no-repeat,
linear-gradient(90deg,rgba(255,255,255,0),var(--tile) 70%) right / 30px 100% no-repeat,
radial-gradient(farthest-side at 0 50%,rgba(16,35,59,.16),transparent) left / 14px 100% no-repeat,
radial-gradient(farthest-side at 100% 50%,rgba(16,35,59,.16),transparent) right / 14px 100% no-repeat;
background-attachment:local,local,scroll,scroll}
.tabsar::-webkit-scrollbar{height:11px;width:11px}
.tabsar::-webkit-scrollbar-thumb{background:var(--silik);border-radius:99px;
border:3px solid var(--tile)}
.tabsar::-webkit-scrollbar-track{background:var(--tile-duz)}
table{width:100%;min-width:1020px;border-collapse:separate;border-spacing:0;table-layout:fixed}
th{font-size:11px;letter-spacing:.3px;color:var(--gri);font-weight:700;text-align:left;
padding:8px 10px;border-bottom:1px solid var(--cizgi);background:var(--tile-duz);
position:sticky;top:0;z-index:6;white-space:nowrap;
overflow:hidden;text-overflow:ellipsis}
td{padding:6px 10px;border-bottom:1px solid var(--cizgi2);font-size:12px;
line-height:1.38;vertical-align:top;
word-break:break-word;overflow-wrap:anywhere;hyphens:auto;background:var(--tile)}
tbody tr:nth-child(even) td{background:var(--satir)}
tbody tr:hover td{background:var(--satir-hover)}
/* Yapışkan başlık + sabit ilk sütun için tablo kabı kendi kaydırma alanıdır */
@media(min-width:1000px){
  .tabsar{max-height:calc(100vh - var(--ustyuk) - 208px);overflow:auto}
  th:first-child,td:first-child{position:sticky;left:0;z-index:7;
    box-shadow:1px 0 0 var(--cizgi)}
  td:first-child{z-index:5}
}
td:has(> .khap){word-break:keep-all;overflow-wrap:normal}
td.kod{font-family:var(--mono);font-size:11.5px;white-space:normal;max-width:0;
overflow-wrap:anywhere;word-break:break-word;line-height:1.4;color:var(--ink)}
td.mv{font-family:var(--mono);font-size:11.5px;color:var(--ink2)}
td.og{font-family:var(--mono);font-size:11.5px;color:var(--ink);font-weight:600}
td.oneri{font-family:var(--mono);font-size:11.5px;color:var(--teal-koyu);
box-shadow:inset 3px 0 0 var(--teal);background-clip:padding-box}
body.gece td.oneri{color:#4FD8C9}
.khap{display:inline-block;font-family:var(--mono);font-size:11px;font-weight:700;color:#fff;
padding:3px 8px;border-radius:5px;letter-spacing:.3px;
white-space:nowrap;word-break:keep-all;overflow-wrap:normal;hyphens:none}
.cokhata{display:block;padding:1px 0 1px 8px;border-left:2px solid var(--cizgi);margin:2px 0}
.cokhata:first-child{border-left-color:var(--teal)}
.cokhata:nth-child(2){border-left-color:var(--mercan)}
/* İKİ KOD SÜTUNLU MOD (Tümü) — toplam %100 */
table.ciftkod th:nth-child(1),table.ciftkod td:nth-child(1){width:15%}
table.ciftkod th:nth-child(2),table.ciftkod td:nth-child(2){width:8%}
table.ciftkod th:nth-child(3),table.ciftkod td:nth-child(3){width:8%}
table.ciftkod th:nth-child(4),table.ciftkod td:nth-child(4){width:6%}
table.ciftkod th:nth-child(5),table.ciftkod td:nth-child(5){width:18%}
table.ciftkod th:nth-child(6),table.ciftkod td:nth-child(6){width:10%}
table.ciftkod th:nth-child(7),table.ciftkod td:nth-child(7){width:11%}
table.ciftkod th:nth-child(8),table.ciftkod td:nth-child(8){width:9%}
table.ciftkod th:nth-child(9),table.ciftkod td:nth-child(9){width:4%}
table.ciftkod th:nth-child(10),table.ciftkod td:nth-child(10){width:11%}
/* TEK KOD SÜTUNLU MOD (Ürün Kodu / Üst Kod) — toplam %100 */
table.tekkod th:nth-child(1),table.tekkod td:nth-child(1){width:16%}
table.tekkod th:nth-child(2),table.tekkod td:nth-child(2){width:9%}
table.tekkod th:nth-child(3),table.tekkod td:nth-child(3){width:6%}
table.tekkod th:nth-child(4),table.tekkod td:nth-child(4){width:19%}
table.tekkod th:nth-child(5),table.tekkod td:nth-child(5){width:11%}
table.tekkod th:nth-child(6),table.tekkod td:nth-child(6){width:12%}
table.tekkod th:nth-child(7),table.tekkod td:nth-child(7){width:10%}
table.tekkod th:nth-child(8),table.tekkod td:nth-child(8){width:5%}
table.tekkod th:nth-child(9),table.tekkod td:nth-child(9){width:12%}

.daha{display:flex;align-items:center;gap:8px;margin:16px auto 0;cursor:pointer;
border:1px solid var(--cizgi);border-radius:9px;background:var(--tile);color:var(--ink2);
padding:10px 22px;font-size:12px;font-weight:600;box-shadow:var(--golge);transition:all .15s}
.daha:hover{border-color:var(--marka-cizgi);color:var(--marka);transform:translateY(-1px)}
.bos{text-align:center;color:var(--gri);padding:26px 12px;font-size:12.5px}
.dipnot{margin-top:14px;color:var(--gri);font-size:11.5px;text-align:center}
.altnot{color:var(--gri);font-size:11.5px;margin-top:6px;line-height:1.5}
.buyuk{font-size:40px;font-weight:300;line-height:1.1;margin-top:6px}
.oneri{font-family:var(--mono);font-size:11.5px;color:var(--teal-koyu)}

/* =====================================================================
   TARİHÇE
   ===================================================================== */
.gecmis-grafik{margin:4px 0 16px;padding:12px 4px 4px;background:var(--tile-duz);
border:1px solid var(--cizgi2);border-radius:10px}
.gecmis-grafik:empty{display:none}
.gecmis-grafik .iz{stroke:var(--iz)}
.gecmis-grafik text{fill:var(--gri)}
.gbaslik,.gsatir{display:grid;
grid-template-columns:minmax(140px,1.3fr) 92px minmax(120px,2fr) 96px 82px 38px;
gap:12px;align-items:center}
.gbaslik{padding:8px 10px;border-bottom:1px solid var(--cizgi)}
.gbaslik span{font-size:11px;letter-spacing:.3px;color:var(--gri);font-weight:600}
.gbaslik span:nth-child(4),.gbaslik span:nth-child(5){text-align:right}
.gsatir{padding:11px 10px;border-bottom:1px solid var(--cizgi2);position:relative;cursor:help;
border-left:3px solid transparent;border-radius:0 8px 8px 0;transition:background .15s}
.gsatir:hover{background:var(--satir-hover)}
.gsatir.son{background:var(--marka-yumusak);border-left-color:var(--marka)}
.gsatir.son:hover{background:var(--marka-yumusak)}
.gtarih{font-family:var(--mono);font-size:12px;color:var(--ink)}
.gkaynak{display:inline-block;font-family:var(--f);font-size:11px;font-weight:600;
letter-spacing:.5px;color:var(--gri);background:var(--tile-duz);border:1px solid var(--cizgi);
border-radius:4px;padding:1px 6px;margin-top:4px}
.gsaglik{font-size:17px;font-weight:600;font-family:var(--mono);letter-spacing:-.5px}
.gray{height:9px;background:var(--iz);border-radius:99px;overflow:hidden;display:block}
.gray div{height:100%;border-radius:99px}
.gfark{font-size:12.5px;font-weight:600;text-align:right;font-family:var(--mono)}
.gbulgu{font-size:12px;color:var(--ink2);text-align:right;font-family:var(--mono)}
.gsil{display:inline-flex;align-items:center;justify-content:center;cursor:pointer;
border:1px solid var(--cizgi);border-radius:7px;background:transparent;color:var(--silik);
padding:5px;transition:all .15s}
.gsil:hover{border-color:var(--mercan);color:var(--mercan);background:var(--mercan-yumusak)}
.gsil:disabled{opacity:.32;cursor:not-allowed}

/* =====================================================================
   İPUCU BALONU
   ===================================================================== */
.ipuc{position:relative;cursor:help}
.ipuc::after{content:attr(data-ipucu);position:absolute;left:0;top:calc(100% + 8px);z-index:80;
width:340px;max-width:min(340px,calc(100vw - 40px));padding:12px 14px;box-sizing:border-box;
background:#12233B;color:#EDF2F8;font-family:var(--f);font-size:11.5px;font-weight:400;
line-height:1.55;letter-spacing:0;text-align:left;text-transform:none;border-radius:8px;
box-shadow:0 10px 30px rgba(0,0,0,.34);opacity:0;visibility:hidden;transform:translateY(-4px);
transition:opacity .15s,transform .15s,visibility .15s;pointer-events:none;white-space:pre-line}
body.gece .ipuc::after{background:#28374B;box-shadow:0 10px 30px rgba(0,0,0,.5)}
.ipuc::before{content:"";position:absolute;left:14px;top:calc(100% + 2px);z-index:81;
border:6px solid transparent;border-bottom-color:#12233B;opacity:0;visibility:hidden;
transition:opacity .15s,visibility .15s}
body.gece .ipuc::before{border-bottom-color:#28374B}
.ipuc:hover::after,.ipuc:focus::after{opacity:1;visibility:visible;transform:translateY(0)}
.ipuc:hover::before,.ipuc:focus::before{opacity:1;visibility:visible}
tbody .ipuc::after,.arac .ipuc::after{left:auto;right:0}
tbody .ipuc::before,.arac .ipuc::before{left:auto;right:14px}
.gsatir.ipuc::after{left:16px;right:auto}
.minisutlar{position:relative}
/* eski sürüm uyumluluğu */
.rzt{background:var(--tile);border:1px solid var(--cizgi);border-radius:10px;padding:10px 14px}
.rzt b{display:block;font-size:20px;font-weight:600}
.rzt span{font-size:11px;color:var(--gri)}

/* =====================================================================
   VERİ YÜKLEME SAYFASI
   ===================================================================== */
#p0{padding:0}
.p0-hero{min-height:100vh;position:relative;overflow:hidden;display:flex;align-items:center;
justify-content:center;background:linear-gradient(160deg,#0E2144 0%,#0B1B33 48%,#08192F 100%);
padding:38px 16px}
.p0-hero .g-izgara{opacity:.9}
.p0-kart{position:relative;width:640px;max-width:96vw;background:var(--tile);
border:1px solid var(--cizgi);border-radius:16px;padding:26px 30px 28px;
box-shadow:0 24px 60px rgba(0,0,0,.38)}
.p0-ust{display:flex;align-items:center;justify-content:space-between;gap:14px;
padding-bottom:16px;margin-bottom:18px;border-bottom:1px solid var(--cizgi);flex-wrap:wrap}
.etiket2{display:inline-flex;align-items:center;gap:7px;font-size:11px;font-weight:700;
letter-spacing:1.4px;color:var(--marka);text-transform:uppercase}
.p0-baslik{font-family:var(--fb);font-size:26px;font-weight:700;margin:10px 0 6px;
color:var(--ink);letter-spacing:-.6px}
.p0-alt{color:var(--gri);font-size:12.5px;line-height:1.65;margin-bottom:8px}
.adim-bas{display:flex;align-items:center;gap:9px;margin:18px 0 10px;
font-size:11.5px;font-weight:700;letter-spacing:.6px;color:var(--ink2);text-transform:uppercase}
.adim-bas::after{content:"";flex:1;height:1px;background:var(--cizgi2)}
.adim-bas .rozet-adim{display:inline-flex;align-items:center;justify-content:center;
width:20px;height:20px;border-radius:6px;background:var(--marka);color:#fff;font-size:11px;
font-weight:700;flex:none}
body.gece .adim-bas .rozet-adim{color:#0B1220}
.adim-bas .istege{font-weight:600;letter-spacing:.2px;text-transform:none;color:var(--gri);
background:var(--tile-duz);border:1px solid var(--cizgi);border-radius:99px;padding:3px 9px;
font-size:11px}
.dosya-kutu{display:flex;align-items:center;gap:13px;border:1px solid var(--cizgi);
border-radius:10px;padding:12px 14px;margin-bottom:8px;cursor:pointer;
background:var(--tile-duz);transition:border-color .15s,background .15s,transform .15s}
.dosya-kutu:hover{border-color:var(--marka-cizgi);background:var(--marka-yumusak);
transform:translateY(-1px)}
.dosya-kutu.dolu{border-color:var(--teal);background:var(--teal-yumusak)}
.dosya-kutu.dolu .dk-no{background:var(--teal);border-color:var(--teal);color:#fff}
.dosya-kutu input{display:none}
.dosya-kutu.birlesik{border-style:dashed;border-color:var(--marka-cizgi)}
.dosya-kutu.pasif{opacity:.45;cursor:not-allowed;filter:grayscale(.6);transform:none}
.dosya-kutu.pasif:hover{border-color:var(--cizgi);background:var(--tile-duz);transform:none}
.dosya-kutu.pasif .dk-no{background:var(--silik);border-color:var(--silik);color:#fff}
.dk-no{flex:none;width:32px;height:32px;border-radius:9px;background:var(--marka-yumusak);
border:1px solid var(--marka-cizgi);color:var(--marka);font-weight:700;font-size:13.5px;
display:flex;align-items:center;justify-content:center}
.dk-metin{flex:1;min-width:0}
.dk-metin b{display:block;font-size:13.5px;color:var(--ink);font-weight:600}
.dk-metin small{display:block;color:var(--gri);font-size:11.5px;margin-top:2px;line-height:1.45}
.dosya-kutu.dolu .dk-metin small{color:var(--teal-koyu);font-family:var(--mono);font-weight:600}
body.gece .dosya-kutu.dolu .dk-metin small{color:#4FD8C9}
.kutu-not{display:block;font-size:11px;color:var(--amber);margin-top:3px}
.kutu-not:empty{display:none}
.dk-sil{flex:none;width:26px;height:26px;border-radius:7px;border:1px solid var(--cizgi);
background:var(--tile);color:var(--silik);cursor:pointer;display:none;align-items:center;
justify-content:center;transition:all .15s;padding:0}
.dk-sil:hover{border-color:var(--mercan);color:var(--mercan);background:var(--mercan-yumusak)}
.dosya-kutu.dolu .dk-sil{display:flex}
.ayirici{display:flex;align-items:center;gap:10px;margin:14px 0 9px;color:var(--gri);
font-size:11px;font-weight:600;letter-spacing:.6px;text-transform:uppercase}
.ayirici::before,.ayirici::after{content:"";flex:1;height:1px;background:var(--cizgi2)}
.p0-butonlar{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin:18px 0 4px}
.analiz{display:inline-flex;align-items:center;gap:9px;cursor:pointer;border:0;border-radius:9px;
background:var(--marka);color:#fff;padding:12px 24px;font-size:13.5px;font-weight:600;
box-shadow:0 4px 14px rgba(0,51,161,.28);transition:transform .15s,background .15s}
.analiz:hover{transform:translateY(-2px);background:var(--marka2)}
.analiz:disabled{opacity:.55;cursor:wait;transform:none}
body.gece .analiz{color:#0B1220}
.ghost{display:inline-flex;align-items:center;gap:8px;cursor:pointer;border:1px solid var(--cizgi);
border-radius:9px;background:var(--tile);color:var(--ink2);padding:11px 17px;font-size:12.5px;
font-weight:600;transition:all .15s}
.ghost:hover{border-color:var(--marka-cizgi);color:var(--marka);background:var(--marka-yumusak)}
.guven{display:flex;align-items:center;gap:9px;margin-top:14px;padding:10px 12px;
background:var(--teal-yumusak);border:1px solid var(--cizgi2);border-radius:9px;
font-size:11.5px;color:var(--ink2);line-height:1.5}
.guven .ikn{color:var(--teal-koyu);flex:none}
body.gece .guven .ikn{color:#4FD8C9}
.teknik-bas{display:flex;align-items:center;gap:8px;margin:16px 0 7px;font-size:11px;
font-weight:700;letter-spacing:.6px;color:var(--gri);text-transform:uppercase}
.teknik-bas::after{content:"";flex:1;height:1px;background:var(--cizgi2)}
.logk{font-family:var(--mono);font-size:11.5px;line-height:1.7;color:var(--ink2);
white-space:pre-wrap;max-height:180px;overflow:auto;background:var(--tile-duz);
border:1px solid var(--cizgi2);border-radius:9px;padding:11px 13px}

/* =====================================================================
   DUYARLI DÜZEN
   ===================================================================== */
@media(max-width:1200px){
  .kpi-serit{grid-template-columns:1.5fr repeat(2,1fr)}
  .s3,.s4,.s5{grid-column:span 6}.s7{grid-column:span 12}
}
@media(max-width:1000px){
  :root{--ustyuk:0px}
  .ustbar{position:static}
  .duzen{flex-direction:column}
  .yan{width:100%;position:static;display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));
    gap:0 16px;align-items:start}
  .yan .marka-blok{grid-column:1/-1}
  .yan .f-grup,.yan .f-grup:first-of-type{border-top:0;padding-top:12px;margin-top:0}
  .icerik{width:100%;min-width:0}
  .arac{position:static;margin:0 0 12px;padding:0;background:transparent;border-bottom:0}
  .tabsar{max-height:none;overflow-x:auto;overflow-y:visible}
  th{position:static}
  .g-kahraman{padding:28px 0 24px;gap:28px}
  .g-sag{width:100%}
}
@media(max-width:820px){
  .kabuk{padding:16px 16px 0}
  .ub-sar{padding:0 16px}
  .kpi-serit{grid-template-columns:repeat(2,1fr)}
  .kpi.ana{grid-column:1/-1}
  .s2,.s3,.s4,.s5,.s7,.s12,.s24{grid-column:span 12}
  .gbaslik{display:none}
  .gsatir{grid-template-columns:1fr auto;gap:6px 12px;padding:12px 10px}
  .gsatir .gray{grid-column:1/-1}
  .g-metrik{grid-template-columns:1fr;max-width:none}
  .p0-kart{padding:22px 18px 24px}
}
@media(max-width:560px){
  .kpi-serit{grid-template-columns:1fr}
  .kabuk{padding:14px 12px 0}
  .ub-sar{padding:0 12px}
  .tile{padding:14px 14px 15px}
  .g-ic{padding:16px 18px 24px}
  .halka{gap:12px}
}
@media(prefers-reduced-motion:reduce){
  *{animation-duration:.01ms!important;transition-duration:.01ms!important}
}

/* =====================================================================
   YAZDIRMA / PDF
   ===================================================================== */
.baski-ust{display:none}
@media print{
  @page{margin:12mm}
  :root{--ustyuk:0px}
  body{background:#fff;color:#12233B;font-size:10.5px}
  body.gece{--bg:#fff;--tile:#fff;--tile-ust:#fff;--tile-duz:#fff;--ink:#12233B;
    --ink2:#3E5169;--gri:#68798F;--cizgi:#D6DEE8;--cizgi2:#E6ECF3;--iz:#E4EAF1;
    --satir:#F6F8FB;--marka:#0033A1}
  .baski-ust{display:flex!important;align-items:center;gap:14px;padding:0 0 10px;
    margin-bottom:14px;border-bottom:2px solid #0033A1}
  .baski-ust .kilit-logo img,.baski-ust .kilit-logo svg{height:26px}
  .baski-ust .baski-bilgi{margin-left:auto;text-align:right;font-size:10px;color:#3E5169;
    line-height:1.5}
  .baski-ust b{font-size:14px;font-family:var(--fb)}
  .ustbar,.sekmeler,.komut,.ub-arac,.ub-cip,.arac,.daha,.yan,.g-ustdugme,.dk-sil,.gsil,
  .p0-butonlar,#indirGrup{display:none!important}
  .sayfa{padding:0}
  .kabuk{padding:0;max-width:none}
  .duzen{display:block}
  .icerik{width:100%}
  .tile,.kpi,.yan{box-shadow:none!important;border:1px solid #D6DEE8;
    break-inside:avoid;page-break-inside:avoid}
  .izg{gap:10px}
  .sayfa-bas{margin-bottom:10px}
  .sayfa-bas-yan{display:none}
  .tbaslik::before,.tbaslik::after{display:none}
  .tbaslik{padding-left:0;min-height:0}
  .tabsar{overflow:visible;max-height:none;border:1px solid #D6DEE8;background:none}
  table{min-width:0}
  th{position:static;background:#F1F4F8}
  th:first-child,td:first-child{position:static;box-shadow:none}
  tbody tr{break-inside:avoid;page-break-inside:avoid}
  .ipuc::after,.ipuc::before{display:none!important}
  .gsatir{break-inside:avoid}
}
"""
    khap_stil = (
        ":root{"
        + "".join(f"--k{i+1}:{RENK[f'K{i+1}']};" for i in range(n_kural))
        + "}"
        + "".join(f'.khap.k{i+1}{{background:var(--k{i+1})}}' for i in range(n_kural)))

    js = r"""
<script src="https://cdn.jsdelivr.net/npm/xlsx-js-style@1.2.0/dist/xlsx.bundle.js"
 onerror="var s=document.createElement('script');s.src='https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js';document.head.appendChild(s);"></script>
<script>
const CFG = __CFG__;
const GECMIS_GOMULU = __GECMIS__;
/* Satır içi SVG ikon seti — Python tarafındaki ik() ile aynı yolları kullanır,
   böylece yeniden çizilen kartlar da aynı görsel dili korur. */
const IKONLAR = {
  saglik:"<path d='M3 12h4l2.5-6 4 12 2.5-6H21'/>",
  bulgu:"<path d='M12 4.5 21 20H3Z'/><path d='M12 10v4.5'/><path d='M12 17.4h.01'/>",
  liste:"<rect x='3' y='4' width='18' height='16' rx='2'/><path d='M3 9h18'/><path d='M9 9v11'/><path d='M3 14.5h18'/>",
  malzeme:"<path d='m12 3 8.5 4.6v8.8L12 21l-8.5-4.6V7.6Z'/><path d='m3.5 7.6 8.5 4.7 8.5-4.7'/><path d='M12 12.3V21'/>",
  onay:"<circle cx='12' cy='12' r='8.5'/><path d='m8.3 12.2 2.6 2.6 4.8-5.2'/>",
  kapat:"<path d='m6.5 6.5 11 11'/><path d='m17.5 6.5-11 11'/>",
  ok:"<path d='m9 5 7 7-7 7'/>"
};
function IK(ad,boyut,sinif){
  return "<svg class='"+(sinif||"ikn")+"' width='"+(boyut||16)+"' height='"+(boyut||16)
    +"' viewBox='0 0 24 24' fill='none' stroke='currentColor' stroke-width='1.7' "
    +"stroke-linecap='round' stroke-linejoin='round' aria-hidden='true' focusable='false'>"
    +(IKONLAR[ad]||"")+"</svg>";
}
const GECMIS_ANAHTAR = "veriKalitesiGecmis";

function gecmisOku(){
  let yerel = [];
  try { yerel = JSON.parse(localStorage.getItem(GECMIS_ANAHTAR) || "[]"); }
  catch(e){ yerel = []; }
  return (GECMIS_GOMULU || []).concat(yerel);
}
function gecmisYaz(liste){
  try { localStorage.setItem(GECMIS_ANAHTAR, JSON.stringify(liste.slice(-60))); }
  catch(e){ /* depolama kapalıysa sessizce geç */ }
}
function gecmisEkle(S){
  const alanSay = {};
  for(const r0 of S.rows){ if(r0.k!=="ALT"&&r0.a) alanSay[r0.a]=(alanSay[r0.a]||0)+1; }
  const d = new Date();
  const iki = n => String(n).padStart(2,"0");
  const kayit = {
    t: iki(d.getDate())+"."+iki(d.getMonth()+1)+"."+d.getFullYear()+" "+iki(d.getHours())+":"+iki(d.getMinutes()),
    kaynak: "panel",
    saglik: +(S.toplamM ? 100*S.dogru/S.toplamM : 0).toFixed(1),
    toplam: S.toplamM, dogru: S.dogru, hatali: S.hatali,
    sayilar: {K1:S.sayilar.K1[0],K2:S.sayilar.K2[0],K3:S.sayilar.K3[0],
              K4:S.sayilar.K4[0],K5:S.sayilar.K5[0],K6:S.sayilar.K6[0]},
    alanlar: alanSay
  };
  let yerel = [];
  try { yerel = JSON.parse(localStorage.getItem(GECMIS_ANAHTAR) || "[]"); } catch(e){}
  yerel.push(kayit); gecmisYaz(yerel);
  return kayit;
}
function gecmisSil(indeks){
  // Yalnızca tarayıcıda tutulan (PANEL) kayıtlar silinebilir; program
  // koşumları panele gömülüdür ve dosyadan gelir.
  const gomuluN = (GECMIS_GOMULU || []).length;
  if(indeks < gomuluN) return;
  let yerel = [];
  try { yerel = JSON.parse(localStorage.getItem(GECMIS_ANAHTAR) || "[]"); }
  catch(e){ yerel = []; }
  yerel.splice(indeks - gomuluN, 1);
  gecmisYaz(yerel);
  gecmisCiz();
}
function ilkN(nesne, n){
  return Object.entries(nesne||{}).sort((a,b)=>b[1]-a[1]).slice(0,n);
}
function gecmisOzet(kayit, onceki){
  const enCok = ilkN(kayit.alanlar, 3).map(([a,v])=>a+" ("+v+")");
  let metin = "En çok hata alan alanlar: " + (enCok.length?enCok.join(" · "):"—");
  if(onceki){
    const fark = +(kayit.saglik - onceki.saglik).toFixed(1);
    metin += "\n\nÖnceki çalıştırmaya göre veri sağlığı: "
      + (fark>0?"+":"") + fark + " puan";
    const iyi=[], kotu=[];
    const tumAlan = new Set([...Object.keys(kayit.alanlar||{}),
                             ...Object.keys(onceki.alanlar||{})]);
    tumAlan.forEach(a=>{
      const y=(kayit.alanlar||{})[a]||0, e=(onceki.alanlar||{})[a]||0;
      if(y<e) iyi.push(a+" (-"+(e-y)+")");
      else if(y>e) kotu.push(a+" (+"+(y-e)+")");
    });
    if(iyi.length) metin += "\nİyileşen: " + iyi.slice(0,4).join(" · ");
    if(kotu.length) metin += "\nKötüleşen: " + kotu.slice(0,4).join(" · ");
    if(!iyi.length && !kotu.length) metin += "\nAlan bazında değişiklik yok.";
  } else {
    metin += "\n\nKarşılaştırılacak önceki çalıştırma yok.";
  }
  const kurallar = Object.entries(kayit.sayilar||{})
    .map(([k,v])=>k+"="+v).join(" · ");
  if(kurallar) metin += "\n\nKural bazında: " + kurallar;
  return metin;
}
function gecmisCiz(){
  const liste = gecmisOku();
  const yer = document.getElementById("gecmisListe");
  const bilgi = document.getElementById("gecmisBilgi");
  if(!yer) return;
  if(bilgi) bilgi.textContent = liste.length + " çalıştırma kaydı";
  if(!liste.length){
    yer.innerHTML = "<div class='bos'>Henüz kayıt yok — veri yükleyip analiz "
      + "ettiğinizde bu sayfa dolmaya başlar.</div>";
    const g=document.getElementById("gecmisGrafik"); if(g) g.innerHTML="";
    return;
  }
  const gomuluN = (GECMIS_GOMULU || []).length;
  yer.innerHTML = "<div class='gbaslik'><span>Çalıştırma</span><span>Veri sağlığı</span>"
    + "<span>Dağılım</span><span>Değişim</span><span>Bulgu</span><span></span></div>"
    + liste.map((k,i)=>{
      const onceki = i>0 ? liste[i-1] : null;
      const fark = onceki ? +(k.saglik-onceki.saglik).toFixed(1) : null;
      const renk = k.saglik>=95?"#01B8AA":(k.saglik>=85?"#E3A63B":"#FD625E");
      const farkRenk = fark===null?"#9AA7B0":(fark>0?"#0E8A6E":(fark<0?"#C0392B":"#9AA7B0"));
      const farkMetin = fark===null?"—":((fark>0?"▲ +":(fark<0?"▼ ":"● "))+fark);
      const toplamBulgu = Object.values(k.sayilar||{}).reduce((a,b)=>a+b,0);
      return "<div class='gsatir ipuc" + (i===liste.length-1 ? " son" : "")
        + "' data-ipucu=\""
        + gecmisOzet(k,onceki).replace(/"/g,"&quot;") + "\">"
        + "<span class='gtarih'>"+kacis(k.t)+"<br><span class='gkaynak'>"
          +(k.kaynak==="python"?"PROGRAM":"PANEL")+"</span></span>"
        + "<span class='gsaglik' style='color:"+renk+"'>%"+k.saglik.toFixed(1)+"</span>"
        + "<span class='gray'><div style='width:"+k.saglik+"%;background:"+renk+"'></div></span>"
        + "<span class='gfark' style='color:"+farkRenk+"'>"+farkMetin+"</span>"
        + "<span class='gbulgu'>"+fmt(toplamBulgu)+"</span>"
        + (i < gomuluN
            ? "<button class='gsil' disabled title='Program koşumu — panelden silinemez'>"+IK("kapat",14)+"</button>"
            : "<button class='gsil' data-sil='"+i+"' title='Bu kaydı sil'>"+IK("kapat",14)+"</button>")
        + "</div>";
    }).join("");
  grafikCiz(liste);
}
function grafikCiz(liste){
  const yer = document.getElementById("gecmisGrafik");
  if(!yer || yer.dataset.gizli==="1"){ if(yer) yer.innerHTML=""; return; }
  const G=560, Y=170, sol=42, alt=26, ust=12, sag=12;
  const n = liste.length;
  const x = i => n<2 ? sol+(G-sol-sag)/2 : sol+(G-sol-sag)*i/(n-1);
  const y = v => ust+(Y-ust-alt)*(1-v/100);
  const cizgi = liste.map((k,i)=>(i?"L":"M")+x(i).toFixed(1)+","+y(k.saglik).toFixed(1)).join(" ");
  const alan = cizgi + " L"+x(n-1).toFixed(1)+","+(Y-alt)+" L"+x(0).toFixed(1)+","+(Y-alt)+" Z";
  const noktalar = liste.map((k,i)=>
    "<circle cx='"+x(i).toFixed(1)+"' cy='"+y(k.saglik).toFixed(1)+"' r='3.5' fill='#01B8AA'>"
    + "<title>"+kacis(k.t+" — %"+k.saglik.toFixed(1))+"</title></circle>").join("");
  const izgara = [0,25,50,75,100].map(v=>
    "<line class='iz' x1='"+sol+"' y1='"+y(v).toFixed(1)+"' x2='"+(G-sag)+"' y2='"+y(v).toFixed(1)
    +"' stroke='currentColor'/><text x='"+(sol-7)+"' y='"+(y(v)+3.5).toFixed(1)
    +"' text-anchor='end' font-size='10'>%"+v+"</text>").join("");
  const etiketler = liste.map((k,i)=>
    (n<=8||i===0||i===n-1||i%Math.ceil(n/6)===0)
      ? "<text x='"+x(i).toFixed(1)+"' y='"+(Y-8)+"' text-anchor='middle' font-size='10'>"
        + kacis(String(k.t).split(" ")[0]) + "</text>" : "").join("");
  yer.innerHTML = "<svg viewBox='0 0 "+G+" "+Y+"' style='width:100%;max-width:760px'>"
    + izgara
    + "<path d='"+alan+"' fill='rgba(1,184,170,.12)'/>"
    + "<path d='"+cizgi+"' fill='none' stroke='#01B8AA' stroke-width='2'/>"
    + noktalar + etiketler + "</svg>";
}
let BULGULAR = __VERI__;
let filtre = "ALL", kaynakF = "ALL", alanF = "ALL", arama = "", limit = 60;
const tbody = document.getElementById("tb");
const bilgi = document.getElementById("sonucBilgi");
function kacis(t){const d=document.createElement("div");d.textContent=t;return d.innerHTML}
function cokSatir(t){
  const s=String(t==null?"":t);
  if(s.indexOf("\n")<0) return kacis(s);
  return s.split("\n").map(p=>'<span class="cokhata">'+kacis(p)+'</span>').join("");
}
function ciz(){
  const q = arama.toLowerCase();
  const liste = BULGULAR.filter(r =>

    (filtre === "ALL" || r.k === filtre) &&
    (kaynakF === "ALL" || r.kay === kaynakF) &&
    (alanF === "ALL" || r.a === alanF ||
      String(r.a || "").split(/\s*[;/]\s*/).indexOf(alanF) >= 0) &&
    (!q || ((r.t||"")+" "+r.m+" "+r.u+" "+r.n+" "+r.a+" "+r.mv+" "+r.og+" "+r.d).toLowerCase().includes(q)));
  bilgi.textContent = liste.length + " bulgu"
    + (filtre !== "ALL" ? " · " + filtre : "")
    + (kaynakF !== "ALL" ? " · " + kaynakF : "")
    + (alanF !== "ALL" ? " · " + alanF : "")
    + (q ? " · [" + arama + "]" : "");
  // Kaynak seçimine göre kod sütunları: Tümü -> Malzeme + Üst Kod,
  // Ürün Kodu -> tek sütun "Malzeme (Ürün Kodu)", Üst Kod -> tek sütun "Üst Kod"
  const tekSutun = kaynakF !== "ALL";
  const kodBaslik = kaynakF === "Üst Kod" ? "Üst Kod"
                  : (kaynakF === "Ürün Kodu" ? "Malzeme (Ürün Kodu)" : null);
  const tablo = document.querySelector(".tabsar table");
  if (tablo) {
    tablo.classList.toggle("tekkod", tekSutun);
    tablo.classList.toggle("ciftkod", !tekSutun);
  }
  const bas = document.getElementById("basliklar");
  if (bas) {
    const kodlar = tekSutun ? "<th>" + kodBaslik + "</th>"
                            : "<th>Malzeme</th><th>Üst Kod</th>";
    bas.innerHTML = "<th>Kısa Metin</th>" + kodlar +
      "<th>Kural</th><th>Hata Nedeni</th><th>İlgili Alan</th>" +
      "<th>Mevcut Değer</th><th>Olması Gereken</th><th>Alt Tarif</th>" +
      "<th>Yapılacak Düzeltme</th>";
  }
  const sutunSayisi = tekSutun ? 9 : 10;

  const g = liste.slice(0, limit);
  tbody.innerHTML = g.length ? g.map(r =>
    "<tr><td>" + kacis(r.t || "") + "</td>" +
    (tekSutun
      ? "<td class='kod'>" + kacis(r.m) + "</td>"
      : "<td class='kod'>" + kacis(r.m) + "</td>" +
        "<td class='kod'>" + (r.u ? kacis(r.u) : "—") + "</td>") +
    "<td><span class='khap ipuc " + r.k.toLowerCase() + "' data-ipucu=\"" +
      ((CFG.aciklama||{})[r.k]||"").replace(/"/g,"&quot;") + "\">" + r.k + "</span></td>" +
    "<td>" + cokSatir(r.n) + "</td>" +
    "<td class='kod'>" + cokSatir(r.a) + "</td>" +
    "<td class='mv'>" + kacis(r.mv) + "</td>" +
    "<td class='og'>" + cokSatir(r.og) + "</td>" +
    "<td class='kod'" + ((r.alt || 1) > 1 ? " style='font-weight:600'" : "") + ">"
      + (r.alt || 1) + "</td>" +
    "<td class='oneri'>" + (r.d ? cokSatir(r.d) : "—") + "</td></tr>").join("")
    : ("<tr><td colspan='" + sutunSayisi + "' class='bos'>Eşleşen bulgu yok"
       + (kaynakF !== "ALL" ? " — <b>" + kacis(kaynakF) + "</b> kaynağında"
          + (filtre !== "ALL" ? " " + filtre + " kuralına ait" : "") + " kayıt bulunmuyor" : "")
       + "</td></tr>");
  const dahaBtn = document.getElementById("daha");
  if (dahaBtn) {
    dahaBtn.style.display = liste.length > limit ? "flex" : "none";
    if (liste.length > limit)
      dahaBtn.innerHTML = IK("ok", 14) + "Daha fazla göster ("
        + fmt(liste.length - limit) + " kayıt daha)";
  }
}
function git(s){
  document.querySelectorAll(".sayfa,#giris").forEach(e => e.classList.remove("acik"));
  document.getElementById(s).classList.add("acik");
  document.querySelectorAll(".sekmeler button[data-git]").forEach(
    x => x.classList.toggle("aktif", x.dataset.git === s));
  window.scrollTo(0, 0);
}
function filtreUygula(k){
  filtre = (filtre === k) ? "ALL" : k; limit = 60;
  document.querySelectorAll(".cip").forEach(c => c.classList.toggle("aktif", c.dataset.k === filtre));
  document.querySelectorAll(".minisut").forEach(x => x.classList.toggle("aktif", x.dataset.kural === filtre));
  document.querySelectorAll(".kuralSec").forEach(s => s.value = filtre);
  ciz();
}
document.addEventListener("change", e => {
  if (e.target.classList && e.target.classList.contains("kuralSec")) {
    filtre = e.target.value; limit = 60;
    document.querySelectorAll(".kuralSec").forEach(x => x.value = filtre);
    document.querySelectorAll(".minisut").forEach(x => x.classList.toggle("aktif", x.dataset.kural === filtre));
    document.querySelectorAll(".cip").forEach(c => c.classList.toggle("aktif", c.dataset.k === filtre));
    ciz();
  }
  if (e.target.classList && e.target.classList.contains("kaynakSec")) {
    kaynakF = e.target.value; limit = 60;
    document.querySelectorAll(".kaynakSec").forEach(x => x.value = kaynakF);
    ciz();
  }
  if (e.target.classList && e.target.classList.contains("alanSec")) {
    alanF = e.target.value; limit = 60;
    document.querySelectorAll(".alanSec").forEach(x => x.value = alanF);
    ciz();
  }
});
document.addEventListener("click", e => {
  const daha = e.target.closest("#daha");
  if (daha) { e.preventDefault(); limit += 200; ciz(); return; }
  const g = e.target.closest("[data-git]");
  if (g) { git(g.dataset.git); return; }
  const ms = e.target.closest(".minisut");
  if (ms) { filtreUygula(ms.dataset.kural); git("p3"); return; }
  const c = e.target.closest(".cip");
  if (c) { filtreUygula(c.dataset.k === filtre ? filtre : c.dataset.k); }
});
document.getElementById("ara").addEventListener("input", e => {
  arama = e.target.value; limit = 60;
  document.querySelectorAll(".soru").forEach(x => x.value = arama);
  ciz();
});
document.addEventListener("input", e => {
  if (e.target.classList && e.target.classList.contains("soru")) {
    arama = e.target.value; limit = 60;
    document.querySelectorAll(".soru").forEach(x => x.value = arama);
    const ar = document.getElementById("ara");
    if (ar) ar.value = arama;
    ciz();
  }
});
document.addEventListener("keydown", e => {
  if (e.key === "Enter" && e.target.classList &&
      e.target.classList.contains("soru")) git("p3");
});
document.querySelectorAll(".yazdir").forEach(x => x.addEventListener("click", () => window.print()));
document.querySelectorAll(".tema").forEach(x => x.addEventListener("click", () => {
  document.body.classList.toggle("gece");
  document.querySelectorAll(".tema").forEach(t =>
    t.textContent = document.body.classList.contains("gece") ? "Açık tema" : "Koyu tema");
}));

/* ================= TARAYICI İÇİ KURAL MOTORU ================= */
function trU(s){ return String(s??"").replace(/i/g,"İ").replace(/ı/g,"I").toUpperCase(); }
function norm(s){ return trU(String(s??"").trim().replace(/\s+/g," ")); }
function katla(s){ return s.replace(/İ/g,"I"); }
function turNorm(t){ const x=String(t??"").trim().replace(/^0+/,""); return x||"0"; }
function log(m){ const l=document.getElementById("log"); l.textContent+="\n"+m; l.scrollTop=l.scrollHeight; }

function basliklariOnar(kolonlar){
  return kolonlar.map(k=>{
    k=String(k).trim();
    if(CFG.en2tr[k]) return CFG.en2tr[k];
    if(k.includes("?")){
      const desen=new RegExp("^"+k.replace(/[.*+^${}()|[\]\\]/g,"\\$&").replace(/\?/g,".")+"$");
      const aday=CFG.kanonik.filter(c=>desen.test(c));
      if(aday.length===1) return aday[0];
    }
    return k;
  });
}
function dilAlgila(kolonlar){
  let en=0; for(const k of kolonlar) if(CFG.en2tr[String(k).trim()]) en++;
  return en>=3 ? "EN":"TR";
}
function tanimKolonu(kolonlar){
  for(const aday of ["Malzeme kısa metni","Malzeme tanımı","Malzeme kısa metinleri","Malzeme Tanımı"])
    if(kolonlar.includes(aday)) return aday;
  for(const k of kolonlar){
    const u=trU(k);
    if(u.includes("STANDART")) continue;
    if(u.includes("TANIM")||u.includes("METİN")||u.includes("METIN")||u.includes("DESCRIPTION")||u.includes("TEXT")) return k;
  }
  return null;
}
function dosyaOku(file){
  return new Promise((res,rej)=>{
    const r=new FileReader();
    r.onload=e=>{
      try{
        const wb=XLSX.read(e.target.result,{type:"array",codepage:1254});
        const ws=wb.Sheets[wb.SheetNames[0]];
        const ham=XLSX.utils.sheet_to_json(ws,{header:1,defval:""});
        if(!ham.length){res({kolonlar:[],satirlar:[]});return;}
        const kolonlar=basliklariOnar(ham[0]);
        const satirlar=ham.slice(1).filter(s=>s.some(h=>String(h).trim()!=="")).map(s=>{
          const o={}; kolonlar.forEach((k,i)=>o[k]=String(s[i]??"").trim()); return o;});
        res({kolonlar,satirlar,dil:dilAlgila(ham[0])});
      }catch(err){rej(err);}
    };
    r.onerror=()=>rej(new Error("Dosya okunamadı"));
    r.readAsArrayBuffer(file);
  });
}

function motor(urun,ust,dil,marm){
  const T = dil==="EN" ? {
    bosUst:"Parent code (base material) is empty", yokUst:"Parent code not found in list",
    uyusmaz:"Field value does not match parent code",
    barkod:"Barcode flag does not match parent code",
    esitle:(a,v)=>"'"+a+"' field must equal parent value: "+v,
    bagla:"Assign a valid parent code (base material)",
    kontrol:"Check the base material code — not found in parent list",
    sonda:(k,al)=>"Description ends with '"+k+"' but "+al+" field is empty or mismatched",
    gir:(al,k)=>al+" field: enter '"+k+"'",
    secili:(al,k)=>al+" field has '"+k+"' selected but missing at the end of the description",
    ekle:(k)=>"Append '"+k+"' to the end of the description",
    yasak:"Forbidden word in description", cikar:(y)=>"Remove from description: "+y,
    yasaksiz:"Description without forbidden words",
    altTarif:"Sub-descriptions of the parent (info — not an error)",
    altSorun:"Sub-descriptions of the parent — same description on multiple codes (info)",
    paylas:(n,l)=>"Codes sharing this description ("+n+"): "+l,
  } : {
    bosUst:"Üst kod (Temel malzeme) boş", yokUst:"Üst kod listede bulunamadı",
    uyusmaz:"Alan değeri üst kod ile uyuşmuyor",
    barkod:"Barkod işareti üst kod ile uyuşmuyor",
    esitle:(a,v)=>"'"+a+"' alanını üst kodun değeriyle eşitleyin: "+v,
    bagla:"Malzemeye geçerli bir üst kod (Temel malzeme) bağlayın",
    kontrol:"Temel malzeme alanındaki kodu kontrol edin — üst kod listesinde yok",
    sonda:(k,al)=>"Tanımın sonunda '"+k+"' var ama "+al+" alanı doldurulmamış veya uyuşmuyor",
    gir:(al,k)=>al+" alanına '"+k+"' girilmeli",
    secili:(al,k)=>al+" alanında '"+k+"' seçili ama tanımın sonunda yok",
    ekle:(k)=>"Tanımın sonuna '"+k+"' eklenmeli",
    yasak:"Tanımda yasaklı kelime var", cikar:(y)=>"Tanımdan şu ifadeleri çıkarın: "+y,
    yasaksiz:"Yasaklı kelimesiz tanım",
    altTarif:"Üst kodun alt tarifleri (bilgi — hata değil)",
    altSorun:"Üst kodun alt tarifleri — aynı tanım birden fazla kodda (bilgi)",
    paylas:(n,l)=>"Aynı tanımı paylaşan "+n+" kod: "+l,
  };
  // Kapsam filtresi
  const kapsam=(liste,gecerli)=>{
    const icin=[],disi=[];
    for(const s of liste)((gecerli.includes(turNorm(s["Malzeme türü"])))?icin:disi).push(s);
    return [icin,disi.length];
  };
  let kapsamDisi=0, x;
  [urun.satirlar,x]=kapsam(urun.satirlar,CFG.urunTur); kapsamDisi+=x;
  [ust.satirlar,x]=kapsam(ust.satirlar,CFG.ustTur); kapsamDisi+=x;

  const uTanim=tanimKolonu(urun.kolonlar), pTanim=tanimKolonu(ust.kolonlar);
  const tanimMap={};
  for(const s of urun.satirlar) if(uTanim) tanimMap[s["Malzeme"]]=s[uTanim]||"";
  for(const s of ust.satirlar) if(pTanim) tanimMap[s["Malzeme"]]=s[pTanim]||"";
  const ustKodMap={};
  for(const s of urun.satirlar) ustKodMap[s["Malzeme"]]=s["Temel malzeme"]||"";
  const turMap={};
  for(const s of urun.satirlar) turMap[s["Malzeme"]]=s["Malzeme türü"]||"";
  for(const s of ust.satirlar) turMap[s["Malzeme"]]=s["Malzeme türü"]||"";

  // --- Otomatik düzeltme günlüğü ---
  const duz=[];
  const duzEkle=(kaynak,m0,kural,alan,eski,yeni,durum,aciklama)=>
    duz.push({kaynak,m:String(m0),kural,alan,
              eski:eski==null?"":String(eski), yeni:yeni==null?"":String(yeni),
              durum,aciklama});

  const rows=[], setK={K1:new Set(),K2:new Set(),K3:new Set(),
                       K4:new Set(),K5:new Set(),K6:new Set()}, k4set=new Set();
  const kaySet=new Set(ust.satirlar.map(s=>String(s["Malzeme"])));
  const ekle=(t,m,u,k,n,a,mv,og,d)=>rows.push({t,m,u,k,n,a,mv,og,d,
    kay:kaySet.has(String(m))?"Üst Kod":"Ürün Kodu"});

  // ---- K1 ----
  const pIdx={}; for(const s of ust.satirlar) pIdx[norm(s["Malzeme"])]=s;
  const ortak=urun.kolonlar.filter(c=>ust.kolonlar.includes(c));
  const silmeOrtak=(()=>{
    for(const c of ortak){
      const u=trU(String(c));
      if((CFG.silmeAnahtar||[]).some(a=>u.indexOf(a)>=0)) return c;
    }
    return null;
  })();
  const kiyas=[];
  for(const c of CFG.kiyas){
    if(!ortak.includes(c)||c===uTanim||c===pTanim) continue;
    if(ust.satirlar.every(s=>norm(s[c])==="")){ log("  Bilgi: '"+c+"' üst kodlarda hiç dolu değil — K1 dışı."); continue; }
    kiyas.push(c);
  }
  const onekler=(CFG.ustkodsuzOnek||[]);
  for(const s of urun.satirlar){
    const m=s["Malzeme"], t=tanimMap[m]||"", temel=norm(s["Temel malzeme"]);
    const onekli=onekler.some(o=>String(m).trim().indexOf(o)===0);
    if(onekli){
      if(temel){
        ekle(t,m,s["Temel malzeme"],"K1",
          dil==="EN"?(onekler.join(" / ")+" prefixed material must not have a parent code")
                    :(onekler.join(" / ")+" ile başlayan malzemenin üst kodu olmamalı"),
          "Temel malzeme", s["Temel malzeme"],
          dil==="EN"?"(must be empty)":"(boş olmalı)",
          dil==="EN"?"Remove the parent code link":"Üst kod bağını kaldırın");
        setK.K1.add(m);
        duzEkle("Ürün Kodu",m,"K1","Temel malzeme",s["Temel malzeme"],"",
                "DÜZELTİLDİ","68 ile başlayan kodun üst kod bağı kaldırıldı");
      }
      continue;                      // 68'li kodlar alan karşılaştırmasına girmez
    }
    if(!temel){ ekle(t,m,"","K1",T.bosUst,"Temel malzeme","","",T.bagla); setK.K1.add(m);
      duzEkle("Ürün Kodu",m,"K1","Temel malzeme","","","ELLE",
              "Geçerli üst kod iş birimi tarafından belirlenmeli"); continue; }
    const p=pIdx[temel];
    if(!p){
      // RC formatındaki üst kodlar geçerli referanstır; listede olmasa da
      // "bulunamadı" bulgusu üretilmez (yalnızca RC dışı kodlar bildirilir).
      const rcOnek=trU(CFG.ustKodOneki||"RC");
      if(CFG.ustYoksaRcMuaf!==false && temel.indexOf(rcOnek)===0){ continue; }
      ekle(t,m,s["Temel malzeme"],"K1",T.yokUst,"Temel malzeme",s["Temel malzeme"],"",T.kontrol); setK.K1.add(m);
      duzEkle("Ürün Kodu",m,"K1","Temel malzeme",s["Temel malzeme"],"","ELLE",
              "Geçerli üst kod iş birimi tarafından belirlenmeli"); continue; }
    // Üst birim düzeyi silme işareti: üst kod işaretli ise ürün de işaretli olmalı
    if(silmeOrtak){
      const ui=norm(s[silmeOrtak]), pi=norm(p[silmeOrtak]);
      // Üst kod silinecekse ona bağlı TÜM ürünler bulgu verir
      if(pi!==""){
        ekle(t,m,s["Temel malzeme"],"K1",
          ui===""
            ? (dil==="EN"?("Parent code is flagged for deletion — this product must be deleted too, "
                          +"but the deletion flag is missing")
                        :("Üst kod silinmek üzere işaretli — bu ürün de silinmeli, "
                          +"ancak silme işareti konulmamış"))
            : (dil==="EN"?("Parent code is flagged for deletion — this product will also be deleted "
                          +"(flag set, deletion must be completed)")
                        :("Üst kod silinmek üzere işaretli — bu ürün de silinecek "
                          +"(işareti konulmuş, silme işlemi tamamlanmalı)")),
          silmeOrtak,
          ui==="" ? (dil==="EN"?"(not flagged)":"(işaretsiz)")
                  : (dil==="EN"?"X (flagged)":"X (işaretli)"),
          dil==="EN"?"X (flagged)":"X (işaretli)",
          dil==="EN"?"Set the deletion flag on the product as well"
                    :"Ürün kodunda da silme işaretini koyun");
        setK.K1.add(m);
        if(ui===""){
          duzEkle("Ürün Kodu",m,"K1",silmeOrtak,"","X","DÜZELTİLDİ",
                  "Üst kod silinecek — ürüne silme işareti konuldu");
        } else {
          duzEkle("Ürün Kodu",m,"K1",silmeOrtak,"X","X","ELLE",
                  "Silme işlemi SAP tarafında tamamlanmalı");
        }
      }
    }
    for(const alan of kiyas){
      let u=norm(s[alan]), v=norm(p[alan]);
      if(CFG.bos.includes(alan)&&(u===""||v==="")) continue;
      const bayrak=CFG.bayrak.includes(alan);
      if(bayrak){ u=u?"X":""; v=v?"X":""; }
      if(u!==v){
        const uy=bayrak?(u?"X (işaretli)":"(işaretsiz)"):s[alan];
        const vy=bayrak?(v?"X (işaretli)":"(işaretsiz)"):p[alan];
        ekle(t,m,s["Temel malzeme"],"K1",bayrak?T.barkod:T.uyusmaz,alan,uy,vy,T.esitle(alan,vy));
        setK.K1.add(m);
        duzEkle("Ürün Kodu",m,"K1",alan,s[alan],p[alan],"DÜZELTİLDİ",
                "Üst kodun değeriyle eşitlendi");
      }
    }
  }

  // ---- K2 ----
  const kodAlan={}, tek=new Set(), cift=new Set();
  const sonKat=CFG.sonKategori||Object.keys(CFG.k2);
  const yasakSon=(CFG.tanimSonuYasak||[]);
  for(const alan of sonKat){
    const kodlar=CFG.k2[alan]||{};
    for(const kod in kodlar){
      const k=trU(kod);
      if(yasakSon.indexOf(k)>=0) continue;   // tanım sonunda yasak; K1 bildirir
      (kodAlan[k]=kodAlan[k]||[]).push(alan);
      (k.includes(" ")?cift:tek).add(k);
    }
  }
  const alanDolu=(deger,kod,alan)=>{
    if(!deger) return false;
    const d=katla(deger), k=katla(kod);
    if(d===k||d.split(" ").includes(k)) return true;
    const metin=katla(trU(CFG.k2[alan][kod]||""));
    return !!metin&&d===metin;
  };
  const tumKod=[...tek,...cift].sort((a,b)=>b.length-a.length);
  const olcuSet=(CFG.olcuBirimi||[]);
  const bitisikBul=(parca)=>{
    for(const k of tumKod){
      if(olcuSet.indexOf(k)>=0) continue;   // "500ML" ölçü birimidir
      if(parca.length>k.length&&parca.endsWith(k)){
        const onceki=parca[parca.length-k.length-1];
        if(/[0-9]/.test(onceki)||CFG.bitisik.map(trU).indexOf(k)>=0) return k;
      }
    }
    return "";
  };
  const sonKodlar=(tanim)=>{
    const p=tanim.split(/[\s\-_/.]+/).filter(Boolean); const bul=[]; let i=p.length;
    while(i>1){
      if(i>=2&&cift.has(p[i-2]+" "+p[i-1])){bul.push(p[i-2]+" "+p[i-1]);i-=2;}
      else if(tek.has(p[i-1])){
        // "…500 ML" -> ML burada ölçü birimidir, kod değil
        if(olcuSet.indexOf(p[i-1])>=0&&i>=2&&p[i-2]&&/[0-9]$/.test(p[i-2])) break;
        bul.push(p[i-1]);i-=1;
      }
      else{
        const b=bitisikBul(p[i-1]);
        if(b) bul.push(b);
        break;
      }
    }
    return bul;
  };
  const yasakSonKod=(CFG.tanimSonuYasak||[]);
  const k2kaynak=(liste,tKol)=>{
    if(!tKol) return;
    for(const s of liste){
      const m=s["Malzeme"], ham=s[tKol]||"", tanim=norm(ham);
      if(!tanim) continue;
      // Tanım sonunda yasak kod (T26 vb.) — K2 bulgusu
      let yasakBulundu=false;
      for(const k of yasakSonKod){
        const re=new RegExp("(?:^|[\\s\\-_/.]|[0-9])"+k+"$");
        if(re.test(tanim)){
          const temiz=String(ham).replace(new RegExp("[\\s\\-_/.]*"+k+"\\s*$","i"),"").trim();
          ekle(ham,m,ustKodMap[m]||"","K2",
            dil==="EN"?("Description ends with '"+k+"' — this code cannot be used at the end")
                      :("Malzeme kısa metni '"+k+"' ile bitiyor — bu kod tanımın sonunda kullanılamaz"),
            dil==="EN"?"End of description":"Tanım sonu",
            ham, temiz,
            dil==="EN"?("Remove '"+k+"' from the end of the description")
                      :("Tanımın sonundaki '"+k+"' ifadesini kaldırın"));
          setK.K2.add(m);
          duzEkle(kaySet.has(String(m).trim())?"Üst Kod":"Ürün Kodu",m,"K2",tKol,ham,temiz,
                  "DÜZELTİLDİ","Tanım sonundaki yasak kod çıkarıldı");
          yasakBulundu=true;
          break;
        }
      }
      if(yasakBulundu) continue;
      const degerler={}; for(const a in CFG.k2) if(a in s) degerler[a]=norm(s[a]);
      const tail=sonKodlar(tanim);
      for(const kod of tail){
        const alanlar=(kodAlan[kod]||[]).filter(a=>a in degerler);
        if(!alanlar.length) continue;
        if(!alanlar.some(a=>alanDolu(degerler[a],kod,a))){
          const al=alanlar.join(" / ");
          ekle(ham,m,ustKodMap[m]||"","K2",T.sonda(kod,al),al,ham,kod,T.gir(al,kod));
          setK.K2.add(m);
          duzEkle(kaySet.has(String(m).trim())?"Üst Kod":"Ürün Kodu",m,"K2",
                  alanlar[0],s[alanlar[0]]||"",kod,"DÜZELTİLDİ",
                  "Tanımdaki koda göre alan dolduruldu");
        }
      }
      if(CFG.ters){
        for(const alan of CFG.tersAlan){
          const d=degerler[alan]; if(!d) continue;
          for(const kod in CFG.k2[alan]||{}){
            const k=trU(kod);
            if((CFG.tanimSonuYasak||[]).indexOf(k)>=0) continue;
            if(alanDolu(d,k,alan)&&!tail.includes(k)){
              ekle(ham,m,ustKodMap[m]||"","K2",T.secili(alan,k),alan,ham,ham.trim()+" "+k,T.ekle(k));
              setK.K2.add(m);
              duzEkle(kaySet.has(String(m).trim())?"Üst Kod":"Ürün Kodu",m,"K2",tKol,
                      ham,ham.trim()+" "+k,"DÜZELTİLDİ","Tanım programın önerdiği hâle getirildi");
              break;
            }
          }
        }
      }
    }
  };
  k2kaynak(urun.satirlar,uTanim); k2kaynak(ust.satirlar,pTanim);

  // ---- K3 ----
  const k3kaynak=(liste,tKol)=>{
    if(!tKol) return;
    for(const s of liste){
      const m=s["Malzeme"], ham=s[tKol]||"", tanim=norm(ham);
      if(!tanim) continue;
      const bulunan=[];
      for(const kelime of CFG.yasakli){
        const k=trU(kelime);
        if(CFG.yasakliSonda.includes(kelime)){
          const p=tanim.split(/\s+/); if(p.length&&p[p.length-1]===k) bulunan.push(kelime);
        }else if(/^\d+$/.test(k)){
          if(tanim.split(/\s+/).includes(k)) bulunan.push(kelime);
        }else{
          const re=new RegExp("(^|[^A-Z0-9_ÇĞİÖŞÜ])"+k+"($|[^A-Z0-9_ÇĞİÖŞÜ])");
          if(re.test(tanim)) bulunan.push(kelime);
        }
      }
      if(bulunan.length){
        let temizT=String(ham);
        for(const kelime of bulunan){
          const kk=trU(kelime);
          if((CFG.yasakliSonda||[]).includes(kelime))
            temizT=temizT.replace(new RegExp("[\\s\\-_/.]*"+kk+"\\s*$","i"),"");
          else
            temizT=temizT.replace(new RegExp("(^|[^A-Za-z0-9ÇĞİÖŞÜçğıöşü])"+kk+"($|[^A-Za-z0-9ÇĞİÖŞÜçğıöşü])","ig"),"$1$2");
        }
        temizT=temizT.replace(/\s{2,}/g," ").trim();
        ekle(ham,m,ustKodMap[m]||"","K3",T.yasak,dil==="EN"?"Description":"Tanım",ham,temizT||T.yasaksiz,T.cikar(bulunan.join(", ")));
        setK.K3.add(m);
        if(temizT&&temizT!==ham)
          duzEkle(kaySet.has(String(m).trim())?"Üst Kod":"Ürün Kodu",m,"K3",tKol,ham,temizT,
                  "DÜZELTİLDİ","Yasaklı kelime tanımdan çıkarıldı");
      }
    }
  };
  k3kaynak(urun.satirlar,uTanim); k3kaynak(ust.satirlar,pTanim);

  // ---- K4: çokluk (xN) ↔ MARM adet birimi sayacı ----
  const cokluk=(tanim)=>{
    const p=String(tanim||"").split(/[\s\-_/.]+/).filter(Boolean);
    let i=p.length;
    while(i>1){
      if(i>=2&&cift.has(p[i-2]+" "+p[i-1])){i-=2;continue;}
      if(tek.has(p[i-1])){
        if(olcuSet.indexOf(p[i-1])>=0&&i>=2&&p[i-2]&&/[0-9]$/.test(p[i-2])) break;
        i-=1;continue;
      }
      break;
    }
    if(i<1) return null;
    let son=p[i-1];
    const b=bitisikBul(son);
    if(b) son=son.slice(0,son.length-b.length);
    const m2=son.match(/[X*](\d{1,4})$/);
    return m2?parseInt(m2[1],10):null;
  };
  if(marm&&marm.satirlar&&marm.satirlar.length){
    const haricB=(CFG.haricBirim||[]);
    const kolonBul=(kolonlar,anahtarlar,haric)=>{
      for(const c of kolonlar){
        const u=trU(String(c));
        if(haric&&haric.some(x=>u.indexOf(x)>=0)) continue;
        if(anahtarlar.some(a=>u.indexOf(a)>=0)) return c;
      }
      return null;
    };
    const kM=kolonBul(marm.kolonlar,["MALZEME","MATERIAL","MATNR"]);
    const kB=kolonBul(marm.kolonlar,["ALTERNATİF","ALTERNATIF","AÖB","MEINH",
      "ÖLÇÜ BİRİMİ","OLCU BIRIMI","UOM","BİRİM","BIRIM"],["TEMEL","BASE"]);
    const kS=kolonBul(marm.kolonlar,["SAYAÇ","SAYAC","UMREZ","NUMERATOR"],["PAYDA"]);
    const kP=kolonBul(marm.kolonlar,["PAYDA","UMREN","DENOMINATOR"]);
    if(kM&&kB&&kS){
      // Her MARM satırı bir çevrim faktörü taşır: "1 koli = 12 adet" -> 12.
      // Adet satırının sayacı 1 olduğu için ona bakılmaz; ağırlık/hacim
      // birimleri tamamen dışarıda tutulur.
      // MARM'da o malzeme için geçen SAYAÇ değerleri kümesi
      const coklukMap={}, birimMap={};
      const sayi=(v)=>{
        const f=parseFloat(String(v==null?"":v).replace(",","."));
        return isNaN(f)?null:f;
      };
      for(const s of marm.satirlar){
        const birim=norm(s[kB]);
        if(haricB.indexOf(birim)>=0) continue;
        const sayac=sayi(s[kS]);
        if(sayac===null||sayac<=0) continue;
        if(Math.abs(sayac-Math.round(sayac))>1e-6) continue;
        const d=Math.round(sayac);
        const kod=String(s[kM]).trim();
        (coklukMap[kod]=coklukMap[kod]||new Set()).add(d);
        birimMap[kod]=birimMap[kod]||{};
        if(!birimMap[kod][d]) birimMap[kod][d]=birim;
      }
      // MARA ürün hiyerarşisi muafiyeti — bu kodlar K4 dışıdır.
      const muafHiy=new Set(CFG.k4MuafHiyerarsiler||[]);
      const hiyNorm=(v)=>{ let x=String(v==null?"":v).trim();
        if(x.slice(-2)===".0") x=x.slice(0,-2); return x; };
      const hiyKolon=(muafHiy.size?(function(){
        for(const a of (CFG.hiyerarsiAdaylari||[]))
          if(urun.kolonlar.indexOf(a)>=0) return a;
        for(const c of urun.kolonlar){ const u=trU(String(c));
          if(u.indexOf("HİYERARŞİ")>=0||u.indexOf("HIYERARSI")>=0
            ||u.indexOf("PRDHA")>=0||u.indexOf("PRODUCT HIERARCHY")>=0) return c; }
        return null; })():null);
      let dogrulanamayan=0, muafSay=0;
      for(const s of urun.satirlar){
        if(!uTanim) break;
        const m0=String(s["Malzeme"]).trim();
        if(hiyKolon && muafHiy.has(hiyNorm(s[hiyKolon]))){ muafSay++; continue; }
        const n=cokluk(norm(s[uTanim]));
        if(n===null) continue;
        const gecerli=coklukMap[m0];
        if(!gecerli||!gecerli.size){ dogrulanamayan++; continue; }
        if(!gecerli.has(n)){
          const bek=[...gecerli].sort((a,b)=>a-b);
          const bekMetin=bek.map(d=>"x"+d).join(", ");
          const bekBirim=bek.map(d=>(birimMap[m0][d]||"")+"="+d).join(", ");
          ekle(s[uTanim],m0,ustKodMap[m0]||"","K4",
            dil==="EN"?"Pack size does not match any MARM numerator value"
                      :"Tanımdaki çokluk, MARM'daki hiçbir sayaç değeriyle eşleşmiyor",
            dil==="EN"?"Pack size (xN) / MARM":"Çokluk (xN) / MARM sayaç",
            (dil==="EN"?"description: x":"tanım: x")+n,
            bekMetin+" ("+bekBirim+")",
            dil==="EN"?("MARM values: "+bekMetin+" — align the description with one of them or fix MARM")
                      :("MARM'da tanımlı sayaç değerleri: "+bekMetin+" — tanımı bunlardan biriyle eşitleyin veya MARM kaydını düzeltin"));
          setK.K4.add(m0);
          if(bek.length===1){
            const yeniT=String(s[uTanim]).replace(/([Xx*])(\d{1,4})\s*$/, "$1"+bek[0]);
            duzEkle("Ürün Kodu",m0,"K4",uTanim,s[uTanim],yeniT,"DÜZELTİLDİ",
                    "Çokluk MARM'daki tek sayaç değeriyle eşitlendi");
          } else {
            duzEkle("Ürün Kodu",m0,"K4","Çokluk (xN)","x"+n,bekMetin,"ELLE",
                    "MARM'da birden çok sayaç var — doğrusu iş birimince seçilmeli");
          }
        }
      }
      log("  K4: MARM çevrim faktörleriyle karşılaştırıldı · doğrulanamayan "+dogrulanamayan
        + (hiyKolon ? " · muaf hiyerarşi "+muafSay : ""));
    } else {
      log("  K4 atlandı: MARM sütunları tanınamadı ("+marm.kolonlar.slice(0,6).join(", ")+"…)");
    }
  }

  // ---- K5: ayırt edici alanlar + koli içi adet kombinasyonu benzersiz olmalı ----
  {
    const alanlar=(CFG.k5Alanlar||[]).filter(c=>ust.kolonlar.indexOf(c)>=0);
    if(alanlar.length){
      // MARM sayaç kümeleri
      const sayacKume={};
      if(marm&&marm.satirlar&&marm.satirlar.length){
        const kb=(kolonlar,anahtarlar,haric)=>{
          for(const c of kolonlar){
            const u=trU(String(c));
            if(haric&&haric.some(x=>u.indexOf(x)>=0)) continue;
            if(anahtarlar.some(a=>u.indexOf(a)>=0)) return c;
          }
          return null;
        };
        const mM=kb(marm.kolonlar,["MALZEME","MATERIAL","MATNR"]);
        const mS=kb(marm.kolonlar,["SAYAÇ","SAYAC","UMREZ","NUMERATOR"],["PAYDA"]);
        if(mM&&mS){
          for(const s of marm.satirlar){
            const v=parseFloat(String(s[mS]).replace(",","."));
            if(isNaN(v)||v<=0) continue;
            const kod=String(s[mM]).trim();
            (sayacKume[kod]=sayacKume[kod]||new Set()).add(Math.round(v));
          }
        }
      }
      // Üst koda bağlı ürün (alt tarif) sayısı
      const altSayi={};
      for(const s of urun.satirlar){
        const uk=String(s["Temel malzeme"]||"").trim();
        if(!uk) continue;
        (altSayi[uk]=altSayi[uk]||new Set()).add(String(s["Malzeme"]).trim());
      }
      const imzaMap={}; let dogrulanamayan=0;
      for(const s of ust.satirlar){
        const kod=String(s["Malzeme"]).trim();
        const n=pTanim?cokluk(norm(s[pTanim])):null;
        const kume=sayacKume[kod];
        // Koli içi adet: tanımdaki çokluk MARM sayacında da varsa doğrulanmış
        if(n===null||!kume||!kume.has(n)){ dogrulanamayan++; continue; }
        const parcalar=alanlar.map(c=>norm(s[c]));
        parcalar.push(String(n));
        const imza=parcalar.join("|");
        (imzaMap[imza]=imzaMap[imza]||new Set()).add(kod);
      }
      if(dogrulanamayan)
        log("  Koli içi adedi doğrulanamayan üst kod: "+dogrulanamayan+" (K5 dışı)");
      for(const imza in imzaMap){
        const kodlar=[...imzaMap[imza]].sort();
        if(kodlar.length<2) continue;
        const parcalar=imza.split("|");
        let okunur=alanlar.map((a,i)=>parcalar[i]?a+"="+parcalar[i]:"")
          .filter(Boolean).join("; ");
        okunur+="; koli içi adet="+parcalar[parcalar.length-1];
        for(const kod of kodlar){
          const digerleri=kodlar.filter(k=>k!==kod);
          const altN=(altSayi[kod]||new Set()).size;
          const digerAlt=digerleri.map(k=>k+"="+((altSayi[k]||new Set()).size)).join(", ");
          ekle(tanimMap[kod]||"",kod,digerleri.join(", "),"K5",
            dil==="EN"?("Parent code shares all "+(alanlar.length+1)
                        +" distinguishing components with another parent code")
                      :("Üst kodun ayırt edici "+(alanlar.length+1)
                        +" bileşeni başka bir üst kodla birebir aynı"),
            dil==="EN"?"Distinguishing attributes + units per case"
                      :(alanlar.join(" / ")+" + koli içi adet"),
            okunur,
            dil==="EN"?"Unique combination per parent code"
                      :"Her üst kod için benzersiz kombinasyon",
            dil==="EN"?("Same combination on "+kodlar.length+" parent codes — this one has "
                        +altN+" sub-descriptions; matches: "+digerAlt)
                      :("Aynı kombinasyona sahip "+kodlar.length+" üst kod — bu kodun "
                        +altN+" alt tarifi var; eşleşenler: "+digerAlt
                        +". Birini kaldırın veya ayırt edici alanları farklılaştırın"));
          setK.K5.add(kod);
          duzEkle("Üst Kod",kod,"K5","Ayırt edici alanlar",okunur,"","ELLE",
                  "Aynı kombinasyon: "+digerleri.join(", ")+" — hangi kodun kalacağı iş birimince belirlenmeli");
        }
      }
    }
  }

  // ---- K6: ürün kodunun son iki hanesi ↔ Menşei ----
  {
    const menKod = (CFG.k2 && CFG.k2["Menşei"]) ? CFG.k2["Menşei"] : {};
    const gecerli = new Set(Object.keys(menKod).map(k=>String(k).trim()));
    const bildirGecersiz = CFG.k6GecersizBildir === true;
    const menCevir = (v)=>{
      let h = String(v==null?"":v).trim();
      if(h.slice(-2)===".0") h = h.slice(0,-2);
      if(!h) return "";
      if(gecerli.has(h)) return h;
      if(/^\d$/.test(h) && gecerli.has("0"+h)) return "0"+h;
      const hedef = katla(trU(h));
      for(const k in menKod){
        const m = menKod[k];
        if(m && katla(trU(String(m)))===hedef) return String(k);
      }
      return h;
    };
    const metin = (k)=>{ const a=menKod[k]; return a ? k+" ("+a+")" : String(k); };
    let tanimsiz = 0;
    if(CFG.k6Aktif !== false && urun.kolonlar.indexOf("Menşei") >= 0){
      for(const s of urun.satirlar){
        const kod = String(s["Malzeme"]).trim();
        if(!/^\d{3,}$/.test(kod)) continue;      // yalnızca sayısal ürün kodu
        const son2 = kod.slice(-2);
        if(!gecerli.has(son2)){
          tanimsiz++;
          if(!bildirGecersiz) continue;
          ekle(uTanim?s[uTanim]:"",kod,ustKodMap[kod]||"","K6",
            dil==="EN"?"Last two digits of the code are not a defined origin code"
                      :"Malzeme kodunun son iki hanesi tanımlı bir menşei kodu değil",
            dil==="EN"?"Origin / last 2 digits":"Menşei / Malzeme kodu son 2 hane",
            (dil==="EN"?"origin: ":"menşei: ")+(menCevir(s["Menşei"])||"(boş)")
              +(dil==="EN"?" · code end: ":" · kod sonu: ")+son2,
            dil==="EN"?"(defined origin code)":"(tanımlı menşei kodu)",
            dil==="EN"?"Check the material code or the origin code list"
                      :"Malzeme kodunu veya menşei kod listesini kontrol edin");
          setK.K6.add(kod);
          duzEkle("Ürün Kodu",kod,"K6","Menşei",s["Menşei"]||"","","ELLE",
                  "Kodun son iki hanesi tanımlı menşei değil");
          continue;
        }
        const men = menCevir(s["Menşei"]);
        if(men===""){
          ekle(uTanim?s[uTanim]:"",kod,ustKodMap[kod]||"","K6",
            dil==="EN"?"Origin field is empty — the code's last two digits are set"
                      :"Menşei alanı boş — malzeme kodunun son iki hanesi dolu",
            dil==="EN"?"Origin / last 2 digits":"Menşei / Malzeme kodu son 2 hane",
            (dil==="EN"?"origin: (empty) · code end: ":"menşei: (boş) · kod sonu: ")+son2,
            metin(son2),
            (dil==="EN"?"Enter the code's last two digits into the Origin field: "
                      :"Menşei alanına kodun son iki hanesini girin: ")+metin(son2));
          setK.K6.add(kod);
          duzEkle("Ürün Kodu",kod,"K6","Menşei","",son2,"DÜZELTİLDİ",
                  "Menşei kodun son iki hanesiyle dolduruldu");
        } else if(men!==son2){
          ekle(uTanim?s[uTanim]:"",kod,ustKodMap[kod]||"","K6",
            dil==="EN"?"Origin does not match the last two digits of the material code"
                      :"Menşei, malzeme kodunun son iki hanesiyle uyuşmuyor",
            dil==="EN"?"Origin / last 2 digits":"Menşei / Malzeme kodu son 2 hane",
            (dil==="EN"?"origin: ":"menşei: ")+metin(men)
              +(dil==="EN"?" · code end: ":" · kod sonu: ")+son2,
            metin(son2),
            (dil==="EN"?"Align the Origin field with the code's last two digits: "
                      :"Menşei alanını kodun son iki hanesiyle eşitleyin: ")+metin(son2)
              +(dil==="EN"?" — or review the material code"
                         :" — ya da malzeme kodunu gözden geçirin"));
          setK.K6.add(kod);
          duzEkle("Ürün Kodu",kod,"K6","Menşei",s["Menşei"]||"",son2,"DÜZELTİLDİ",
                  "Menşei kodun son iki hanesiyle eşitlendi");
        }
      }
      if(tanimsiz && !bildirGecersiz)
        log("  K6: "+tanimsiz+" üründe kodun son iki hanesi tanımlı menşei değil — K6 dışı");
    } else if(CFG.k6Aktif !== false){
      log("  K6 atlandı: 'Menşei' sütunu bulunamadı.");
    }
  }

  // ---- Alt tarifler (bilgi) ----
  const k4kaynak=(liste,tKol)=>{
    if(!tKol) return;
    const grup={};
    for(const s of liste){ const n=norm(s[tKol]); if(n)(grup[n]=grup[n]||[]).push(s); }
    for(const n in grup){
      const g=grup[n]; const kodlar=[...new Set(g.map(s=>String(s["Malzeme"])))].sort();
      if(kodlar.length<2) continue;
      for(const s of g){
        const m=s["Malzeme"];
        ekle(s[tKol],m,ustKodMap[m]||"","ALT",T.altTarif,dil==="EN"?"Description":"Tanım",s[tKol],"—",T.paylas(kodlar.length,kodlar.join(", ")));
        k4set.add(m);
      }
    }
  };
  k4kaynak(urun.satirlar,uTanim); k4kaynak(ust.satirlar,pTanim);

  // ÜB dzy.silme iştr. dolu olan malzemeler silinmiş sayılır: hiçbir
  // kuralda bulgu üretmezler ve sağlık hesabına girmezler.
  const silinmisSet=new Set();
  [[urun,"urun"],[ust,"ust"]].forEach(([tablo])=>{
    let kolon=null;
    for(const c of tablo.kolonlar){
      const u=trU(String(c));
      if((CFG.silmeAnahtar||[]).some(a=>u.indexOf(a)>=0)){ kolon=c; break; }
    }
    if(!kolon) return;
    for(const s of tablo.satirlar)
      if(norm(s[kolon])!=="") silinmisSet.add(String(s["Malzeme"]).trim());
  });
  if(silinmisSet.size){
    const oncesi=rows.length;
    // Silme kaskadı bulguları muaf: ürün işaretli olsa da takip gerekir
    const kaskadMi=r0=>{
      const a=trU(String(r0.a||""));
      return a.indexOf("SİLME")>=0 || a.indexOf("SILME")>=0;
    };
    const kalan=rows.filter(r0=>!silinmisSet.has(String(r0.m).trim())||kaskadMi(r0));
    rows.length=0; rows.push(...kalan);
    ["K1","K2","K3","K4","K5","K6"].forEach(k=>{
      [...setK[k]].forEach(m0=>{ if(silinmisSet.has(String(m0).trim())) setK[k].delete(m0); });
    });
    const duzKalan=duz.filter(d=>
      !silinmisSet.has(String(d.m).trim()) ||
      (d.alan && trU(String(d.alan)).indexOf("SİLME")>=0) ||
      (d.alan && trU(String(d.alan)).indexOf("SILME")>=0));
    duz.length=0; duz.push(...duzKalan);
    log("  "+silinmisSet.size+" malzeme silinmiş işaretli — "+(oncesi-rows.length)+" bulgu düşüldü.");
  }

  // Aynı malzemedeki birden çok K2 bulgusu TEK satırda birleştirilir
  // (örn. "…24RU LF" -> 1) LF/Ek Alan  2) RU/Pazar)
  const k2grup={}, k2sira=[];
  const digerRows=[];
  for(const r0 of rows){
    if(r0.k!=="K2"){ digerRows.push(r0); continue; }
    if(!k2grup[r0.m]){ k2grup[r0.m]=[]; k2sira.push(r0.m); }
    k2grup[r0.m].push(r0);
  }
  const k2birlesik=k2sira.map(mm=>{
    const g=k2grup[mm];
    if(g.length===1) return g[0];
    const num=(liste)=>liste.map((v,i)=>(i+1)+") "+v).join("\n");
    const b=Object.assign({},g[0]);
    b.n=num(g.map(x=>x.n));
    b.a=num(g.map(x=>x.a));
    b.og=num(g.map(x=>x.og));
    b.d=num(g.map(x=>x.d));
    return b;
  });
  rows.length=0;
  rows.push(...digerRows, ...k2birlesik);

  // Alt tarif sayısı: aynı tanımı paylaşan kod adedi (K4 bilgisi sütuna taşınır)
  const altMap={};
  for(const r0 of rows) if(r0.k==="ALT"){
    const n=(r0.d.match(/(\d+)/)||[0,1])[1];
    altMap[r0.m]=parseInt(n,10)||1;
  }
  for(const r0 of rows) r0.alt=altMap[r0.m]||1;

  // ---- Özet ----
  const say=k=>rows.filter(r=>r.k===k).length;
  const malz=k=>new Set(rows.filter(r=>r.k===k).map(r=>r.m)).size;
  const tekilKod=new Set();
  urun.satirlar.forEach(s=>tekilKod.add(String(s["Malzeme"]).trim()));
  ust.satirlar.forEach(s=>tekilKod.add(String(s["Malzeme"]).trim()));
  silinmisSet.forEach(k=>tekilKod.delete(k));
  const toplamM=tekilKod.size;
  const hataliSet=new Set([...setK.K1,...setK.K2,...setK.K3,
                           ...setK.K4,...setK.K5,...setK.K6]);
  const marka={}, tur={};
  for(const r0 of rows){
    const mk=(tanimMap[r0.m]||"").split(" ")[0];
    if(mk && r0.k!=="ALT"){
      (marka[mk]=marka[mk]||{K1:0,K2:0,K3:0,K4:0,K5:0,K6:0});
      if(marka[mk][r0.k]===undefined) marka[mk][r0.k]=0;
      marka[mk][r0.k]++;
    }
    const tt=String(turMap[r0.m]||"");
    (tur[r0.k]=tur[r0.k]||{}); tur[r0.k][tt]=(tur[r0.k][tt]||0)+1;
  }
  const etki=[];
  const tumK=["K1","K2","K3","K4","K5","K6"];
  for(const kk of tumK){
    let sadece=0;
    for(const m of setK[kk]){
      if(tumK.every(o=>o===kk||!setK[o].has(m))) sadece++;
    }
    etki.push([kk,100*(toplamM-hataliSet.size+sadece)/Math.max(toplamM,1)]);
  }
  const hazir=rows.filter(r=>r.k==="K2"&&r.og&&r.og!==r.mv&&r.og.length>r.a.length&&r.d&&(r.d.includes("sonuna")||r.d.includes("Append"))).length;
  const matlar=[]; const gorulen=new Set();
  for(const s of urun.satirlar){
    const mk=String(s["Malzeme"]).trim();
    if(gorulen.has(mk)||silinmisSet.has(mk)) continue; gorulen.add(mk);
    matlar.push({kay:"Ürün Kodu",m:mk,t:tanimMap[s["Malzeme"]]||""});
  }
  for(const s of ust.satirlar){
    const mk=String(s["Malzeme"]).trim();
    if(gorulen.has(mk)||silinmisSet.has(mk)) continue; gorulen.add(mk);
    matlar.push({kay:"Üst Kod",m:mk,t:tanimMap[s["Malzeme"]]||""});
  }
  return {rows,dil,kapsamDisi,toplamM,matlar,setK,k4set,duz,
    _urun:urun.satirlar,_ust:ust.satirlar,
    urunN:matlar.filter(x=>x.kay==="Ürün Kodu").length,
    ustN:matlar.filter(x=>x.kay==="Üst Kod").length,
    dogru:toplamM-hataliSet.size,hatali:hataliSet.size,
    sayilar:{K1:[say("K1"),malz("K1")],K2:[say("K2"),malz("K2")],
             K3:[say("K3"),malz("K3")],K4:[say("K4"),malz("K4")],
             K5:[say("K5"),malz("K5")],K6:[say("K6"),malz("K6")]},
    marka,tur,etki,hazir};
}

/* ---- Panelleri yeniden çiz ---- */
function fmt(n){ return n.toLocaleString("tr-TR"); }
/* Kural listesi — panel çizimi ve Excel dışa aktarımı ortak kullanır */
const KL=["K1","K2","K3","K4","K5","K6"];
function panoGuncelle(S){
  const yuzde=S.toplamM?100*S.dogru/S.toplamM:0;
  const toplamBulgu=Object.values(S.sayilar).reduce((a,b)=>a+b[0],0);
  const adlar={K1:"Akıllı kod eşleşmesi",K2:"Tanım ↔ ek veri alanı",
               K3:"Yasaklı kelimeler",K4:"Çokluk ↔ MARM sayacı",
               K5:"Üst kod benzersizliği",K6:"Kod sonu ↔ Menşei"};
  // KPI şeridi
  const ckpi=document.getElementById("c_kpi");
  if(ckpi) ckpi.innerHTML=
    '<div class="kpi ana"><div class="kpi-ust"><span>Veri Sağlığı</span>'+IK("saglik",17,"ikn kpi-ikon")+'</div>'
      +'<b>%'+yuzde.toFixed(1)+'</b><div class="kpi-ray"><div style="width:'+yuzde.toFixed(1)+'%"></div></div>'
      +'<em>kurallardan geçen malzeme oranı</em></div>'
    +'<div class="kpi uyari"><div class="kpi-ust"><span>Toplam Bulgu</span>'+IK("bulgu",16,"ikn kpi-ikon")+'</div>'
      +'<b>'+fmt(toplamBulgu)+'</b><em>hata + bilgi kaydı</em></div>'
    +'<div class="kpi"><div class="kpi-ust"><span>Hatalı Malzeme</span>'+IK("liste",16,"ikn kpi-ikon")+'</div>'
      +'<b>'+fmt(S.hatali)+'</b><em>en az bir kurala takılan</em></div>'
    +'<div class="kpi"><div class="kpi-ust"><span>Taranan Malzeme</span>'+IK("malzeme",16,"ikn kpi-ikon")+'</div>'
      +'<b>'+fmt(S.toplamM)+'</b><em>'+fmt(S.urunN)+' ürün + '+fmt(S.ustN)+' üst kod</em></div>'
    +'<div class="kpi"><div class="kpi-ust"><span>Hazır Düzeltme</span>'+IK("onay",16,"ikn kpi-ikon")+'</div>'
      +'<b>'+fmt(S.hazir)+'</b><em>program önerisi (K2)</em></div>';
  // Alan filtresi seçeneklerini veriye göre yenile
  const alanKume=new Set();
  S.rows.filter(r=>r.k!=="ALT"&&r.a).forEach(r=>
    String(r.a).split(/\s*[;/]\s*/).forEach(p=>{ if(p.trim()) alanKume.add(p.trim()); }));
  const alanlar=[...alanKume].sort();
  document.querySelectorAll(".alanSec").forEach(sel=>{
    const eski=sel.value;
    sel.innerHTML='<option value="ALL">Tümü</option>'
      + alanlar.map(a=>'<option value="'+kacis(a)+'">'+kacis(a)+'</option>').join("");
    sel.value = alanlar.includes(eski) ? eski : "ALL";
    if(sel.value==="ALL") alanF="ALL";
  });

  // Veri kalitesi boyutları
  const cb=document.getElementById("c_boyut");
  if(cb){
    const kos=[
      ["Tamlık","Zorunlu alan ve üst kod bağı dolu mu",
        r=>(r.k==="K1"&&r.a==="Temel malzeme")||(r.k==="K2"&&/doldurulmam|empty/i.test(r.n))],
      ["Tutarlılık","Ürün alanları üst koduyla uyumlu mu",
        r=>r.k==="K1"&&r.a!=="Temel malzeme"],
      ["Geçerlilik","Tanım ile ek veri alanı örtüşüyor mu", r=>r.k==="K2"],
      ["Standart","Tanım adlandırma kuralına uygun mu", r=>r.k==="K3"],
      ["Doğruluk","Tanımdaki çokluk MARM sayacıyla örtüşüyor mu", r=>r.k==="K4"],
      ["Benzersizlik","Her üst kodun ayırt edici kombinasyonu tekil mi",
        r=>r.k==="K5"],
      ["Uyum","Kodun son iki hanesi menşei ile örtüşüyor mu", r=>r.k==="K6"],
    ];
    const renk=p=>p>=95?"#01B8AA":(p>=85?"#E3A63B":"#FD625E");
    cb.innerHTML='<div class="tbaslik">Veri Kalitesi Boyutları'
      +'<small>Her boyut 100 üzerinden · yeşil ≥95 · sarı ≥85</small></div>'
      +kos.map(([ad,ac,f])=>{
        const n=new Set(S.rows.filter(f).map(x=>x.m)).size;
        const p=S.toplamM?100*(1-n/S.toplamM):100;
        return '<div class="boyut"><div class="bsatir"><span class="bad">'+ad+'</span>'
          +'<span class="bpuan" style="color:'+renk(p)+'">'+p.toFixed(1)+'</span></div>'
          +'<div class="bray"><div style="width:'+p.toFixed(1)+'%;background:'+renk(p)+'"></div></div>'
          +'<div class="baciklama">'+ac+' · '+fmt(n)+' malzeme etkilendi</div></div>';
      }).join("");
  }
  // En sorunlu alanlar
  const ca=document.getElementById("c_alan");
  if(ca){
    const say={};
    for(const x of S.rows) if(x.k!=="K4"&&x.a) say[x.a]=(say[x.a]||0)+1;
    const ilk=Object.entries(say).sort((a,b)=>b[1]-a[1]).slice(0,6);
    const enA=ilk.length?ilk[0][1]:1;
    ca.innerHTML='<div class="tbaslik">En Sorunlu Alanlar'
      +'<small>Bulgu sayısına göre ilk 6 · aksiyon önceliği</small></div>'
      +ilk.map(([a,v])=>'<div class="alanb"><div class="asatir"><span>'+kacis(a)+'</span>'
        +'<span class="mono">'+fmt(v)+'</span></div>'
        +'<div class="aray"><div style="width:'+(100*v/enA).toFixed(0)+'%"></div></div></div>').join("");
  }
  // Kural özet tablosu
  const ctab=document.getElementById("c_tablo");
  if(ctab){
    const cv=2*Math.PI*8;
    ctab.innerHTML='<div class="tbaslik">Kural Özeti'
      +'<small>Bulgu · etkilenen malzeme · toplam içindeki pay</small></div>'
      +'<table class="ozet-tablo"><thead><tr><th>Kural</th><th style="text-align:right">Bulgu</th>'
      +'<th style="text-align:right">Malzeme</th><th style="text-align:right">Pay</th></tr></thead><tbody>'
      +KL.map((k,i)=>{
        const pay=toplamBulgu?100*S.sayilar[k][0]/toplamBulgu:0;
        return '<tr><td><span class="khap '+k.toLowerCase()+'" style="background:'+CFG.renk[k]+'">'+k+'</span> '+adlar[k]+'</td>'
          +'<td class="say">'+fmt(S.sayilar[k][0])+'</td><td class="say">'+fmt(S.sayilar[k][1])+'</td>'
          +'<td class="say"><svg class="minidonut" width="20" height="20" viewBox="0 0 20 20">'
          +'<circle cx="10" cy="10" r="8" fill="none" class="iz" stroke-width="4"/>'
          +'<circle cx="10" cy="10" r="8" fill="none" stroke="'+CFG.renk[k]+'" stroke-width="4" stroke-dasharray="'
          +(cv*pay/100).toFixed(1)+' '+cv.toFixed(1)+'" stroke-dashoffset="'+(cv/4).toFixed(1)+'"/></svg> %'+pay.toFixed(0)+'</td></tr>';
      }).join("")+'</tbody></table>';
  }
  // kural mini
  const ck=document.getElementById("c_kural");
  if(ck){
    const enB=Math.max(...Object.values(S.sayilar).map(v=>v[0]),1);
    const kutu=ck.parentElement;
    const lej=kutu&&kutu.querySelector(".lejlist");
    if(lej) lej.innerHTML=KL.map(k=>
      '<span class="lj"><i style="background:'+CFG.renk[k]+'"></i>'+k+' <b>'+fmt(S.sayilar[k][0])+'</b></span>').join("");
    const eks=kutu&&kutu.querySelector(".eksen");
    if(eks) eks.innerHTML='<span>0</span><span>'+fmt(enB)+' bulgu</span>';
    const ipu=k=>((CFG.aciklama||{})[k]||"").replace(/"/g,"&quot;");
    ck.innerHTML=KL.map(k=>
      '<div class="minisut ipuc" data-kural="'+k+'" data-ipucu="'+ipu(k)+'">'
      +'<span class="msdeger">'+S.sayilar[k][0]+'</span>'
      +'<div class="msbar" style="height:'+Math.max(6,100*S.sayilar[k][0]/enB)+'px;background:'+CFG.renk[k]+'"></div>'
      +'<span class="msad">'+k+'</span></div>').join("");
  }
  // donut
  const cd=document.getElementById("c_donut");
  if(cd){
    const cevre=2*Math.PI*44, pay=cevre*(S.toplamM?S.dogru/S.toplamM:0);
    cd.innerHTML='<div class="tbaslik">Veri Durumu<small>Doğru / hatalı dağılımı</small></div>'
      +'<div class="halka"><svg width="118" height="118" viewBox="0 0 112 112">'
      +'<circle cx="56" cy="56" r="44" fill="none" stroke="currentColor" stroke-width="16" opacity=".13"/>'
      +'<g transform="rotate(-90 56 56)">'
      +'<circle cx="56" cy="56" r="44" fill="none" stroke="#01B8AA" stroke-width="16" stroke-dasharray="'+pay.toFixed(2)+' '+(cevre-pay).toFixed(2)+'"/>'
      +'<circle cx="56" cy="56" r="44" fill="none" stroke="#FD625E" stroke-width="16" stroke-dasharray="'+(cevre-pay).toFixed(2)+' '+pay.toFixed(2)+'" stroke-dashoffset="'+(-pay).toFixed(2)+'"/></g>'
      +'<text x="56" y="54" text-anchor="middle" fill="currentColor" font-size="19" font-weight="600">%'+yuzde.toFixed(1)+'</text>'
      +'<text x="56" y="70" text-anchor="middle" fill="currentColor" opacity=".62" font-size="10">SAĞLIKLI</text></svg>'
      +'<div class="lejlist"><span class="lj"><i style="background:#01B8AA"></i>Doğru <b>'+fmt(S.dogru)+'</b></span>'
      +'<span class="lj"><i style="background:#FD625E"></i>Hatalı <b>'+fmt(S.hatali)+'</b></span>'
      +'<span class="lj">Toplam <b>'+fmt(S.toplamM)+'</b> malzeme</span></div></div>';
  }
  // ---- ANA EKRAN (giriş) ----
  const gm=document.getElementById("g_metrik");
  if(gm) gm.innerHTML=
    '<div class="g-mkart"><span>Toplam bulgu</span><b>'+fmt(toplamBulgu)+'</b></div>'
    +'<div class="g-mkart"><span>Hazır düzeltme</span><b>'+fmt(S.hazir)+'</b></div>'
    +'<div class="g-mkart"><span>Taranan</span><b>'+fmt(S.toplamM)+'</b></div>';
  const gs=document.getElementById("g_skor");
  if(gs){
    const cev=2*Math.PI*62, yay=cev*(yuzde/100);
    const durumMetni = yuzde>=95 ? "Veri sağlığı hedefin üzerinde"
                     : (yuzde>=85 ? "Veri sağlığı kabul aralığında"
                                  : "Veri sağlığı müdahale gerektiriyor");
    gs.innerHTML='<div class="g-halka">'
      +'<svg viewBox="0 0 152 152" width="152" height="152">'
      +'<circle cx="76" cy="76" r="62" fill="none" stroke="rgba(255,255,255,.06)" stroke-width="18"/>'
      +'<g transform="rotate(-90 76 76)">'
      +'<circle cx="76" cy="76" r="62" fill="none" stroke="rgba(255,255,255,.14)" stroke-width="11"/>'
      +'<circle cx="76" cy="76" r="62" fill="none" stroke="#01B8AA" stroke-width="11" '
      +'stroke-linecap="round" stroke-dasharray="'+yay.toFixed(1)+' '+cev.toFixed(1)+'"/></g>'
      +'<text x="76" y="72" text-anchor="middle" fill="#FFFFFF" font-size="34" font-weight="600" '
      +'letter-spacing="-1" font-family="Segoe UI Variable Display, Segoe UI, sans-serif">%'+yuzde.toFixed(1)+'</text>'
      +'<text x="76" y="92" text-anchor="middle" fill="#8FD4CC" font-size="10" letter-spacing="1.6" '
      +'font-family="Segoe UI, sans-serif">VERİ SAĞLIĞI</text></svg></div>'
      +'<div class="g-skorbilgi">'
      +'<div class="g-skorsatir"><i class="ok"></i><b>'+fmt(S.dogru)+'</b> kurallardan geçti</div>'
      +'<div class="g-skorsatir"><i class="hata"></i><b>'+fmt(S.hatali)+'</b> bulgusu var</div></div>'
      +'<div class="g-durum">'+durumMetni+'</div>';
  }
  // rozetler (eski sürüm uyumluluğu)
  const rz=document.getElementById("rozetler");
  if(rz) rz.innerHTML='<div class="rzt"><b>%'+yuzde.toFixed(1)+'</b><span>VERİ SAĞLIĞI</span></div>'
    +'<div class="rzt"><b>'+fmt(toplamBulgu)+'</b><span>BULGU</span></div>'
    +'<div class="rzt"><b>'+fmt(S.toplamM)+'</b><span>MALZEME TARANDI</span></div>'
    +'<div class="rzt"><b>'+S.hazir+'</b><span>HAZIR DÜZELTME</span></div>';
  // marka
  const cm=document.getElementById("c_marka");
  if(cm){
    const sirali=Object.entries(S.marka).map(
        ([mk,v])=>[mk,v,KL.reduce((a,k)=>a+(v[k]||0),0)])
      .sort((a,b)=>b[2]-a[2]).slice(0,6);
    const enT=Math.max(...sirali.map(x=>x[2]),1);
    const ort=sirali.reduce((a,x)=>a+x[2],0)/Math.max(sirali.length,1);
    const ortAlt=18+160*ort/enT;
    cm.innerHTML='<div class="tbaslik">Bulgu Sayısı<small>Markaya ve kurala göre · kesikli çizgi = ortalama</small></div>'
      +'<div class="yiginlar ort-sarici"><div class="ort-cizgi" style="bottom:'+ortAlt.toFixed(0)+'px"></div>'
      +'<div class="ort-etiket" style="bottom:'+(ortAlt+2).toFixed(0)+'px">Ort: '+ort.toFixed(0)+'</div>'
      +sirali.map(([mk,v,t])=>
        '<div class="yigin"><span class="ydeger">'+t+'</span><div class="ycubuk">'
        +KL.filter(k=>v[k]>0).map(k=>'<div style="height:'+(160*v[k]/enT).toFixed(1)+'px;background:'+CFG.renk[k]+'"></div>').join("")
        +'</div><span class="yad">'+kacis(mk)+'</span></div>').join("")
      +'</div><div class="lejantlar">'+KL.map(k=>'<span class="lj"><i style="background:'+CFG.renk[k]+'"></i>'+k+'</span>').join("")+'</div>';
  }
  // tür
  const ctur=document.getElementById("c_tur");
  if(ctur){
    const tumTur={}; for(const k in S.tur) for(const t in S.tur[k]) tumTur[t]=(tumTur[t]||0)+S.tur[k][t];
    const ilk=Object.entries(tumTur).sort((a,b)=>b[1]-a[1]).slice(0,2).map(x=>x[0]);
    ctur.innerHTML='<div class="tbaslik">Bulgu Dağılımı<small>Kural × malzeme türü</small></div>'
      +KL.filter(k=>S.sayilar[k][0]>0).map(k=>{
        const n=S.sayilar[k][0], t1=(S.tur[k]&&S.tur[k][ilk[0]])||0, p1=100*t1/n;
        return '<div class="turb"><div class="turust"><span>'+k+'</span><span class="mono">'+n+'</span></div>'
          +'<div class="turray"><div style="width:'+p1.toFixed(0)+'%;background:var(--koyu)"></div><div style="width:'+(100-p1).toFixed(0)+'%;background:var(--koyu2)"></div></div></div>';
      }).join("")
      +(ilk.length>=2?'<div class="lejantlar"><span class="lj"><i style="background:var(--koyu)"></i>Tür '+kacis(ilk[0])+'</span><span class="lj"><i style="background:var(--koyu2)"></i>Tür '+kacis(ilk[1])+' / diğer</span></div>':"");
  }
  // etki
  const ce=document.getElementById("c_etki");
  if(ce){
    ce.innerHTML='<div class="tbaslik">Düzeltme Etkisi<small>İlgili kural çözülürse ulaşılacak veri sağlığı · şu an %'+yuzde.toFixed(1)+'</small></div>'
      +S.etki.map(([k,y])=>'<div class="etkib"><span class="ead">'+k+'</span><div class="eray"><div style="width:'+y.toFixed(0)+'%;background:#01B8AA"></div></div><span class="mono">%'+y.toFixed(1)+'</span></div>').join("")
      +'<div class="etkib"><span class="ead"><b>Tümü</b></span><div class="eray"><div style="width:100%;background:var(--koyu)"></div></div><span class="mono"><b>%100</b></span></div>';
  }
}

/* ---- Yükleme akışı ---- */
const VARSAYILAN_ETIKET = {
  Urun:"XLSX veya CSV seçin",
  Ust:"XLSX veya CSV seçin",
  Marm:"Kural 4 (çokluk kontrolü) için — isteğe bağlı",
  Ayar:"Yeni kod/kelime eklemek için — mevcut kurallar korunur",
  Birlesik:"Ürün + üst kod aynı dosyada — RC ile başlayanlar üst kod, rakamla başlayanlar ürün"
};
function dosyaVar(ad){
  const i=document.getElementById("f"+ad);
  return !!(i && i.files && i.files[0]);
}
function kutuKilitle(){
  const tekliSecili = dosyaVar("Urun") || dosyaVar("Ust");
  const birlesikSecili = dosyaVar("Birlesik");
  // Tekli dosya seçiliyse birleşik liste devre dışı; birleşik seçiliyse tekliler
  [["Urun", birlesikSecili], ["Ust", birlesikSecili],
   ["Birlesik", tekliSecili]].forEach(([ad, pasif])=>{
    const kutu=document.getElementById("kutu"+ad);
    const inp=document.getElementById("f"+ad);
    if(!kutu||!inp) return;
    kutu.classList.toggle("pasif", pasif);
    inp.disabled = pasif;
    const not=document.getElementById("not"+ad);
    if(not) not.textContent = pasif
      ? (ad==="Birlesik" ? "Ayrı dosya seçildiği için devre dışı"
                         : "Birleşik liste seçildiği için devre dışı")
      : "";
  });
}
function dosyaTemizle(ad){
  const inp=document.getElementById("f"+ad);
  if(!inp) return;
  inp.value="";
  const et=document.getElementById("ad"+ad);
  if(et) et.textContent=VARSAYILAN_ETIKET[ad];
  const kutu=document.getElementById("kutu"+ad);
  if(kutu) kutu.classList.remove("dolu");
  kutuKilitle();
}
document.addEventListener("click", e=>{
  const d=e.target.closest("[data-temizle]");
  if(d){ e.preventDefault(); e.stopPropagation(); dosyaTemizle(d.dataset.temizle); }
});
["Urun","Ust","Marm","Ayar","Birlesik"].forEach(ad=>{
  const inp=document.getElementById("f"+ad);
  if(inp) inp.addEventListener("change",()=>{
    const f=inp.files[0];
    document.getElementById("ad"+ad).textContent=f?f.name:VARSAYILAN_ETIKET[ad];
    document.getElementById("kutu"+ad).classList.toggle("dolu",!!f);
    kutuKilitle();
  });
  // Pasif kutuya tıklanınca dosya seçici açılmasın
  const kutu=document.getElementById("kutu"+ad);
  if(kutu) kutu.addEventListener("click", e=>{
    if(kutu.classList.contains("pasif")) e.preventDefault();
  });
});
kutuKilitle();

function calisma(ad,veri,sayfa){
  const ws=XLSX.utils.json_to_sheet(veri);
  const wb=XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb,ws,sayfa);
  return wb;
}
const KRENK={K1:"3599B8",K2:"F2C80F",K3:"FD625E",K4:"A66999",
             K5:"118DFF",K6:"0E8A7E"};
function stilUygula(ws,veri,ozellik){
  // xlsx-js-style yüklüyse modern biçim; düz SheetJS'te sessizce yok sayılır
  if(!veri.length) return;
  const basliklar=Object.keys(veri[0]);
  const ince={style:"thin",color:{rgb:"DFE6EA"}};
  const kenar={bottom:ince};
  basliklar.forEach((b,c)=>{
    const h=ws[XLSX.utils.encode_cell({r:0,c})];
    if(h) h.s={fill:{fgColor:{rgb:"1B2A3A"}},font:{color:{rgb:"FFFFFF"},bold:true,sz:10},
      alignment:{vertical:"center"},border:{bottom:{style:"medium",color:{rgb:"E3A63B"}}}};
  });
  const kuralC=basliklar.findIndex(b=>b==="Kural"||b==="Rule");
  for(let r0=0;r0<veri.length;r0++){
    for(let c=0;c<basliklar.length;c++){
      const h=ws[XLSX.utils.encode_cell({r:r0+1,c})];
      if(!h) continue;
      h.s={border:kenar,alignment:{vertical:"top",wrapText:(ozellik&&ozellik.sarma||[]).includes(basliklar[c])},
        font:{sz:10}};
      if(r0%2===1) h.s.fill={fgColor:{rgb:"F5F8FA"}};
      if(c===kuralC&&KRENK[String(h.v)])
        h.s={...h.s,font:{color:{rgb:"FFFFFF"},bold:true,sz:10},
          fill:{fgColor:{rgb:KRENK[String(h.v)]}},alignment:{horizontal:"center"}};
    }
  }
  ws["!rows"]=[{hpt:22}];
  ws["!cols"]=basliklar.map(b=>({wch:(ozellik&&ozellik.gen&&ozellik.gen[b])||Math.min(Math.max(b.length+2,10),16)}));
  ws["!autofilter"]={ref:XLSX.utils.encode_range({s:{r:0,c:0},e:{r:veri.length,c:basliklar.length-1}})};
}
function tumunuIndir(S){
  const EN=S.dil==="EN";
  // 1) Düzeltme listesi
  bulgularIndir(S);
  // 2) Düzeltilmiş veri seti
  duzeltilmisIndir(S);
  // 3) Sonuç: Özet + Genel Durum
  const KADI=EN?{K1:"Parent code match",K2:"Description / field",
                 K3:"Forbidden words",K4:"Pack size / MARM",
                 K5:"Parent code uniqueness",K6:"Code ending / Origin"}
               :{K1:"Akıllı kod eşleşmesi",K2:"Tanım / ek veri alanı",
                 K3:"Yasaklı kelimeler",K4:"Çokluk / MARM sayacı",
                 K5:"Üst kod benzersizliği",K6:"Kod sonu / Menşei"};
  const ozet=["K1","K2","K3","K4","K5","K6"].map(k=>({
    [EN?"Rule":"Kural"]:k+" - "+KADI[k],
    [EN?"Findings":"Bulgu"]:S.sayilar[k][0],
    [EN?"Materials":"Malzeme Sayısı"]:S.sayilar[k][1]}));
  // Alt tarifler bir kural değil, bilgi kaydıdır: özetin sonuna ayrı eklenir
  ozet.push({[EN?"Rule":"Kural"]:EN?"Info - Sub-descriptions":"Bilgi - Alt Tarifler",
    [EN?"Findings":"Bulgu"]:S.rows.filter(r=>r.k==="ALT").length,
    [EN?"Materials":"Malzeme Sayısı"]:S.k4set.size});
  const wb2=calisma("s",ozet,EN?"Summary":"Özet");
  stilUygula(wb2.Sheets[EN?"Summary":"Özet"],ozet,{});
  const H=EN?["Source","Material","Description","K1","K2","K3","K4","K5","K6",
              "Sub-desc. (info)","Status"]
            :["Kaynak","Malzeme","Tanım","K1","K2","K3","K4","K5","K6",
              "Alt Tarif (bilgi)","Genel Durum"];
  const D=EN?["OK","ERROR"]:["DOĞRU","HATALI"];
  const genel=S.matlar.map(x=>{
    const h=KL.filter(k=>S.setK[k].has(x.m));
    const satir={[H[0]]:EN?(x.kay==="Ürün Kodu"?"Product":"Parent"):x.kay,
                 [H[1]]:x.m,[H[2]]:x.t};
    KL.forEach((k,i)=>{ satir[H[3+i]]=S.setK[k].has(x.m)?D[1]:D[0]; });
    satir[H[9]]=S.k4set.has(x.m)?(EN?"SUB":"ALT TARİF"):"-";
    satir[H[10]]=h.length?D[1]:D[0];
    return satir;});
  const wsGenel=XLSX.utils.json_to_sheet(genel);
  stilUygula(wsGenel,genel,{gen:{[H[2]]:34}});
  XLSX.utils.book_append_sheet(wb2,wsGenel,EN?"Overall":"Genel_Durum");
  // 3) Kategori analizi (uzun format)
  const kats=["Malzeme türü","Mal grubu","Raporlama Markası","Raporlama Alt Markası",
    "Menşei","Pazar","Ek Alan","Ambalaj Tipi","S&OP Kategorisi","Ürün Boyutu"];
  const hset=new Set(KL.flatMap(k=>[...S.setK[k]]));
  const tum=[...S._urun,...S._ust];
  const kayit=[];
  for(const kat of kats){
    if(!tum.some(s=>kat in s)) continue;
    const g={};
    for(const s of tum){
      const d=String(s[kat]??"").trim()||"(boş)"; const m=String(s["Malzeme"]);
      (g[d]=g[d]||{t:new Set(),h:0}); g[d].t.add(m);
    }
    for(const s of tum){
      const d=String(s[kat]??"").trim()||"(boş)";
      if(hset.has(String(s["Malzeme"]))) g[d].h++;
    }
    const sirali=Object.entries(g).sort((a,b)=>b[1].t.size-a[1].t.size).slice(0,25);
    for(const [d,v] of sirali)
      kayit.push({[EN?"Category":"Kategori"]:kat,[EN?"Value":"Değer"]:d,
        [EN?"Total":"Toplam"]:v.t.size,[EN?"Faulty":"Hatalı"]:v.h,
        [EN?"Error %":"Hata %"]:v.t.size?+((100*v.h/v.t.size).toFixed(1)):0});
  }
  const wsKat=XLSX.utils.json_to_sheet(kayit);
  stilUygula(wsKat,kayit,{});
  XLSX.utils.book_append_sheet(wb2,wsKat,EN?"Category_Analysis":"Kategori_Analizi");
  XLSX.writeFile(wb2,"kalite_kontrol_sonuc.xlsx");
}
function duzeltilmisIndir(S){
  const EN=S.dil==="EN";
  const G=S.duz||[];
  // 1) Düzeltilmiş malzeme listeleri
  const uygula=(satirlar,kaynak)=>{
    const kopya=satirlar.map(s=>Object.assign({},s));
    const idx={}; kopya.forEach((s,i)=>idx[String(s["Malzeme"]).trim()]=i);
    for(const d of G){
      if(d.durum!=="DÜZELTİLDİ"||d.kaynak!==kaynak) continue;
      const i=idx[String(d.m).trim()];
      if(i===undefined) continue;
      if(d.alan && (d.alan in kopya[i])) kopya[i][d.alan]=d.yeni;
    }
    return kopya;
  };
  const uD=uygula(S._urun||[],"Ürün Kodu");
  const pD=uygula(S._ust||[],"Üst Kod");

  const H=EN?["Source","Material","Rule","Field","Old Value","New Value","Status","Note"]
            :["Kaynak","Malzeme","Kural","Alan","Eski Değer","Yeni Değer","Durum","Açıklama"];
  const gunluk=G.map(d=>({[H[0]]:d.kaynak,[H[1]]:d.m,[H[2]]:d.kural,[H[3]]:d.alan,
    [H[4]]:d.eski,[H[5]]:d.yeni,[H[6]]:d.durum,[H[7]]:d.aciklama}));
  const elle=gunluk.filter(x=>x[H[6]]==="ELLE");
  const duzeltildi=G.filter(d=>d.durum==="DÜZELTİLDİ").length;

  const ozet=[
    {[EN?"Metric":"Ölçüt"]:EN?"Total processed findings":"Toplam işlenen bulgu",
     [EN?"Value":"Değer"]:G.length},
    {[EN?"Metric":"Ölçüt"]:EN?"Auto-corrected":"Otomatik düzeltilen",
     [EN?"Value":"Değer"]:duzeltildi},
    {[EN?"Metric":"Ölçüt"]:EN?"Manual decision needed":"Elle karar gereken",
     [EN?"Value":"Değer"]:G.length-duzeltildi},
    {[EN?"Metric":"Ölçüt"]:EN?"Corrected products":"Düzeltilen ürün kodu",
     [EN?"Value":"Değer"]:new Set(G.filter(d=>d.durum==="DÜZELTİLDİ"&&d.kaynak==="Ürün Kodu").map(d=>d.m)).size},
    {[EN?"Metric":"Ölçüt"]:EN?"Corrected parents":"Düzeltilen üst kod",
     [EN?"Value":"Değer"]:new Set(G.filter(d=>d.durum==="DÜZELTİLDİ"&&d.kaynak==="Üst Kod").map(d=>d.m)).size},
  ];

  const wb=XLSX.utils.book_new();
  const say=(ad,veri,ozellik)=>{
    const ws=XLSX.utils.json_to_sheet(veri);
    stilUygula(ws,veri,ozellik||{});
    XLSX.utils.book_append_sheet(wb,ws,ad);
  };
  say(EN?"Summary":"Özet",ozet,{});
  say(EN?"Change_Log":"Değişiklik_Günlüğü",gunluk,
      {sarma:[H[4],H[5],H[7]],gen:{[H[4]]:34,[H[5]]:34,[H[7]]:36,[H[3]]:20}});
  say(EN?"Manual_Decisions":"Elle_Karar_Gerekenler",
      elle.length?elle:[{[H[0]]:"",[H[1]]:EN?"None":"Kayıt yok"}],
      {sarma:[H[4],H[5],H[7]],gen:{[H[4]]:34,[H[5]]:34,[H[7]]:36}});
  if(uD.length) say(EN?"Corrected_Products":"Düzeltilmiş_Ürün_Kodları",uD,{});
  if(pD.length) say(EN?"Corrected_Parents":"Düzeltilmiş_Üst_Kodlar",pD,{});
  XLSX.writeFile(wb,"powerbi_bulgular_duzeltilmis.xlsx");
}

function bulgularIndir(S){
  const EN=S.dil==="EN";
  const H=EN?["Material Description","Material","Parent Code","Rule","Error Reason","Related Field","Current Value","Expected Value","Sub-description Count","Correction To Apply"]
            :["Malzeme Kısa Metni","Malzeme","Üst Kod","Kural","Hata Nedeni","İlgili Alan","Mevcut Değer","Olması Gereken","Alt Tarif Sayısı","Yapılacak Düzeltme"];
  const veri=S.rows.filter(r=>r.k!=="ALT").map(r=>({[H[0]]:r.t,[H[1]]:r.m,[H[2]]:r.u,[H[3]]:r.k,[H[4]]:r.n,[H[5]]:r.a,[H[6]]:r.mv,[H[7]]:r.og,[H[8]]:r.alt||1,[H[9]]:r.d}));
  const wb=calisma("b",veri,EN?"Findings":"Bulgular");
  stilUygula(wb.Sheets[EN?"Findings":"Bulgular"],veri,
    {sarma:[H[0],H[4],H[5],H[7],H[9]],
     gen:{[H[0]]:30,[H[4]]:40,[H[5]]:14,[H[7]]:16,[H[9]]:42,[H[1]]:11,[H[2]]:11,[H[8]]:9}});
  XLSX.writeFile(wb,"powerbi_bulgular.xlsx");
}

function kuralSetiBirlestir(ayar){
  // Gömülü kurallar korunur; Excel'deki fazladan satırlar YENİ KURAL olur.
  if(!ayar || !ayar.satirlar || !ayar.satirlar.length) return 0;
  const kolonBul=(anahtarlar)=>{
    for(const c of ayar.kolonlar){
      const u=trU(String(c));
      if(anahtarlar.some(a=>u.indexOf(a)>=0)) return c;
    }
    return null;
  };
  const kG=kolonBul(["GRUP","GROUP","KATEGOR"]);
  const kD=kolonBul(["DEĞER","DEGER","VALUE","KOD","CODE"]);
  const kM=kolonBul(["METİN","METIN","TEXT","AÇIKLAMA","ACIKLAMA"]);
  if(!kG||!kD){ log("  Kural seti atlandı: 'Grup' ve 'Değer' sütunları bulunamadı."); return 0; }

  const alanAdi={"MENŞEİ":"Menşei","MENSEI":"Menşei","PAZAR":"Pazar",
    "EK ALAN":"Ek Alan","AMBALAJ TİPİ":"Ambalaj Tipi","AMBALAJ TIPI":"Ambalaj Tipi"};
  let yeni=0; const eklenen=[];
  for(const s of ayar.satirlar){
    const grup=trU(String(s[kG]||"")).trim();
    const deger=String(s[kD]==null?"":s[kD]).trim();
    const metin=kM?String(s[kM]==null?"":s[kM]).trim():"";
    if(!grup||!deger) continue;
    const D=trU(deger);
    if(grup==="YASAKLI"){
      if(!CFG.yasakli.map(trU).includes(D)){ CFG.yasakli.push(deger); yeni++; eklenen.push(grup+": "+deger); }
    } else if(alanAdi[grup]){
      const alan=alanAdi[grup];
      CFG.k2[alan]=CFG.k2[alan]||{};
      if(!Object.keys(CFG.k2[alan]).map(trU).includes(D)){
        CFG.k2[alan][D]=trU(metin); yeni++; eklenen.push(alan+": "+D);
      }
    } else if(grup==="ÜRÜN TÜRÜ"||grup==="URUN TÜRÜ"||grup==="URUN TURU"){
      const t=deger.replace(/^0+/,"")||"0";
      if(CFG.urunTur.indexOf(t)<0){ CFG.urunTur.push(t); yeni++; eklenen.push("Ürün türü: "+t); }
    } else if(grup==="ÜST TÜRÜ"||grup==="UST TURU"){
      const t=deger.replace(/^0+/,"")||"0";
      if(CFG.ustTur.indexOf(t)<0){ CFG.ustTur.push(t); yeni++; eklenen.push("Üst türü: "+t); }
    } else if(grup.indexOf("TANIM SONU")>=0){
      CFG.tanimSonuYasak=CFG.tanimSonuYasak||[];
      if(CFG.tanimSonuYasak.indexOf(D)<0){ CFG.tanimSonuYasak.push(D); yeni++; eklenen.push("Tanım sonu yasak: "+D); }
    }
  }
  if(yeni){
    log("  "+yeni+" YENİ KURAL eklendi: "+eklenen.slice(0,8).join(" · ")
        +(eklenen.length>8?" …":""));
  } else {
    log("  Kural seti okundu — yeni kural yok, tüm satırlar zaten tanımlı.");
  }
  return yeni;
}

let SON_SONUC=null;
const btn=document.getElementById("btnAnaliz");
if(btn) btn.addEventListener("click", async ()=>{
  const fu=document.getElementById("fUrun").files[0];
  const fp=document.getElementById("fUst").files[0];
  const fb=document.getElementById("fBirlesik").files[0];
  const l=document.getElementById("log");
  if(!fb&&(!fu||!fp)){
    l.textContent="Lütfen ya iki ayrı dosyayı (ürün + üst kod) ya da birleşik listeyi seçin.";
    return;
  }
  btn.disabled=true; l.textContent="Analiz başladı…";
  try{
    let urun, ust;
    if(fb){
      log("Birleşik liste okunuyor: "+fb.name);
      const hepsi=await dosyaOku(fb);
      log("  "+hepsi.satirlar.length+" satır, "+hepsi.kolonlar.length+" sütun");
      const uS=[], pS=[]; let disi=0;
      for(const s of hepsi.satirlar){
        const kod=String(s["Malzeme"]==null?"":s["Malzeme"]).trim().toUpperCase();
        if(kod.indexOf("RC")===0) pS.push(s);
        else if(/^[0-9]/.test(kod)) uS.push(s);
        else disi++;
      }
      urun={kolonlar:hepsi.kolonlar,satirlar:uS,dil:hepsi.dil};
      ust={kolonlar:hepsi.kolonlar,satirlar:pS,dil:hepsi.dil};
      log("  Ayrıştırıldı: "+uS.length+" ürün kodu (rakamla başlayan) + "
          +pS.length+" üst kod (RC ile başlayan)"
          +(disi?" · "+disi+" kayıt tanınmadı":""));
    } else {
      log("Ürün dosyası okunuyor: "+fu.name);
      urun=await dosyaOku(fu);
      log("  "+urun.satirlar.length+" satır, "+urun.kolonlar.length+" sütun");
      log("Üst kod dosyası okunuyor: "+fp.name);
      ust=await dosyaOku(fp);
      log("  "+ust.satirlar.length+" satır, "+ust.kolonlar.length+" sütun");
    }
    const fa=document.getElementById("fAyar").files[0];
    if(fa){
      log("Kural seti okunuyor: "+fa.name);
      const ayar=await dosyaOku(fa);
      log("  "+ayar.satirlar.length+" satır");
      kuralSetiBirlestir(ayar);
    }
    let marm=null;
    const fm=document.getElementById("fMarm").files[0];
    if(fm){
      log("MARM dosyası okunuyor: "+fm.name);
      marm=await dosyaOku(fm);
      log("  "+marm.satirlar.length+" satır, "+marm.kolonlar.length+" sütun");
    } else {
      log("MARM yüklenmedi — Kural 4 (çokluk kontrolü) atlanacak.");
    }
    const dil=(urun.dil==="EN"||ust.dil==="EN")?"EN":"TR";
    log("Dil algılandı: "+(dil==="EN"?"İngilizce":"Türkçe"));
    const S=motor(urun,ust,dil,marm);
    SON_SONUC=S;
    if(S.kapsamDisi) log("Kapsam dışı tür: "+S.kapsamDisi+" kayıt atlandı (hata değil).");
    log("K1: "+S.sayilar.K1[0]+" · K2: "+S.sayilar.K2[0]+" · K3: "+S.sayilar.K3[0]
        +" · K4: "+S.sayilar.K4[0]+" · K5: "+S.sayilar.K5[0]+" bulgu");
    log("Veri sağlığı: %"+(S.toplamM?(100*S.dogru/S.toplamM).toFixed(1):"0")+" ("+S.dogru+" doğru / "+S.hatali+" hatalı / "+S.toplamM+")");
    BULGULAR=S.rows; filtre="ALL"; limit=60; ciz(); panoGuncelle(S);
    gecmisEkle(S); gecmisCiz();
    log("Çalıştırma tarihçeye kaydedildi (4 · Tarihçe sekmesi).");
    log("Sonuç dosyaları indiriliyor (3 Excel: sonuç + bulgular + düzeltilmiş veri)…");
    try{ tumunuIndir(S); }catch(e){ log("İndirme uyarısı: "+e.message+" — aşağıdaki düğmeleri kullanın."); }
    document.getElementById("indirGrup").style.display="flex";
    log("Panel güncellendi — ana ekrana yönlendiriliyorsunuz…");
    setTimeout(()=>git("giris"), 1200);
  }catch(err){ log("HATA: "+err.message); }
  btn.disabled=false;
});
const b1=document.getElementById("ind1"), b2=document.getElementById("ind2");
if(b1) b1.addEventListener("click",()=>SON_SONUC&&tumunuIndir(SON_SONUC));
if(b2) b2.addEventListener("click",()=>SON_SONUC&&bulgularIndir(SON_SONUC));
const b3=document.getElementById("ind3");
if(b3) b3.addEventListener("click",()=>SON_SONUC&&duzeltilmisIndir(SON_SONUC));
const gAc=document.getElementById("grafikAc");
if(gAc) gAc.addEventListener("click",()=>{
  const yer=document.getElementById("gecmisGrafik");
  const gizli=yer.dataset.gizli==="1";
  yer.dataset.gizli = gizli ? "0" : "1";
  gAc.textContent = gizli ? "Grafiği gizle" : "Grafiği göster";
  gAc.classList.toggle("aktif", gizli);
  grafikCiz(gecmisOku());
});
const gListe=document.getElementById("gecmisListe");
if(gListe) gListe.addEventListener("click", e => {
  const d = e.target.closest("[data-sil]");
  if(d) gecmisSil(parseInt(d.dataset.sil, 10));
});
gecmisCiz();
ciz();
</script>"""

    # ---------------------------------------------------------------
    # FİLTRE RAYI (sticky sol panel)
    # ---------------------------------------------------------------
    _yan_bas = f"""<aside class="yan">
<div class="marka-blok">{ik("filtre", 15)}Filtreler</div>
<div class="marka-alt">SAP MM · MALZEME MASTER</div>
<div class="f-grup"><label>Çalıştırma tarihi</label>
<div class="tarihk">{ik("takvim", 14)}{tarih}</div></div>
<div class="f-grup"><label>Kural</label>
<select class="kuralSec" aria-label="Kural filtresi"><option value="ALL">Tümü</option>
<option value="K1">K1 · Akıllı kod</option><option value="K2">K2 · Tanım/Ek alan</option>
<option value="K3">K3 · Yasaklı kelime</option>
<option value="K4">K4 · Çokluk / MARM</option>
<option value="K5">K5 · Üst kod tekilliği</option></select></div>"""
    _alan_secenek = ""
    if not basit.empty:
        _alanlar = set()
        for _a in basit["İlgili Alan"].dropna().astype(str).unique():
            for _p in _re.split(r"\s*[;/]\s*", _a):   # "Menşei; Pazar" -> ayrı seçenekler
                if _p.strip():
                    _alanlar.add(_p.strip())
        for _a in sorted(_alanlar):
            _alan_secenek += f'<option value="{_a}">{_a}</option>'
    _yan_alan = ('<div class="f-grup"><label>Alan</label>'
                 '<select class="alanSec" aria-label="Alan filtresi">'
                 '<option value="ALL">Tümü</option>'
                 + _alan_secenek + '</select></div>')
    _yan_kaynak = """<div class="f-grup"><label>Kaynak</label>
<select class="kaynakSec" aria-label="Kaynak filtresi"><option value="ALL">Tümü</option>
<option value="Ürün Kodu">Ürün Kodu</option><option value="Üst Kod">Üst Kod</option></select></div>"""
    _yan_son = """<div class="f-grup"><label>Kapsam</label>
<div class="bilgi-panel"><b>Ürün:</b> tür 6/7<br><b>Üst kod:</b> tür 15<br>
Diğer türler analiz dışı</div></div>
</aside>"""
    yan = _yan_bas + _yan_son               # Kontrol Paneli ve Analiz
    # Tarihçe sayfasında kural/kaynak filtresi anlamsız; yalnızca kimlik ve
    # kapsam bilgisi gösterilir.
    yan4 = f"""<aside class="yan">
<div class="marka-blok">{ik("filtre", 15)}Filtreler</div>
<div class="marka-alt">SAP MM · MALZEME MASTER</div>
<div class="f-grup"><label>Çalıştırma tarihi</label>
<div class="tarihk">{ik("takvim", 14)}{tarih}</div></div>
<div class="f-grup"><label>Tarihçe hakkında</label>
<div class="bilgi-panel">Her analiz otomatik kaydedilir. Satırın üzerine gelin:
en çok hata alan alanlar ve önceki koşuma göre iyileşen/kötüleşen alanlar görünür.</div></div>
<div class="f-grup"><label>Kapsam</label>
<div class="bilgi-panel"><b>Ürün:</b> tür 6/7<br><b>Üst kod:</b> tür 15<br>
Silinmiş kayıtlar hariç</div></div>
</aside>"""
    yan3 = _yan_bas + _yan_kaynak + _yan_alan + _yan_son  # Düzeltme Listesi

    # ---------------------------------------------------------------
    # İKİ KATMANLI KURUMSAL BAŞLIK
    # ---------------------------------------------------------------
    _sekme_tanim = [("giris", "ana", "Ana Ekran"), ("p0", "yukle", "Veri Yükle"),
                    ("p1", "panel", "Kontrol Paneli"), ("p2", "analiz", "Analiz"),
                    ("p3", "liste", "Düzeltme Listesi"), ("p4", "tarihce", "Tarihçe")]
    _sekmeler = "".join(f'<button data-git="{_p}">{ik(_i, 15)}{_ad}</button>'
                        for _p, _i, _ad in _sekme_tanim)

    ustbar = f"""<header class="ustbar"><div class="ub-sar">
<div class="ub-ust">{marka_kilidi()}
<div class="ub-cip">
<span class="ub-durum"><span class="nokta"></span>SAP MM</span>
<span class="ub-durum">{ik("takvim", 14)}Çalıştırma <span class="mono">{tarih}</span></span>
<span class="ub-durum">{ik("kitap", 14)}{n_kural} kural</span>
</div>
<div class="ub-arac">
<button class="ub-dugme tema">Koyu tema</button>
<button class="ub-dugme yazdir">PDF kaydet</button>
</div></div>
<nav class="sekmeler" aria-label="Sayfa gezinmesi">{_sekmeler}</nav>
<div class="komut">
<input class="soru" type="text" aria-label="Panel içinde ara"
 placeholder="Bu paneldeki veri hakkında ara — malzeme kodu, alan adı veya sorun yazın…"></div>
</div></header>"""

    hazir_oneri = 0
    if not k2.empty and "Önerilen Tanım" in k2.columns:
        hazir_oneri = int((k2["Önerilen Tanım"].astype(str).str.strip() != "").sum())

    _durum_metni = ("Veri sağlığı hedefin üzerinde" if yuzde >= 95 else
                    ("Veri sağlığı kabul aralığında" if yuzde >= 85 else
                     "Veri sağlığı müdahale gerektiriyor"))

    html = f"""<!DOCTYPE html><html lang="tr"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Evyap · Malzeme Veri Kalitesi — Kontrol Paneli</title>
<style>{stil}{khap_stil}</style></head><body>

<div class="baski-ust">{marka_kilidi()}
<div class="baski-bilgi"><b>Malzeme Veri Kalitesi Raporu</b><br>
SAP MM · Malzeme Master · Bilgi Teknolojileri<br>Çalıştırma tarihi: {tarih}</div></div>

<div id="giris">
<div class="g-arka">
<span class="g-parlak"></span><span class="g-parlak b"></span>
<span class="g-izgara"></span>
<svg class="g-kayit" viewBox="0 0 1200 620" preserveAspectRatio="none" aria-hidden="true">
<g stroke="rgba(255,255,255,.07)" fill="none" stroke-width="1">
<rect x="742" y="86" width="392" height="34" rx="4"/>
<rect x="742" y="128" width="392" height="34" rx="4"/>
<rect x="742" y="170" width="392" height="34" rx="4"/>
<rect x="742" y="212" width="392" height="34" rx="4"/>
<path d="M862 86v160M962 86v160M1052 86v160"/>
</g>
<g fill="rgba(1,184,170,.16)">
<rect x="750" y="96" width="86" height="12" rx="2"/>
<rect x="750" y="138" width="70" height="12" rx="2"/>
<rect x="750" y="180" width="94" height="12" rx="2"/>
<rect x="750" y="222" width="62" height="12" rx="2"/>
</g>
</svg>
</div>

<div class="g-ic">
<header class="g-ust">
{marka_kilidi(koyu=True)}
<div class="g-ustsag">
<span class="g-rozet-mini">SAP MM · Malzeme Master</span>
<span class="g-rozet-mini">Çalıştırma <span class="mono">{tarih}</span></span>
<button class="g-ustdugme" data-git="p0">{ik("yukle", 15)}Yeni veri yükle</button>
</div>
</header>

<section class="g-kahraman">
<div class="g-sol">
<div class="g-etiket">MALZEME MASTER VERİ KALİTESİ</div>
<h1 class="g-baslik">Veriniz ne kadar<br><span class="g-vurgu">sağlıklı?</span></h1>
<p class="g-alt">Ürün kodu ↔ üst kod tutarlılığını {n_kural} kuralla denetleyen,
hataları açıklayan ve düzeltme önerisini kendisi üreten otomatik kontrol sistemi.</p>
<div class="g-metrik" id="g_metrik">
<div class="g-mkart"><span>Toplam bulgu</span><b>{toplam_bulgu:,}</b></div>
<div class="g-mkart"><span>Hazır düzeltme</span><b>{hazir_oneri:,}</b></div>
<div class="g-mkart"><span>Taranan</span><b>{toplam:,}</b></div>
</div>
</div>
<div class="g-sag" id="g_skor">
<div class="g-halka">
<svg viewBox="0 0 152 152" width="152" height="152" role="img"
 aria-label="Veri sağlığı yüzde {yuzde:.1f}">
<circle cx="76" cy="76" r="62" fill="none" stroke="rgba(255,255,255,.06)" stroke-width="18"/>
<g transform="rotate(-90 76 76)">
<circle cx="76" cy="76" r="62" fill="none" stroke="rgba(255,255,255,.14)" stroke-width="11"/>
<circle cx="76" cy="76" r="62" fill="none" stroke="{TEAL}" stroke-width="11"
 stroke-linecap="round" stroke-dasharray="{skor_yay:.1f} {skor_cevre:.1f}"/>
</g>
<text x="76" y="72" text-anchor="middle" fill="#FFFFFF" font-size="34"
 font-weight="600" letter-spacing="-1" font-family="Segoe UI Variable Display, Segoe UI, sans-serif">%{yuzde:.1f}</text>
<text x="76" y="92" text-anchor="middle" fill="#8FD4CC" font-size="10"
 letter-spacing="1.6" font-family="Segoe UI, sans-serif">VERİ SAĞLIĞI</text>
</svg>
</div>
<div class="g-skorbilgi">
<div class="g-skorsatir"><i class="ok"></i><b>{dogru:,}</b> kurallardan geçti</div>
<div class="g-skorsatir"><i class="hata"></i><b>{hatali:,}</b> bulgusu var</div>
</div>
<div class="g-durum">{_durum_metni}</div>
</div>
</section>

<section>
<div class="g-bolum-bas"><h2>Modüller</h2><span>Panelin dört ana bölümü</span></div>
<div class="g-kartlar">{giris_nav}</div>
</section>

<footer class="g-dip"><span>malzeme_kalite_kontrol.py</span><span>·</span>
<span>{tarih}</span><span>·</span>
<span>panel her çalıştırmada güncel veriden üretilir</span></footer>
</div>
</div>

<div id="p0" class="sayfa acik">
<div class="p0-hero">
<span class="g-izgara"></span>
<div class="p0-kart">
<div class="p0-ust">{marka_kilidi()}
<span class="etiket2">{ik("veritabani", 14)}SAP MM veri aktarımı</span></div>
<h1 class="p0-baslik">Veri Yükleme</h1>
<p class="p0-alt">Ürün kodları ve üst kodlar dosyalarını seçin — analiz tarayıcınızda çalışır,
sonuç dosyaları otomatik iner ve panel açılır.</p>

<div class="p0-adimlar">
<div class="adim-bas"><span class="rozet-adim">1</span>Zorunlu kaynak dosyalar</div>
<label class="dosya-kutu" id="kutuUrun"><span class="dk-no">1</span>
<span class="dk-metin"><b>Ürün kodları</b><small id="adUrun">XLSX veya CSV seçin</small>
<span class="kutu-not" id="notUrun"></span></span>
<input type="file" id="fUrun" accept=".xlsx,.xls,.csv"><button class="dk-sil" data-temizle="Urun" title="Ürün kodları seçimini kaldır" aria-label="Ürün kodları seçimini kaldır">{ik("kapat", 14)}</button></label>
<label class="dosya-kutu" id="kutuUst"><span class="dk-no">2</span>
<span class="dk-metin"><b>Üst kodlar</b><small id="adUst">XLSX veya CSV seçin</small>
<span class="kutu-not" id="notUst"></span></span>
<input type="file" id="fUst" accept=".xlsx,.xls,.csv"><button class="dk-sil" data-temizle="Ust" title="Üst kodlar seçimini kaldır" aria-label="Üst kodlar seçimini kaldır">{ik("kapat", 14)}</button></label>
<div class="ayirici"><span>ya da tek dosyada</span></div>
<label class="dosya-kutu birlesik" id="kutuBirlesik"><span class="dk-no">{ik("katman", 16)}</span>
<span class="dk-metin"><b>Birleşik liste</b>
<small id="adBirlesik">Ürün + üst kod aynı dosyada — RC ile başlayanlar üst kod, rakamla başlayanlar ürün</small>
<span class="kutu-not" id="notBirlesik"></span></span>
<input type="file" id="fBirlesik" accept=".xlsx,.xls,.csv"><button class="dk-sil" data-temizle="Birlesik" title="Birleşik liste seçimini kaldır" aria-label="Birleşik liste seçimini kaldır">{ik("kapat", 14)}</button></label>

<div class="adim-bas"><span class="rozet-adim">2</span>Ek dosyalar
<span class="istege">isteğe bağlı</span></div>
<label class="dosya-kutu" id="kutuMarm"><span class="dk-no">3</span>
<span class="dk-metin"><b>MARM · alternatif ölçü birimleri</b>
<small id="adMarm">Kural 4 (çokluk kontrolü) için — isteğe bağlı</small></span>
<input type="file" id="fMarm" accept=".xlsx,.xls,.csv"><button class="dk-sil" data-temizle="Marm" title="MARM seçimini kaldır" aria-label="MARM seçimini kaldır">{ik("kapat", 14)}</button></label>
<label class="dosya-kutu" id="kutuAyar"><span class="dk-no">4</span>
<span class="dk-metin"><b>Kural seti · ayarlar.xlsx</b>
<small id="adAyar">Yeni kod/kelime eklemek için — mevcut kurallar korunur</small>
<span class="kutu-not" id="notAyar"></span></span>
<input type="file" id="fAyar" accept=".xlsx,.xls,.csv"><button class="dk-sil" data-temizle="Ayar" title="Kural seti seçimini kaldır" aria-label="Kural seti seçimini kaldır">{ik("kapat", 14)}</button></label>
</div>

<div class="p0-butonlar">
<button class="analiz" id="btnAnaliz">{ik("saglik", 17)}Analiz Et ve Panele Geç</button>
<button class="ghost" data-git="giris">Yüklemeden devam{ik("ok", 14)}</button>
</div>
<div class="guven">{ik("kalkan", 17)}<span><b>Verileriniz bilgisayarınızdan çıkmaz.</b>
Dosyalar sunucuya gönderilmez; tüm analiz tarayıcınızın belleğinde çalışır.</span></div>

<div class="teknik-bas">{ik("bilgi", 14)}İşlem günlüğü</div>
<div id="log" class="logk">Başlıklar Türkçe veya İngilizce olabilir — otomatik algılanır.</div>
<div class="p0-butonlar" id="indirGrup" style="display:none">
<button class="ghost" id="ind1">{ik("indir", 14)}kalite_kontrol_sonuc.xlsx</button>
<button class="ghost" id="ind2">{ik("indir", 14)}powerbi_bulgular.xlsx</button>
<button class="ghost" id="ind3">{ik("indir", 14)}powerbi_bulgular_duzeltilmis.xlsx</button>
</div>
</div></div></div>

<div id="p1" class="sayfa">{ustbar}
<div class="kabuk">
<div class="sayfa-bas"><div><span class="goz">Modül 01 · SAP MM</span>
<h1>Kontrol Paneli</h1>
<p>Malzeme master verisinin genel sağlık özeti: kural bazlı bulgu dağılımı,
veri kalitesi boyutları ve aksiyon önceliği.</p></div>
<div class="sayfa-bas-yan">
<div class="ozet-kutu iyi"><span>Veri sağlığı</span><b>%{yuzde:.1f}</b></div>
<div class="ozet-kutu"><span>Bulgu</span><b>{toplam_bulgu:,}</b></div>
<div class="ozet-kutu"><span>Malzeme</span><b>{toplam:,}</b></div></div></div>
<div class="duzen">{yan}<div class="icerik">
<div class="kpi-serit" id="c_kpi">
<div class="kpi ana"><div class="kpi-ust"><span>Veri Sağlığı</span>{ik("saglik", 17, "ikn kpi-ikon")}</div>
<b>%{yuzde:.1f}</b><div class="kpi-ray"><div style="width:{yuzde:.1f}%"></div></div>
<em>kurallardan geçen malzeme oranı</em></div>
<div class="kpi uyari"><div class="kpi-ust"><span>Toplam Bulgu</span>{ik("bulgu", 16, "ikn kpi-ikon")}</div>
<b>{toplam_bulgu:,}</b><em>hata + bilgi kaydı</em></div>
<div class="kpi"><div class="kpi-ust"><span>Hatalı Malzeme</span>{ik("liste", 16, "ikn kpi-ikon")}</div>
<b>{hatali:,}</b><em>en az bir kurala takılan</em></div>
<div class="kpi"><div class="kpi-ust"><span>Taranan Malzeme</span>{ik("malzeme", 16, "ikn kpi-ikon")}</div>
<b>{toplam:,}</b><em>{urun_n:,} ürün + {ust_n:,} üst kod</em></div>
<div class="kpi"><div class="kpi-ust"><span>Hazır Düzeltme</span>{ik("onay", 16, "ikn kpi-ikon")}</div>
<b>{hazir_oneri:,}</b><em>program önerisi (K2)</em></div>
</div>
<div class="izg">
<div class="tile s5 ana-kart" id="c_boyut"><div class="tbaslik">Veri Kalitesi Boyutları
<small>Her boyut 100 üzerinden · yeşil ≥95 · sarı ≥85</small></div>
{boyut_html}</div>
<div class="tile s4 t-kural"><div class="tbaslik">Bulgu Sayısı<small>Kurallara göre · sütuna tıklayın, liste filtrelensin</small></div>
<div class="lejlist">{kural_lejant}</div>
<div class="minisutlar" id="c_kural">{kural_bar}</div>
<div class="eksen"><span>0</span><span>{en_b:,} bulgu</span></div></div>
<div class="tile s3" id="c_donut"><div class="tbaslik">Veri Durumu<small>Doğru / hatalı dağılımı</small></div>
<div class="halka"><svg width="118" height="118" viewBox="0 0 112 112">
<circle cx="56" cy="56" r="44" fill="none" stroke="currentColor" stroke-width="16" opacity=".13"/>
<g transform="rotate(-90 56 56)">
<circle cx="56" cy="56" r="44" fill="none" stroke="{TEAL}" stroke-width="16"
 stroke-dasharray="{pay:.2f} {cevre - pay:.2f}"/>
<circle cx="56" cy="56" r="44" fill="none" stroke="{MERCAN}" stroke-width="16"
 stroke-dasharray="{cevre - pay:.2f} {pay:.2f}" stroke-dashoffset="{-pay:.2f}"/>
</g>
<text x="56" y="54" text-anchor="middle" fill="currentColor" font-size="19" font-weight="600">%{yuzde:.1f}</text>
<text x="56" y="70" text-anchor="middle" fill="currentColor" opacity=".62" font-size="10">SAĞLIKLI</text></svg>
<div class="lejlist"><span class="lj"><i style="background:{TEAL}"></i>Doğru <b>{dogru:,}</b></span>
<span class="lj"><i style="background:{MERCAN}"></i>Hatalı <b>{hatali:,}</b></span>
<span class="lj">Toplam <b>{toplam:,}</b> malzeme</span></div></div></div>
<div class="tile s5" id="c_tablo"><div class="tbaslik">Kural Özeti<small>Bulgu · etkilenen malzeme · toplam içindeki pay</small></div>
{ozet_tablo}</div>
<div class="tile s4" id="c_alan"><div class="tbaslik">En Sorunlu Alanlar
<small>Bulgu sayısına göre ilk 6 · aksiyon önceliği</small></div>
{alan_html}</div>
<div class="tile s3 duz t-kapsam"><div class="tbaslik">Analiz Kapsamı<small>Neler değerlendirildi</small></div>
<div class="kapsamk"><b>{urun_n:,}</b> ürün kodu · tür 6/7</div>
<div class="kapsamk"><b>{ust_n:,}</b> üst kod · tür 15</div>
<div class="kapsamk"><b>{kapsam_disi:,}</b> kayıt kapsam dışı — analiz edilmedi</div>
<div class="kapsamk"><b>{istisna_n:,}</b> bulgu onaylı istisna olarak gizlendi</div>
</div>
{rehber}
</div></div></div></div></div>

<div id="p2" class="sayfa">{ustbar}
<div class="kabuk">
<div class="sayfa-bas"><div><span class="goz">Modül 02 · SAP MM</span>
<h1>Analiz</h1>
<p>Bulguların marka, malzeme türü ve kural kırılımı; hangi kuralın çözülmesinin
veri sağlığını ne kadar yükselteceği ve kural rehberi.</p></div>
<div class="sayfa-bas-yan">
<div class="ozet-kutu"><span>Bulgu</span><b>{toplam_bulgu:,}</b></div>
<div class="ozet-kutu"><span>Kural</span><b>{n_kural}</b></div></div></div>
<div class="duzen">{yan}<div class="icerik">
<div class="izg">{marka_html}{tur_html}{etki_html}{rehber}</div>
</div></div></div></div>

<div id="p3" class="sayfa">{ustbar}
<div class="kabuk">
<div class="sayfa-bas"><div><span class="goz">Modül 03 · SAP MM</span>
<h1>Düzeltme Listesi</h1>
<p>Neyi, neyle ve nasıl düzelteceğinizi satır satır gösteren çalışma listesi.
Mevcut değer, olması gereken değer ve önerilen düzeltme yan yana.</p></div>
<div class="sayfa-bas-yan">
<div class="ozet-kutu"><span>Satır</span><b>{len(kayitlar):,}</b></div>
<div class="ozet-kutu iyi"><span>Hazır öneri</span><b>{hazir_oneri:,}</b></div></div></div>
<div class="duzen">{yan3}<div class="icerik">
<div class="tile s12 t-liste"><div class="tbaslik">Düzeltme Listesi<small>Neyi, neyle, nasıl — satır satır ({len(kayitlar):,} bulgu)</small></div>
<div class="arac">
<span class="arac-sar"><input class="arama" id="ara" type="text" aria-label="Düzeltme listesinde ara"
 placeholder="Ara: malzeme, üst kod, sorun, alan, düzeltme…"></span>
<span class="cipler">{cipler}</span><span class="rozet-say" id="sonucBilgi"></span></div>
<div class="tabsar"><table class="ciftkod"><thead><tr id="basliklar"><th>Kısa Metin</th><th>Malzeme</th><th>Üst Kod</th><th>Kural</th><th>Hata Nedeni</th><th>İlgili Alan</th><th>Mevcut Değer</th><th>Olması Gereken</th><th>Alt Tarif</th><th>Yapılacak Düzeltme</th></tr></thead>
<tbody id="tb"></tbody></table></div>
<button class="daha" id="daha">{ik("ok", 14)}Daha fazla göster</button></div>
<div class="dipnot">Bu liste powerbi_bulgular.xlsx ile birebir aynıdır · filtreli hali PDF olarak kaydedilebilir</div>
</div></div></div></div>

<div id="p4" class="sayfa">{ustbar}
<div class="kabuk">
<div class="sayfa-bas"><div><span class="goz">Modül 04 · SAP MM</span>
<h1>Veri Sağlığı Tarihçesi</h1>
<p>Çalıştırmalar arasındaki değişim: hangi alanlar iyileşti, hangileri kötüleşti.
Her analiz otomatik kaydedilir.</p></div>
<div class="sayfa-bas-yan">
<div class="ozet-kutu iyi"><span>Son sağlık</span><b>%{yuzde:.1f}</b></div></div></div>
<div class="duzen">{yan4}<div class="icerik">
<div class="izg">
<div class="tile s12 ana-kart t-tarihce"><div class="tbaslik">Veri Sağlığı Tarihçesi
<small>Her çalıştırma kaydedilir · satırın üzerine gelin, koşum özetini görün</small></div>
<div class="arac">
<button class="arac-dugme aktif" id="grafikAc">Grafiği göster</button>
<span class="rozet-say" id="gecmisBilgi"></span>
<span class="altnot" style="margin:0">Bir kaydı kaldırmak için satırın sonundaki × düğmesini kullanın.</span>
</div>
<div id="gecmisGrafik" class="gecmis-grafik"></div>
<div id="gecmisListe"></div>
</div></div>
</div></div></div></div>

{js.replace("__CFG__", cfg_json).replace("__GECMIS__", gecmis_json).replace("__VERI__", veri_json)}
</body></html>"""
    with open(dosya, "w", encoding="utf-8") as f:
        f.write(html)


def grafikler_sayfasi(wb, ozet, durum, kategoriler):
    """Sonuç dosyasına yerleşik Excel grafikleri ekler: kural bulguları,
    doğru/hatalı dağılımı ve öne çıkan kategori analizleri."""
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.styles import Font

    ws = wb.create_sheet(L("Grafikler") if DIL != "EN" else "Charts", 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = ("Malzeme Veri Kalitesi — Grafik Özeti" if DIL != "EN"
                else "Material Data Quality — Chart Summary")
    ws["A1"].font = Font(bold=True, size=14, color="1B2A3A")

    # ---- Yardımcı veri blokları (grafiklerin kaynağı, sağda AA sütununda) ----
    K = 27  # AA
    def blok(bas_satir, baslik, satirlar):
        ws.cell(row=bas_satir, column=K, value=baslik).font = Font(bold=True, size=9)
        for i, (a, b) in enumerate(satirlar, start=1):
            ws.cell(row=bas_satir + i, column=K, value=a)
            ws.cell(row=bas_satir + i, column=K + 1, value=b)
        return bas_satir + 1, bas_satir + len(satirlar)

    # 1) Kural bulguları
    b1a, b1b = blok(1, "Kural bulguları",
                    [(str(s["Kural"]), int(s["Bulgu Sayısı"]))
                     for _, s in ozet.iterrows()])
    ch1 = BarChart(); ch1.type = "col"; ch1.title = (
        "Kurallara Göre Bulgu Sayısı" if DIL != "EN" else "Findings by Rule")
    veri = Reference(ws, min_col=K + 1, min_row=b1a, max_row=b1b)
    kat = Reference(ws, min_col=K, min_row=b1a, max_row=b1b)
    ch1.add_data(veri); ch1.set_categories(kat)
    ch1.legend = None; ch1.height = 8; ch1.width = 15
    renkler = ["3599B8", "F2C80F", "FD625E", "A66999"]
    s = ch1.series[0]
    for i, renk in enumerate(renkler[:b1b - b1a + 1]):
        dp = DataPoint(idx=i); dp.graphicalProperties.solidFill = renk
        s.data_points.append(dp)
    ws.add_chart(ch1, "A3")

    # 2) Doğru / Hatalı
    dogru = int((durum["Genel Durum"] == "DOĞRU").sum())
    hatali = len(durum) - dogru
    b2a, b2b = blok(8, "Durum", [
        ("Doğru" if DIL != "EN" else "OK", dogru),
        ("Hatalı" if DIL != "EN" else "Error", hatali)])
    ch2 = PieChart(); ch2.title = (
        "Veri Durumu" if DIL != "EN" else "Data Status")
    ch2.add_data(Reference(ws, min_col=K + 1, min_row=b2a, max_row=b2b))
    ch2.set_categories(Reference(ws, min_col=K, min_row=b2a, max_row=b2b))
    ch2.height = 8; ch2.width = 10
    for i, renk in enumerate(("0E8A6E", "D5484A")):
        dp = DataPoint(idx=i); dp.graphicalProperties.solidFill = renk
        ch2.series[0].data_points.append(dp)
    ws.add_chart(ch2, "J3")

    # 3-4) Öne çıkan iki kategori: hata % çubukları
    konum = ["A20", "J20"]
    sec = [k for k in ("Malzeme türü", "Raporlama Markası", "Mal grubu",
                       "S&OP Kategorisi") if k in kategoriler][:2]
    bas = 12
    for ci, kat_ad in enumerate(sec):
        g = kategoriler[kat_ad].head(8)
        satirlar = [(str(d), round(100 * s["Hatalı"] / s["Toplam"], 1)
                     if s["Toplam"] else 0) for d, s in g.iterrows()]
        ba, bb = blok(bas, kat_ad, satirlar); bas = bb + 2
        ch = BarChart(); ch.type = "bar"
        ch.title = (f"{kat_ad} — Hata %" if DIL != "EN"
                    else f"{kat_ad} — Error %")
        ch.add_data(Reference(ws, min_col=K + 1, min_row=ba, max_row=bb))
        ch.set_categories(Reference(ws, min_col=K, min_row=ba, max_row=bb))
        ch.legend = None; ch.height = 9; ch.width = 11
        ch.series[0].graphicalProperties.solidFill = "33628C"
        ws.add_chart(ch, konum[ci])

    # 5) Histogram — malzeme başına düşen hata sayısı dağılımı
    if "Hatalı Kural Sayısı" in durum.columns:
        dagilim = durum["Hatalı Kural Sayısı"].astype(int).value_counts().sort_index()
        satirlar = [(f"{int(k)} kural" if DIL != "EN" else f"{int(k)} rules",
                     int(v)) for k, v in dagilim.items()]
        ba, bb = blok(bas, "Hata sayısı dağılımı", satirlar)
        ch = BarChart(); ch.type = "col"; ch.gapWidth = 8
        ch.title = ("Malzeme başına hata sayısı (histogram)" if DIL != "EN"
                    else "Findings per material (histogram)")
        ch.add_data(Reference(ws, min_col=K + 1, min_row=ba, max_row=bb))
        ch.set_categories(Reference(ws, min_col=K, min_row=ba, max_row=bb))
        ch.legend = None; ch.height = 9; ch.width = 24
        ch.series[0].graphicalProperties.solidFill = "1B2A3A"
        ws.add_chart(ch, "A38")

    # Yardımcı sütunları gizle
    for col in ("AA", "AB"):
        ws.column_dimensions[col].hidden = True


# =====================================================================
# VERİ KALİTESİ BOYUTLARI (panel ve Power BI modeli aynı hesabı kullanır)
# =====================================================================

BOYUT_TANIM = [
    ("Tamlık", "Zorunlu alan ve üst kod bağı dolu mu",
     lambda d: ((d["Kural"] == "K1") & (d["İlgili Alan"] == "Temel malzeme")) |
               ((d["Kural"] == "K2") & d["Hata Nedeni"].astype(str).str.contains(
                   "doldurulmamış", na=False))),
    ("Tutarlılık", "Ürün alanları üst koduyla uyumlu mu",
     lambda d: (d["Kural"] == "K1") & (d["İlgili Alan"] != "Temel malzeme")),
    ("Geçerlilik", "Tanım ile ek veri alanı örtüşüyor mu",
     lambda d: d["Kural"] == "K2"),
    ("Standart", "Tanım adlandırma kuralına uygun mu",
     lambda d: d["Kural"] == "K3"),
    ("Doğruluk", "Tanımdaki çokluk MARM sayacıyla örtüşüyor mu",
     lambda d: d["Kural"] == "K4"),
    ("Benzersizlik", "Her üst kodun ayırt edici kombinasyonu tekil mi",
     lambda d: d["Kural"] == "K5"),
    ("Uyum", "Kodun son iki hanesi menşei ile örtüşüyor mu",
     lambda d: d["Kural"] == "K6"),
]

# Kural -> kalite boyutu eşlemesi (Power BI kural tablosunda kullanılır)
KURAL_BOYUT = {"K1": "Tutarlılık / Tamlık", "K2": "Geçerlilik",
               "K3": "Standart", "K4": "Doğruluk", "K5": "Benzersizlik",
               "K6": "Uyum"}
KURAL_RENK = {"K1": "3599B8", "K2": "E3A63B", "K3": "FD625E",
              "K4": "A66999", "K5": "118DFF", "K6": "0E8A7E"}
KURAL_ADI = {"K1": "Akıllı kod eşleşmesi", "K2": "Tanım ↔ ek veri alanı",
             "K3": "Yasaklı kelimeler", "K4": "Çokluk ↔ MARM sayacı",
             "K5": "Üst kod benzersizliği", "K6": "Kod sonu ↔ Menşei"}

# Panelde kural etiketinin üzerine gelindiğinde açılan açıklama metinleri
KURAL_ACIKLAMA = {
    "K1": ("Akıllı kod eşleşmesi — Ürün, bağlı olduğu üst kodla beyaz listedeki "
           "alanlarda (Raporlama Markası, S&OP Kategorisi, Ürün Boyutu, Kare "
           "Barkod, Ek Alan …) birebir aynı olmalıdır. Üst kodun boş olması veya "
           "listede bulunmaması da bulgudur. 68 ile başlayan malzemelerin hiç "
           "üst kodu olmamalıdır. Üst birim düzeyi silme işareti KASKAD "
           "denetlenir: bir üst kod silinmek üzere işaretliyse ona bağlı TÜM "
           "ürün kodları da silinmelidir ve hepsi listelenir — işareti "
           "konulmamış olanlar 'işaret eksik', konulmuş olanlar 'silme işlemi "
           "tamamlanmalı' notuyla. Ürün işaretli ama üst kod işaretsizse bu "
           "hata değildir. Tarih/saat alanları hiçbir zaman karşılaştırılmaz; "
           "SKU Grup boşsa hata sayılmaz."),
    "K2": ("Tanım ↔ ek veri alanı — Malzeme kısa metni T26 gibi yasak bir "
           "kodla bitemez; bittiği anda bulgu üretilir ve kodun çıkarıldığı "
           "düzeltilmiş tanım önerilir. Ayrıca metnin sonundaki kısaltma "
           "ile ilgili alan iki yönde tutarlı olmalıdır: tanımda LF varsa Ek "
           "Alan LF olmalı, Ek Alan LF ise tanım LF ile bitmelidir. Ülke "
           "kısaltmaları (RU, KZ, DE …) Pazar alanında aranır. Bitişik yazım "
           "(…200MLx24RU) ve birden çok kod aynı satırda ayrı ayrı denetlenir."),
    "K3": ("Yasaklı kelimeler — NEW, YENİ, YENI ile 2026 ve 2027 tanımın her "
           "yerinde yasaktır. 6 ve 7 ise yalnızca tanımın SONUNDA yasaklıdır; "
           "ortada geçen 6 ürün varyantıdır (ARKO CREAM 6 150X4X18 ML temizdir) "
           "ve KFR*6 gibi çarpanlar hata sayılmaz."),
    "K4": ("Çokluk ↔ MARM sayacı — Tanımın sonundaki çokluk (…125G*8X12 → 12), "
           "MARM tablosunda o malzeme koduna ait TÜM alternatif ölçü birimi "
           "satırlarındaki Sayaç değerlerinden herhangi biriyle eşleşmelidir "
           "(PAK, SR, PAL, KL, ADT… hepsi değerlendirilir). Hiçbiriyle "
           "eşleşmiyorsa tanım güncellenmiş ama MARM güncellenmemiş (veya "
           "tersi) demektir. MARM'da kaydı olmayan malzeme doğrulanamaz ve "
           "hata sayılmaz."),
    "K5": ("Üst kod benzersizliği — Her üst kodun (RC) ayırt edici alan "
           "kombinasyonu tekil olmalıdır: MARA'daki Marka1, Raporlama Markası, "
           "Raporlama Alt Markası, Varyant, Ürün Boyutu ve S&OP Kategorisi "
           "(altı alan) ile MARM'dan doğrulanan koli içi adet — toplam yedi "
           "bileşen — birlikte değerlendirilir. Yalnızca dosyada mevcut olan "
           "alanlar kullanılır. İki farklı üst kodda bu değerlerin tamamı birebir "
           "aynıysa kodlardan biri gereksiz demektir ve her iki kod da bulgu "
           "olarak listelenir. Ayrıca ÜB dzy.silme iştr. alanı dolu (X) olan "
           "malzemeler silinmiş sayılır ve hiçbir kuralda bulgu üretmez."),
    "K6": ("Kod sonu ↔ Menşei — Ürün kodunun son iki hanesi, Menşei "
           "alanındaki sayısal kodla aynı olmalıdır (10=TR, 11=TR VAS, "
           "20=MY, 21=MY VAS, 30=RU, 31=RU VAS, 40=EG, 41=EG VAS, 50=ID, "
           "99=TG). Örneğin 60002810 kodunun menşei 10 (TR) olmalıdır; "
           "alanda 21 yazıyorsa uyuşmazlık bulgusu üretilir. Menşei alanı "
           "BOŞ bırakılmış malzemeler de bulgu olarak listelenir. Kural "
           "yalnızca sayısal ürün kodlarına uygulanır; üst kodlar (RC…) "
           "kapsam dışıdır."),
}


def boyut_puanlari(basit, durum, toplam):
    """[(boyut, açıklama, puan, etkilenen malzeme)] listesi döndürür."""
    sonuc = []
    for ad_b, aciklama, kosul in BOYUT_TANIM:
        etkilenen = 0
        if not basit.empty:
            etkilenen = basit.loc[kosul(basit), "Malzeme"].astype(str).nunique()
        puan = 100 * (1 - etkilenen / toplam) if toplam else 100.0
        sonuc.append((ad_b, aciklama, puan, etkilenen))
    return sonuc


# =====================================================================
# SONUÇ DOSYASI BİÇİMLENDİRİCİLERİ (özet / genel durum / kategori)
# =====================================================================

PIKTO = "█"          # in-cell çubuk karakteri
PIKTO_BOS = "░"


def _pikto(oran, uzunluk=10):
    """Yüzdeyi karakter çubuğuna çevirir (piktogram)."""
    dolu = int(round(uzunluk * max(0.0, min(oran, 100.0)) / 100))
    return PIKTO * dolu + PIKTO_BOS * (uzunluk - dolu)


def ozet_sayfasi_bicimle(wb, ozet, durum, kapsam_disi=0, istisna_n=0,
                         hazir=0):
    """Özet sayfasını KPI blokları + toplamlı kural tablosu olarak yeniden yazar."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    LACI, TEAL, MERCAN, GRI = "1B2A3A", "0E8A6E", "C0392B", "5F6B6D"
    RENK = {"K1": "3599B8", "K2": "D9A400", "K3": "FD625E", "K4": "A66999"}
    ince = Side(style="thin", color="DFE6EA")

    ad = "Summary" if DIL == "EN" else "Özet"
    if ad in wb.sheetnames:
        wb.remove(wb[ad])
    ws = wb.create_sheet(ad, 1)
    ws.sheet_view.showGridLines = False

    toplam_m = len(durum)
    dogru = int((durum["Genel Durum"] == "DOĞRU").sum())
    hatali = toplam_m - dogru
    saglik = 100 * dogru / toplam_m if toplam_m else 0
    toplam_bulgu = int(ozet["Bulgu Sayısı"].sum())
    bilgi_kayit = int(ozet.loc[ozet["Kural"].astype(str).str.startswith("K4"),
                               "Bulgu Sayısı"].sum())
    hata_bulgu = toplam_bulgu - bilgi_kayit

    # ---- Başlık ----
    ws.merge_cells("B2:I2")
    b = ws["B2"]
    b.value = ("Malzeme Veri Kalitesi — Yönetici Özeti" if DIL != "EN"
               else "Material Data Quality — Executive Summary")
    b.font = Font(bold=True, size=16, color=LACI)
    ws.merge_cells("B3:I3")
    import datetime
    ws["B3"] = (f"SAP MM · {len(ozet)} kural · çalıştırma: "
                f"{datetime.date.today().strftime('%d.%m.%Y')}")
    ws["B3"].font = Font(size=10, color=GRI)

    # ---- KPI blokları (2 sıra x 4) ----
    kpi = [
        (L("Toplam Bulgu"), toplam_bulgu, MERCAN, "hata + bilgi kaydı"),
        (L("Hata Bulgusu"), hata_bulgu, MERCAN, "K1 + K2 + K3"),
        (L("Hatalı Malzeme"), hatali, MERCAN, "en az bir kurala takılan"),
        (L("Veri Sağlığı"), f"%{saglik:.1f}", TEAL, "kurallardan geçen oranı"),
        (L("Taranan Malzeme"), toplam_m, LACI, "ürün + üst kod"),
        (L("Bilgi Kaydı"), bilgi_kayit, "6B4FA1", "K4 alt tarif — hata değil"),
        (L("Kapsam Dışı"), kapsam_disi, GRI, "analiz edilmedi"),
        (L("Hazır Düzeltme"), hazir, TEAL, "program önerisi (K2)"),
    ]
    if DIL == "EN":
        alt_en = ["errors + info", "K1 + K2 + K3", "at least one finding",
                  "share passing all rules", "products + parents",
                  "K4 sub-descriptions", "not analysed", "auto suggestions (K2)"]
        kpi = [(k[0], k[1], k[2], alt_en[i]) for i, k in enumerate(kpi)]

    for i, (etiket, deger, renk, alt) in enumerate(kpi):
        sut = 2 + (i % 4) * 2
        sat = 5 + (i // 4) * 4
        ws.merge_cells(start_row=sat, start_column=sut,
                       end_row=sat, end_column=sut + 1)
        c = ws.cell(row=sat, column=sut, value=tr_upper(etiket))
        c.font = Font(bold=True, size=9, color=GRI)
        ws.merge_cells(start_row=sat + 1, start_column=sut,
                       end_row=sat + 1, end_column=sut + 1)
        d = ws.cell(row=sat + 1, column=sut, value=deger)
        d.font = Font(bold=True, size=20, color=renk)
        d.alignment = Alignment(vertical="center")
        ws.row_dimensions[sat + 1].height = 28
        ws.merge_cells(start_row=sat + 2, start_column=sut,
                       end_row=sat + 2, end_column=sut + 1)
        e = ws.cell(row=sat + 2, column=sut, value=alt)
        e.font = Font(size=9, color=GRI)
        for r in range(sat, sat + 3):
            for cc in range(sut, sut + 2):
                ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor="F4F7F9")
        ws.cell(row=sat, column=sut).border = Border(top=Side(style="medium", color=renk))

    # ---- Kural tablosu ----
    bas = 14
    ws.merge_cells(start_row=bas, start_column=2, end_row=bas, end_column=7)
    t = ws.cell(row=bas, column=2,
                value=("KURAL BAZINDA DAĞILIM" if DIL != "EN" else "BREAKDOWN BY RULE"))
    t.font = Font(bold=True, size=11, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=LACI)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[bas].height = 20

    basliklar = [L("Kural"), L("Bulgu Sayısı"), L("Etkilenen Malzeme Sayısı"),
                 L("Pay %"), L("Dağılım"), L("Tür")]
    if DIL == "EN":
        basliklar = ["Rule", "Findings", "Materials", "Share %", "Distribution", "Type"]
    for j, adx in enumerate(basliklar):
        c = ws.cell(row=bas + 1, column=2 + j, value=adx)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="33628C")
        c.alignment = Alignment(horizontal="center" if j else "left")

    r = bas + 2
    for _, s in ozet.iterrows():
        kural_ad = str(s["Kural"])
        kod = kural_ad[:2]
        bulgu = int(s["Bulgu Sayısı"])
        malz = int(s["Etkilenen Malzeme Sayısı"])
        pay = 100 * bulgu / toplam_bulgu if toplam_bulgu else 0
        bilgi_mi = kod == "K4"
        satir = [kural_ad, bulgu, malz, round(pay, 1), _pikto(pay),
                 (L("Bilgi") if bilgi_mi else L("Hata")) if DIL != "EN"
                 else ("Info" if bilgi_mi else "Error")]
        for j, v in enumerate(satir):
            c = ws.cell(row=r, column=2 + j, value=v)
            c.border = Border(bottom=ince)
            if r % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F7FAFB")
            if j == 0:
                c.font = Font(bold=True, color=RENK.get(kod, LACI))
            elif j == 4:
                c.font = Font(name="Consolas", color=RENK.get(kod, LACI))
            elif j == 5:
                c.font = Font(size=9, color="6B4FA1" if bilgi_mi else MERCAN)
                c.alignment = Alignment(horizontal="center")
            else:
                c.alignment = Alignment(horizontal="center")
        r += 1

    # ---- TOPLAM satırı ----
    toplam_satir = [L("TOPLAM") if DIL != "EN" else "TOTAL", toplam_bulgu,
                    hatali, 100.0, _pikto(100),
                    (f"{hata_bulgu} hata + {bilgi_kayit} bilgi" if DIL != "EN"
                     else f"{hata_bulgu} errors + {bilgi_kayit} info")]
    for j, v in enumerate(toplam_satir):
        c = ws.cell(row=r, column=2 + j, value=v)
        c.font = Font(bold=True, size=11,
                      color="FFFFFF" if j < 5 else "FFFFFF")
        c.fill = PatternFill("solid", fgColor=LACI)
        c.alignment = Alignment(horizontal="left" if j == 0 else "center")
        if j == 4:
            c.font = Font(name="Consolas", bold=True, color="FFFFFF")
    ws.row_dimensions[r].height = 20
    not_r = r + 2
    ws.merge_cells(start_row=not_r, start_column=2, end_row=not_r, end_column=7)
    ws.cell(row=not_r, column=2, value=(
        "Not: Etkilenen malzeme sayısı toplamı, bir malzemenin birden çok kurala "
        "takılabilmesi nedeniyle kural satırlarının toplamından küçüktür."
        if DIL != "EN" else
        "Note: a material may match several rules, so the material total is lower "
        "than the sum of the rule rows.")).font = Font(size=9, italic=True, color=GRI)
    if istisna_n:
        ws.cell(row=not_r + 1, column=2, value=(
            f"{istisna_n} bulgu onaylı istisna olarak listeden düşüldü."
            if DIL != "EN" else
            f"{istisna_n} findings excluded as approved exceptions.")
        ).font = Font(size=9, italic=True, color=GRI)

    genislik = {"A": 2, "B": 30, "C": 13, "D": 22, "E": 10, "F": 14, "G": 18}
    for k, v in genislik.items():
        ws.column_dimensions[k].width = v
    for k in ("H", "I"):
        ws.column_dimensions[k].width = 12
    return ws


def genel_durum_bicimle(wb, durum):
    """Genel_Durum sayfasına gruplandırılmış başlık şeridi ve biçim ekler."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ad = "Overall" if DIL == "EN" else "Genel_Durum"
    ad = ad if ad in wb.sheetnames else "Genel_Durum"
    if ad not in wb.sheetnames:
        return
    ws = wb[ad]
    ws.sheet_view.showGridLines = False
    kolonlar = [c.value for c in ws[1]]
    n = len(kolonlar)

    ws.insert_rows(1, 2)          # 1: sayfa başlığı, 2: grup şeridi
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    t = ws.cell(row=1, column=1, value=(
        "MALZEME BAZINDA DURUM — her satır bir malzemenin tüm kural sonuçlarıdır"
        if DIL != "EN" else
        "STATUS BY MATERIAL — each row shows every rule result for one material"))
    t.font = Font(bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1B2A3A")
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    # Grupları sütun adlarından çıkar
    gruplar, mevcut, bas = [], None, 1
    for i, k in enumerate(kolonlar, start=1):
        ks = str(k)
        if ks.startswith("K4"):
            g = "BİLGİ" if DIL != "EN" else "INFO"
        elif ks[:2] in ("K1", "K2", "K3"):
            g = "KURAL SONUÇLARI" if DIL != "EN" else "RULE RESULTS"
        elif ks in ("Hatalı Kural Sayısı", "Genel Durum", "Faulty Rule Count",
                    "Overall Status"):
            g = "SONUÇ" if DIL != "EN" else "RESULT"
        else:
            g = "MALZEME" if DIL != "EN" else "MATERIAL"
        if g != mevcut:
            if mevcut is not None:
                gruplar.append((mevcut, bas, i - 1))
            mevcut, bas = g, i
    gruplar.append((mevcut, bas, n))

    grup_renk = {"MALZEME": "5F6B6D", "MATERIAL": "5F6B6D",
                 "KURAL SONUÇLARI": "33628C", "RULE RESULTS": "33628C",
                 "BİLGİ": "6B4FA1", "INFO": "6B4FA1",
                 "SONUÇ": "0E8A6E", "RESULT": "0E8A6E"}
    for g, b_, s_ in gruplar:
        if s_ > b_:
            ws.merge_cells(start_row=2, start_column=b_, end_row=2, end_column=s_)
        c = ws.cell(row=2, column=b_, value=g)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=grup_renk.get(g, "5F6B6D"))
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 17

    ince = Side(style="thin", color="E4E9EE")
    for j in range(1, n + 1):
        c = ws.cell(row=3, column=j)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1B2A3A")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    ws.row_dimensions[3].height = 30

    yesil = Font(color="0E8A6E", bold=True)
    kirmizi = Font(color="C0392B", bold=True)
    mor = Font(color="6B4FA1", bold=True)
    for r in range(4, ws.max_row + 1):
        for j in range(1, n + 1):
            c = ws.cell(row=r, column=j)
            c.border = Border(bottom=ince)
            if r % 2 == 1:
                c.fill = PatternFill("solid", fgColor="F7FAFB")
            v = str(c.value)
            if v in ("DOĞRU", "OK"):
                c.font = yesil
                c.alignment = Alignment(horizontal="center")
            elif v in ("HATALI", "ERROR"):
                c.font = kirmizi
                c.alignment = Alignment(horizontal="center")
            elif v in ("ALT TARİF", "SUB"):
                c.font = mor
                c.alignment = Alignment(horizontal="center")
            elif v in ("-", "0"):
                c.alignment = Alignment(horizontal="center")

    for j, k in enumerate(kolonlar, start=1):
        harf = get_column_letter(j)
        ks = str(k)
        if ks in ("Tanım", "Description"):
            ws.column_dimensions[harf].width = 38
        elif ks in ("Malzeme", "Material"):
            ws.column_dimensions[harf].width = 13
        elif ks in ("Kaynak", "Source"):
            ws.column_dimensions[harf].width = 11
        else:
            ws.column_dimensions[harf].width = 15
    ws.freeze_panes = "D4"
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}{ws.max_row}"


def kategori_sayfasi_duzenle(wb, kategoriler):
    """Kategori_Analizi sayfasını iki sütunlu, başlıklı bloklar hâlinde yazar."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ad = "Kategori_Analizi" if DIL != "EN" else "Category_Analysis"
    for s in ("Kategori_Analizi", "Category_Analysis"):
        if s in wb.sheetnames:
            wb.remove(wb[s])
    ws = wb.create_sheet(ad)
    ws.sheet_view.showGridLines = False

    LACI, MAVI, GRI = "1B2A3A", "33628C", "5F6B6D"
    ince = Side(style="thin", color="E4E9EE")

    ws.merge_cells("A1:K1")
    t = ws.cell(row=1, column=1, value=(
        "KATEGORİ ANALİZİ — her kategoride hangi değerlerde ne kadar hata var"
        if DIL != "EN" else
        "CATEGORY ANALYSIS — error concentration by value in each category"))
    t.font = Font(bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=LACI)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24
    ws.merge_cells("A2:K2")
    ws.cell(row=2, column=1, value=(
        "Hata % = o değerdeki hatalı malzeme / toplam malzeme · çubuk uzunluğu oranı gösterir"
        if DIL != "EN" else
        "Error % = faulty materials / total materials for that value · bar length shows the ratio")
    ).font = Font(size=9, italic=True, color=GRI)

    basliklar = ([L("Değer"), L("Toplam"), L("Hatalı"), L("Hata %"), L("Dağılım")]
                 if DIL != "EN" else
                 ["Value", "Total", "Faulty", "Error %", "Distribution"])

    kat_listesi = list(kategoriler.items())
    satir = 4
    for i in range(0, len(kat_listesi), 2):
        band = kat_listesi[i:i + 2]
        yukseklik = 0
        for j, (kategori, gruplar) in enumerate(band):
            sut = 1 + j * 6                      # A.. ve G..
            ws.merge_cells(start_row=satir, start_column=sut,
                           end_row=satir, end_column=sut + 4)
            h = ws.cell(row=satir, column=sut, value=tr_upper(str(kategori)))
            h.font = Font(bold=True, size=10.5, color="FFFFFF")
            h.fill = PatternFill("solid", fgColor=MAVI)
            h.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[satir].height = 19

            for k, adx in enumerate(basliklar):
                c = ws.cell(row=satir + 1, column=sut + k, value=adx)
                c.font = Font(bold=True, size=9, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=LACI)
                c.alignment = Alignment(horizontal="center" if k else "left")

            r = satir + 2
            for deger, s in gruplar.iterrows():
                oran = 100 * s["Hatalı"] / s["Toplam"] if s["Toplam"] else 0
                degerler = [str(deger), int(s["Toplam"]), int(s["Hatalı"]),
                            round(oran, 1), _pikto(oran)]
                for k, v in enumerate(degerler):
                    c = ws.cell(row=r, column=sut + k, value=v)
                    c.border = Border(bottom=ince)
                    if r % 2 == 1:
                        c.fill = PatternFill("solid", fgColor="F7FAFB")
                    if k == 4:
                        c.font = Font(name="Consolas", size=9,
                                      color="C0392B" if oran >= 60 else MAVI)
                    elif k:
                        c.alignment = Alignment(horizontal="center")
                        if k == 3 and oran >= 60:
                            c.font = Font(color="C0392B", bold=True)
                r += 1

            top_t = int(gruplar["Toplam"].sum())
            top_h = int(gruplar["Hatalı"].sum())
            top_o = 100 * top_h / top_t if top_t else 0
            for k, v in enumerate((L("TOPLAM") if DIL != "EN" else "TOTAL",
                                   top_t, top_h, round(top_o, 1), _pikto(top_o))):
                c = ws.cell(row=r, column=sut + k, value=v)
                c.font = (Font(name="Consolas", bold=True, color="FFFFFF")
                          if k == 4 else Font(bold=True, color="FFFFFF"))
                c.fill = PatternFill("solid", fgColor=LACI)
                if k:
                    c.alignment = Alignment(horizontal="center")
            yukseklik = max(yukseklik, r - satir + 1)
        satir += yukseklik + 2

    for j in range(2):
        sut = 1 + j * 6
        ws.column_dimensions[get_column_letter(sut)].width = 26
        for k in (1, 2, 3):
            ws.column_dimensions[get_column_letter(sut + k)].width = 9
        ws.column_dimensions[get_column_letter(sut + 4)].width = 13
        if j == 0:
            ws.column_dimensions[get_column_letter(sut + 5)].width = 3
    ws.freeze_panes = "A4"
    return ws


# =====================================================================
# POWER BI VERİ MODELİ (yıldız şema + hazır DAX ölçüleri)
# =====================================================================

CIKTI_PBI_MODEL = "powerbi_veri_modeli.xlsx"
CIKTI_PBI_REHBER = "powerbi_rehber.md"


def powerbi_modeli_uret(ham, durum, ozet, kategoriler, urun, ust,
                        basit=None, kapsam_disi=0, istisna_n=0, hazir=0):
    """Power BI Desktop'ın doğrudan içe alabileceği yıldız şema üretir.
    Her sayfa gerçek bir Excel Tablosu olarak yazılır; Power BI bunları
    adlandırılmış tablo olarak görür ve ilişkileri kurmak kolaylaşır."""
    import datetime
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    toplam = len(durum)
    dogru = int((durum["Genel Durum"] == "DOĞRU").sum())
    hatali = toplam - dogru
    tarih = datetime.date.today()

    wb = Workbook()
    wb.remove(wb.active)

    def tablo_yaz(ad, df, tablo_adi):
        ws = wb.create_sheet(ad)
        ws.append(list(df.columns))
        for c in ws[1]:
            c.font = Font(bold=True)
        for _, s in df.iterrows():
            ws.append([("" if pd.isna(v) else v) for v in s.tolist()])
        son = f"{get_column_letter(len(df.columns))}{max(len(df) + 1, 2)}"
        t = Table(displayName=tablo_adi, ref=f"A1:{son}")
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9",
                                          showRowStripes=True)
        ws.add_table(t)
        for j, kol in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(j)].width = \
                min(max(len(str(kol)) + 4, 12), 46)
        return ws

    # ---- 1) FACT: Bulgular ----
    basit = basit if basit is not None else ham
    f = ham.copy() if not ham.empty else pd.DataFrame(
        columns=["Malzeme", "Üst Kod", "Kural", "Hata Nedeni", "İlgili Alan",
                 "Mevcut Değer", "Olması Gereken", "Yapılacak Düzeltme"])
    if not f.empty:
        kaynak_map = dict(zip(durum["Malzeme"].astype(str), durum["Kaynak"]))
        f["Kaynak"] = f["Malzeme"].astype(str).map(kaynak_map).fillna("Ürün Kodu")
        # Çok satırlı hücreleri Power BI'da okunur tutmak için " | " ile birleştir
        for kol in ("Hata Nedeni", "İlgili Alan", "Olması Gereken",
                    "Yapılacak Düzeltme"):
            if kol in f.columns:
                f[kol] = f[kol].astype(str).str.replace("\n", " | ", regex=False)
        f["Hata Sayısı"] = 1
        f["Çalıştırma Tarihi"] = tarih
    tablo_yaz("Bulgular", f, "Bulgular")

    # ---- 2) DIM: Malzemeler ----
    d = durum.copy()
    marka = {}
    for df_ in (urun, ust):
        tk = [c for c in df_.columns if c in TANIM_SUTUN_ADAYLARI]
        if tk:
            marka.update(dict(zip(df_["Malzeme"].astype(str),
                                  df_[tk[0]].astype(str).str.split().str[0])))
    tur = {}
    for df_ in (urun, ust):
        if "Malzeme türü" in df_.columns:
            tur.update(dict(zip(df_["Malzeme"].astype(str),
                                df_["Malzeme türü"].astype(str))))
    d["Marka"] = d["Malzeme"].astype(str).map(marka).fillna("")
    d["Malzeme Türü"] = d["Malzeme"].astype(str).map(tur).fillna("")
    d["Hatalı mı"] = (d["Genel Durum"] != "DOĞRU").astype(int)
    tablo_yaz("Malzemeler", d, "Malzemeler")

    # ---- 3) DIM: Kurallar ----
    kurallar = pd.DataFrame([{
        "Kural": str(s["Kural"])[:2],
        "Kural Adı": KURAL_ADI.get(str(s["Kural"])[:2], str(s["Kural"])),
        "Tip": "Bilgi" if str(s["Kural"]).startswith("K4") else "Hata",
        "Kalite Boyutu": KURAL_BOYUT.get(str(s["Kural"])[:2], ""),
        "Renk": "#" + KURAL_RENK.get(str(s["Kural"])[:2], "5F6B6D"),
        "Bulgu Sayısı": int(s["Bulgu Sayısı"]),
        "Etkilenen Malzeme": int(s["Etkilenen Malzeme Sayısı"]),
    } for _, s in ozet.iterrows()])
    tablo_yaz("Kurallar", kurallar, "Kurallar")

    # ---- 4) Kalite boyutları ----
    boyut_df = pd.DataFrame(
        [{"Boyut": a, "Açıklama": b, "Puan": round(p, 1),
          "Etkilenen Malzeme": n,
          "Durum": ("İyi" if p >= 95 else ("Dikkat" if p >= 85 else "Kritik"))}
         for a, b, p, n in boyut_puanlari(basit, durum, toplam)])
    tablo_yaz("Kalite_Boyutlari", boyut_df, "Kalite_Boyutlari")

    # ---- 5) Kategori analizi (uzun format) ----
    kat_satir = []
    for kategori, gruplar in kategoriler.items():
        for deger, s in gruplar.iterrows():
            kat_satir.append({
                "Kategori": kategori, "Değer": str(deger),
                "Toplam": int(s["Toplam"]), "Hatalı": int(s["Hatalı"]),
                "Hata %": round(100 * s["Hatalı"] / s["Toplam"], 1)
                if s["Toplam"] else 0.0})
    tablo_yaz("Kategoriler", pd.DataFrame(kat_satir), "Kategoriler")

    # ---- 6) Çalıştırma özeti (KPI kartları için tek satır) ----
    meta = pd.DataFrame([{
        "Çalıştırma Tarihi": tarih,
        "Taranan Malzeme": toplam,
        "Doğru": dogru, "Hatalı": hatali,
        "Veri Sağlığı %": round(100 * dogru / toplam, 1) if toplam else 0,
        "Toplam Bulgu": int(ozet["Bulgu Sayısı"].sum()),
        "Kapsam Dışı Kayıt": kapsam_disi,
        "Onaylı İstisna": istisna_n,
        "Hazır Düzeltme": hazir,
    }])
    tablo_yaz("Calistirma", meta, "Calistirma")

    wb.save(CIKTI_PBI_MODEL)
    return boyut_df


def powerbi_rehber_yaz(dosya=None):
    """Power BI Desktop kurulum rehberi + kopyala-yapıştır DAX ölçüleri."""
    dosya = dosya or CIKTI_PBI_REHBER
    metin = """# Power BI'da Raporu Açma Rehberi

İki yol var. **A yolu** hazır şablonu açmaktır (önerilen, 1 dakika).
**B yolu** modeli elle kurmaktır (şablon açılmazsa veya raporu sıfırdan
kendin tasarlamak istersen).

---

# A YOLU — Hazır şablon (.pbit)

1. `malzeme_veri_kalitesi.pbit` dosyasına çift tıkla (Power BI Desktop açar)
2. Açılan pencerede **Yükle** de — veri şablonun içinde gömülü olduğu için
   hiçbir dosya yolu veya bağlantı sorulmaz
3. Rapor üç sayfa hâlinde hazır gelir:
   **1 · Kontrol Paneli**, **2 · Analiz**, **3 · Düzeltme Listesi**
4. İsteğe bağlı: **Görünüm > Temalar > Tema ara** ile `powerbi_tema.json`
   dosyasını seç — görseller HTML panelle aynı renklere döner

Şablonun içinde neler hazır geliyor:

- Altı tablo (Bulgular, Malzemeler, Kurallar, Kalite_Boyutlari,
  Kategoriler, Calistirma) ve veriler
- İki ilişki (Bulgular → Malzemeler, Bulgular → Kurallar)
- On DAX ölçüsü (Veri Sağlığı %, Hata Bulgusu, Hazır Düzeltme ...)
- KPI kartları, boyut çubukları, kural sütunları, halka grafik,
  kategori matrisi ve dilimleyicilerle üç sayfa

> Şablon her Python çalıştırmasında güncel veriyle yeniden üretilir.
> Yeni veriyi göstermek için yeni `.pbit` dosyasını açman yeterlidir.

---

# B YOLU — Tarayıcıda (app.powerbi.com), kurulum gerekmez

`.pbit` dosyası **yalnızca Power BI Desktop** ile açılır; tarayıcıdaki
Power BI (app.powerbi.com) şablon dosyası açamaz. Bilgisayara program
kuramıyorsan bu yolu kullan:

1. Tarayıcıda **app.powerbi.com** > sol menüden **Oluştur**
2. **Excel** kutucuğunu seç
3. `powerbi_tarayici.xlsx` dosyasını yükle
4. Gelen listede dört tabloyu işaretle: `Bulgular`, `Kalite_Boyutlari`,
   `Kategoriler`, `Ozet`
5. Görselleri sağdaki alan listesinden sürükleyerek kur

Bu dosyada **ilişki kurmaya gerek yoktur**: malzeme ve kural özellikleri
(Tanım, Marka, Malzeme Türü, Genel Durum, Kural Adı, Tip, Kalite Boyutu)
doğrudan `Bulgular` tablosuna eklenmiştir. Tarayıcıda ölçü yazmak
gerekmez, şu sütunlarla aynı sonuçlara ulaşırsın:

| İstediğin | Nasıl |
|---|---|
| Toplam bulgu | `Hata Sayısı` toplamı |
| Hatalı malzeme | `Hatalı mı` alanını `Malzeme Tekil` ile filtreleyip topla |
| Taranan malzeme | `Malzeme Tekil` toplamı |
| Veri sağlığı | `Genel Durum` alanıyla halka grafik |
| Boyut puanları | `Kalite_Boyutlari` tablosu: Boyut × Puan çubuk grafiği |
| Kategori analizi | `Kategoriler` tablosu: matris görseli |
| Düzeltme listesi | `Bulgular` tablosu: tablo görseli, dilimleyici olarak Kural/Kaynak/İlgili Alan |

---

# C YOLU — Modeli Desktop'ta elle kurmak

## 1. Veriyi içe al

1. Power BI Desktop > **Veri al** > **Excel çalışma kitabı**
2. `powerbi_veri_modeli.xlsx` dosyasını seç
3. Gelen listede **tablo** simgeli altı öğeyi işaretle (sayfa simgeli
   olanları değil): `Bulgular`, `Malzemeler`, `Kurallar`,
   `Kalite_Boyutlari`, `Kategoriler`, `Calistirma`
4. **Yükle**

## 2. İlişkileri kur

**Model** görünümüne geç ve şu iki ilişkiyi oluştur (sürükle-bırak):

| Kaynak | Hedef | Yön | Kardinalite |
|---|---|---|---|
| `Bulgular[Malzeme]` | `Malzemeler[Malzeme]` | tek yön | çoktan bire (*:1) |
| `Bulgular[Kural]` | `Kurallar[Kural]` | tek yön | çoktan bire (*:1) |

`Kalite_Boyutlari`, `Kategoriler` ve `Calistirma` bağımsız tablolardır;
ilişki gerekmez.

## 3. Ölçüleri (measure) ekle

`Bulgular` tablosunu seç > **Yeni ölçü** > aşağıdaki her bloğu tek tek
yapıştır. (Fact tablosunda her bulgu ayrı satırdır: `COUNTROWS(Bulgular)`
kural sayılarının toplamına eşittir — K4 dahil.)

```dax
Toplam Bulgu = COUNTROWS(Bulgular)
```
```dax
Hata Bulgusu =
CALCULATE(COUNTROWS(Bulgular), Kurallar[Tip] = "Hata")
```
```dax
Hatalı Malzeme =
CALCULATE(COUNTROWS(Malzemeler), Malzemeler[Hatalı mı] = 1)
```
```dax
Bilgi Kaydı =
CALCULATE(COUNTROWS(Bulgular), Kurallar[Tip] = "Bilgi")
```
```dax
Taranan Malzeme = COUNTROWS(Malzemeler)
```
```dax
Veri Sağlığı % =
DIVIDE(
    CALCULATE(COUNTROWS(Malzemeler), Malzemeler[Hatalı mı] = 0),
    COUNTROWS(Malzemeler)
) * 100
```
```dax
Hata Oranı % = 100 - [Veri Sağlığı %]
```
```dax
Hazır Düzeltme =
CALCULATE(
    COUNTROWS(Bulgular),
    Bulgular[Kural] = "K2",
    CONTAINSSTRING(Bulgular[Yapılacak Düzeltme], "Tanımın sonuna")
)
```
```dax
Ortalama Boyut Puanı = AVERAGE(Kalite_Boyutlari[Puan])
```
```dax
Düzeltme Etkisi % =
-- Yalnızca seçili kuralın hataları giderilirse veri sağlığı ne olur
VAR SeciliKural = SELECTEDVALUE(Kurallar[Kural])
VAR SadeceBuKural =
    CALCULATE(
        DISTINCTCOUNT(Bulgular[Malzeme]),
        Kurallar[Kural] = SeciliKural,
        Kurallar[Tip] = "Hata"
    )
VAR Dogru = CALCULATE(COUNTROWS(Malzemeler), Malzemeler[Hatalı mı] = 0)
RETURN DIVIDE(Dogru + SadeceBuKural, COUNTROWS(Malzemeler)) * 100
```

## 4. Görselleri kur (HTML panelin karşılığı)

### Sayfa 1 — Kontrol Paneli

| Görsel | Alanlar | Not |
|---|---|---|
| 5 × **Kart** | `Veri Sağlığı %`, `Toplam Bulgu`, `Hatalı Malzeme`, `Taranan Malzeme`, `Hazır Düzeltme` | Üst şerit; kart başlıklarını küçük ve gri yap |
| **Yığılmış çubuk** | Eksen: `Kalite_Boyutlari[Boyut]` · Değer: `Puan` | Veri etiketleri açık, X ekseni 0-100 |
| **Kümelenmiş sütun** | Eksen: `Kurallar[Kural]` · Değer: `Toplam Bulgu` | Renkleri `Kurallar[Renk]` sütununa göre elle ata |
| **Halka (donut)** | Açıklama: `Malzemeler[Genel Durum]` · Değer: `Taranan Malzeme` | Yeşil = DOĞRU, kırmızı = HATALI |
| **Tablo** | `Kurallar[Kural]`, `Kural Adı`, `Bulgu Sayısı`, `Etkilenen Malzeme`, `Tip` | Kural özeti |

### Sayfa 2 — Analiz

| Görsel | Alanlar |
|---|---|
| **Yığılmış sütun** | Eksen: `Malzemeler[Marka]` · Açıklama: `Kurallar[Kural]` · Değer: `Toplam Bulgu` |
| **Kümelenmiş çubuk** | Eksen: `Bulgular[İlgili Alan]` · Değer: `Toplam Bulgu` (ilk 6 için üst N filtresi) |
| **Matris** | Satır: `Kategoriler[Kategori]`, `Değer` · Değer: `Toplam`, `Hatalı`, `Hata %` |
| **Ölçüm (gauge)** | Değer: `Veri Sağlığı %` · Hedef: 95 |

### Sayfa 3 — Düzeltme Listesi

**Tablo** görseli, sütun sırası:
`Malzemeler[Tanım]` · `Bulgular[Malzeme]` · `Üst Kod` · `Kural` ·
`Hata Nedeni` · `İlgili Alan` · `Mevcut Değer` · `Olması Gereken` ·
`Alt Tarif Sayısı` · `Yapılacak Düzeltme`

## 5. Dilimleyiciler (panelin sol filtre paneli)

Her sayfaya ekle:

- `Kurallar[Kural]` — döşeme (tile) biçiminde
- `Bulgular[Kaynak]` — Ürün Kodu / Üst Kod
- `Bulgular[İlgili Alan]` — açılır liste
- `Malzemeler[Marka]` — arama kutulu liste

## 6. Renk paleti (panelle aynı görünmesi için)

| Öğe | Hex |
|---|---|
| K1 | `#3599B8` |
| K2 | `#E3A63B` |
| K3 | `#FD625E` |
| K4 (bilgi) | `#A66999` |
| Olumlu / doğru | `#01B8AA` |
| Olumsuz / hatalı | `#FD625E` |
| Metin — koyu | `#1B2A3A` |
| Zemin | `#F3F5F8` |

**Görünüm** > **Temalar** > **Geçerli temayı özelleştir** menüsünden bu
renkleri sırayla girersen tüm görseller otomatik uyumlu olur.

## 7. Koşullu biçimlendirme (eşikler)

`Kalite_Boyutlari[Puan]` ve `Hata %` sütunlarında:
**Biçim** > **Hücre öğeleri** > **Arka plan rengi** > *Kurallar*

- `>= 95` → yeşil `#01B8AA`
- `85 – 95` → sarı `#E3A63B`
- `< 85` → kırmızı `#FD625E`

## 8. Yenileme

Python programı her çalıştığında `powerbi_veri_modeli.xlsx` yeniden
üretilir. Power BI'da **Giriş** > **Yenile** demek yeterlidir; görseller
ve ölçüler korunur. Dosyayı paylaşılan bir klasöre koyup Power BI
Service'e yayınlarsan zamanlanmış yenileme de kurulabilir
(**Ağ geçidi** gerektirir).

## Notlar

- Model dosyası yalnızca **sonuçları** taşır, ham SAP verisini taşımaz;
  bu yüzden paylaşımı güvenlidir.
- `Bulgular` tablosunda her bulgu **ayrı satırdır**: aynı malzemenin hem
  `RU` hem `LF` hatası varsa iki satır gelir. HTML panelde bunlar tek
  satırda birleştirilmiş görünür; Power BI'da ayrı satır olması
  filtreleme ve sayım için daha doğrudur.
- `Alt Tarif Sayısı` sütunu, o malzemenin tanımını kaç kodun paylaştığını
  gösterir (K4 bilgisi). 1 = benzersiz.
- Aynı klasördeki `kalite_kontrol_sonuc.xlsx` ve `powerbi_bulgular.xlsx`
  dosyalarını da içe alabilirsin, ancak yıldız şema bu model dosyasında
  hazır olduğu için önerilen kaynak budur.
"""
    with open(dosya, "w", encoding="utf-8") as f:
        f.write(metin)


# =====================================================================
# OTOMATİK DÜZELTME — bulguların uygulanmış hâli
# =====================================================================

def duzeltilmis_veri_uret(urun, ust, k1, k2, k3, k4, k5,
                          urun_tanim, ust_tanim, yol=CIKTI_DUZELTILMIS,
                          k6=None):
    """Tüm otomatik düzeltilebilir bulguları veriye uygular ve
    düzeltilmiş malzeme listelerini + değişiklik günlüğünü Excel'e yazar.

    Otomatik uygulananlar
      K1  alan uyuşmazlığı      -> ürün alanı üst kodun değeriyle eşitlenir
      K1  barkod işareti        -> üst kodun işaretiyle eşitlenir
      K1  68'li kodda üst kod   -> Temel malzeme boşaltılır
      K1  silme kaskadı         -> ürüne silme işareti konur
      K2  tanım/alan uyuşmazlığı-> alan doldurulur veya tanım düzeltilir
      K3  yasaklı kelime        -> kelime tanımdan çıkarılır
      K4  çokluk (tek MARM değeri varsa) -> tanımdaki xN düzeltilir
      K6  menşei uyuşmazlığı    -> Menşei kodun son iki hanesiyle eşitlenir

    Elle karar gerekenler (günlükte işaretlenir)
      K1  üst kod boş / listede yok
      K4  birden çok MARM değeri varsa
      K5  üst kod benzersizliği
    """
    u = urun.copy()
    p = ust.copy()
    gunluk = []

    def kaydet(kaynak, malzeme, kural, alan, eski, yeni, durum, aciklama=""):
        gunluk.append({
            "Kaynak": kaynak, "Malzeme": str(malzeme), "Kural": kural,
            "Alan": alan, "Eski Değer": "" if eski is None else str(eski),
            "Yeni Değer": "" if yeni is None else str(yeni),
            "Durum": durum, "Açıklama": aciklama})

    u_idx = {str(m).strip(): i for i, m in enumerate(u["Malzeme"])}
    p_idx = {str(m).strip(): i for i, m in enumerate(p["Malzeme"])}
    silme_kolonu = silme_isareti_kolonu(u)

    # ---------------- K1 ----------------
    if not k1.empty:
        for _, s in k1.iterrows():
            malzeme = str(s["Malzeme"]).strip()
            alan = str(s.get("Alan", ""))
            sorun = str(s.get("Sorun", ""))
            i = u_idx.get(malzeme)
            if i is None:
                continue
            if "üst kodu olmamalı" in sorun:
                eski = u.at[u.index[i], "Temel malzeme"]
                u.at[u.index[i], "Temel malzeme"] = ""
                kaydet("Ürün Kodu", malzeme, "K1", "Temel malzeme", eski, "",
                       "DÜZELTİLDİ", "68 ile başlayan kodun üst kod bağı kaldırıldı")
            elif silme_kolonu and alan == silme_kolonu:
                if normalize(u.at[u.index[i], silme_kolonu]) == "":
                    u.at[u.index[i], silme_kolonu] = "X"
                    kaydet("Ürün Kodu", malzeme, "K1", silme_kolonu, "", "X",
                           "DÜZELTİLDİ", "Üst kod silinecek — ürüne silme işareti konuldu")
                else:
                    kaydet("Ürün Kodu", malzeme, "K1", silme_kolonu, "X", "X",
                           "ELLE", "Silme işlemi SAP tarafında tamamlanmalı")
            elif alan == "Temel malzeme":
                kaydet("Ürün Kodu", malzeme, "K1", "Temel malzeme",
                       s.get("Ürün Değeri", ""), "", "ELLE",
                       "Geçerli üst kod iş birimi tarafından belirlenmeli")
            elif alan in u.columns:
                ust_kod = str(u.at[u.index[i], "Temel malzeme"]).strip()
                j = p_idx.get(ust_kod)
                if j is None:
                    continue
                eski = u.at[u.index[i], alan]
                yeni = p.at[p.index[j], alan]
                u.at[u.index[i], alan] = yeni
                kaydet("Ürün Kodu", malzeme, "K1", alan, eski, yeni,
                       "DÜZELTİLDİ", "Üst kodun değeriyle eşitlendi")

    # ---------------- K2 ----------------
    if not k2.empty:
        for _, s in k2.iterrows():
            malzeme = str(s["Malzeme"]).strip()
            kaynak = str(s.get("Kaynak", "Ürün Kodu"))
            hedef, idx, tanim_kolonu = ((u, u_idx, urun_tanim)
                                        if kaynak == "Ürün Kodu"
                                        else (p, p_idx, ust_tanim))
            i = idx.get(malzeme)
            if i is None:
                continue
            onerilen = str(s.get("Önerilen Tanım", "") or "").strip()
            kod = str(s.get("Anahtar Kelime", "")).strip()
            alanlar = [a.split("=")[0].strip() for a in
                       str(s.get("Kontrol Edilen Alanlar", "")).split(";") if a]
            if onerilen and tanim_kolonu:
                eski = hedef.at[hedef.index[i], tanim_kolonu]
                if str(eski).strip() != onerilen:
                    hedef.at[hedef.index[i], tanim_kolonu] = onerilen
                    kaydet(kaynak, malzeme, "K2", tanim_kolonu, eski, onerilen,
                           "DÜZELTİLDİ", "Tanım programın önerdiği hâle getirildi")
            else:
                for alan in alanlar:
                    if alan in hedef.columns and kod:
                        eski = hedef.at[hedef.index[i], alan]
                        hedef.at[hedef.index[i], alan] = kod
                        kaydet(kaynak, malzeme, "K2", alan, eski, kod,
                               "DÜZELTİLDİ", "Tanımdaki koda göre alan dolduruldu")
                        break

    # ---------------- K3 ----------------
    if not k3.empty:
        for _, s in k3.iterrows():
            malzeme = str(s["Malzeme"]).strip()
            kaynak = str(s.get("Kaynak", "Ürün Kodu"))
            hedef, idx, tanim_kolonu = ((u, u_idx, urun_tanim)
                                        if kaynak == "Ürün Kodu"
                                        else (p, p_idx, ust_tanim))
            i = idx.get(malzeme)
            if i is None or not tanim_kolonu:
                continue
            eski = str(hedef.at[hedef.index[i], tanim_kolonu])
            yeni = eski
            for kelime in str(s.get("Yasaklı Kelimeler", "")).split(","):
                k = kelime.strip()
                if not k:
                    continue
                if k in YASAKLI_SADECE_SONDA:
                    yeni = re.sub(rf"[\s\-_/.]*{re.escape(k)}\s*$", "", yeni,
                                  flags=re.IGNORECASE)
                else:
                    yeni = re.sub(rf"(?<![\wÇĞİÖŞÜçğıöşü]){re.escape(k)}"
                                  rf"(?![\wÇĞİÖŞÜçğıöşü])", "", yeni,
                                  flags=re.IGNORECASE)
            yeni = re.sub(r"\s{2,}", " ", yeni).strip()
            if yeni and yeni != eski:
                hedef.at[hedef.index[i], tanim_kolonu] = yeni
                kaydet(kaynak, malzeme, "K3", tanim_kolonu, eski, yeni,
                       "DÜZELTİLDİ", "Yasaklı kelime tanımdan çıkarıldı")

    # ---------------- K4 ----------------
    if not k4.empty:
        for _, s in k4.iterrows():
            malzeme = str(s["Malzeme"]).strip()
            i = u_idx.get(malzeme)
            degerler = [d.strip().lstrip("x") for d in
                        str(s.get("MARM Çokluk", "")).split(",") if d.strip()]
            eski_tanim = str(s.get("Tanım", ""))
            if i is None or not urun_tanim:
                continue
            if len(degerler) == 1 and degerler[0].isdigit():
                yeni = re.sub(r"([Xx*])(\d{1,4})\s*$",
                              lambda m: m.group(1) + degerler[0], eski_tanim)
                if yeni != eski_tanim:
                    u.at[u.index[i], urun_tanim] = yeni
                    kaydet("Ürün Kodu", malzeme, "K4", urun_tanim, eski_tanim,
                           yeni, "DÜZELTİLDİ",
                           "Çokluk MARM'daki tek sayaç değeriyle eşitlendi")
                    continue
            kaydet("Ürün Kodu", malzeme, "K4", "Çokluk (xN)",
                   f"x{s.get('Tanımdaki Çokluk', '')}",
                   str(s.get("MARM Çokluk", "")), "ELLE",
                   "MARM'da birden çok sayaç var — doğrusu iş birimince seçilmeli")

    # ---------------- K5 ----------------
    if not k5.empty:
        for _, s in k5.iterrows():
            kaydet("Üst Kod", s["Malzeme"], "K5", "Ayırt edici alanlar",
                   str(s.get("Ayırt Edici Değerler", "")),
                   "", "ELLE",
                   f"Aynı kombinasyon: {s.get('Eşleşen Kodlar', '')} — "
                   f"hangi kodun kalacağı iş birimince belirlenmeli")

    # ---------------- K6 ----------------
    # Malzeme kodu SAP'de sabittir; bu yüzden düzeltme Menşei alanına
    # uygulanır: alan, kodun son iki hanesiyle eşitlenir.
    if k6 is not None and not k6.empty and "Menşei" in u.columns:
        u_idx = {str(v).strip(): i for i, v in u["Malzeme"].items()}
        for _, s6 in k6.iterrows():
            kod = str(s6["Malzeme"]).strip()
            beklenen = str(s6.get("Kod Son 2 Hane", "")).strip()
            eski = str(s6.get("Menşei Değeri", ""))
            if not beklenen or "tanımlı bir menşei" in str(s6.get("Sorun", "")):
                kaydet("Ürün Kodu", kod, "K6", "Menşei", eski, "", "ELLE",
                       "Kodun son iki hanesi tanımlı menşei değil — kontrol edin")
                continue
            i = u_idx.get(kod)
            if i is not None:
                u.at[i, "Menşei"] = beklenen
            kaydet("Ürün Kodu", kod, "K6", "Menşei", eski, beklenen,
                   "DÜZELTİLDİ",
                   "Menşei, malzeme kodunun son iki hanesiyle eşitlendi")

    gunluk_df = pd.DataFrame(gunluk, columns=[
        "Kaynak", "Malzeme", "Kural", "Alan", "Eski Değer", "Yeni Değer",
        "Durum", "Açıklama"])

    ozet = pd.DataFrame({
        "Ölçüt": ["Toplam işlenen bulgu", "Otomatik düzeltilen",
                  "Elle karar gereken", "Düzeltilen ürün kodu",
                  "Düzeltilen üst kod"],
        "Değer": [
            len(gunluk_df),
            int((gunluk_df["Durum"] == "DÜZELTİLDİ").sum()) if len(gunluk_df) else 0,
            int((gunluk_df["Durum"] == "ELLE").sum()) if len(gunluk_df) else 0,
            gunluk_df.loc[(gunluk_df["Durum"] == "DÜZELTİLDİ") &
                          (gunluk_df["Kaynak"] == "Ürün Kodu"),
                          "Malzeme"].nunique() if len(gunluk_df) else 0,
            gunluk_df.loc[(gunluk_df["Durum"] == "DÜZELTİLDİ") &
                          (gunluk_df["Kaynak"] == "Üst Kod"),
                          "Malzeme"].nunique() if len(gunluk_df) else 0,
        ]})

    with pd.ExcelWriter(yol, engine="openpyxl") as w:
        ozet.to_excel(w, sheet_name="Özet", index=False)
        gunluk_df.to_excel(w, sheet_name="Değişiklik_Günlüğü", index=False)
        (gunluk_df[gunluk_df["Durum"] == "ELLE"]
         if len(gunluk_df) else gunluk_df).to_excel(
            w, sheet_name="Elle_Karar_Gerekenler", index=False)
        u.to_excel(w, sheet_name="Düzeltilmiş_Ürün_Kodları", index=False)
        p.to_excel(w, sheet_name="Düzeltilmiş_Üst_Kodlar", index=False)

    _duzeltilmis_bicimle(yol)
    return u, p, gunluk_df


def _duzeltilmis_bicimle(yol):
    """Düzeltilmiş veri dosyasına panelle aynı görsel kimliği verir."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(yol)
    ince = Side(style="thin", color="E4E9EE")
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        if ws.max_row < 1:
            continue
        for c in ws[1]:
            c.font = Font(bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1B2A3A")
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        ws.row_dimensions[1].height = 26
        basliklar = [str(c.value) for c in ws[1]]
        durum_i = basliklar.index("Durum") + 1 if "Durum" in basliklar else 0
        for r in range(2, ws.max_row + 1):
            for j in range(1, ws.max_column + 1):
                c = ws.cell(row=r, column=j)
                c.border = Border(bottom=ince)
                if r % 2 == 1:
                    c.fill = PatternFill("solid", fgColor="F7FAFB")
            if durum_i:
                d = ws.cell(row=r, column=durum_i)
                if str(d.value) == "DÜZELTİLDİ":
                    d.font = Font(color="0E8A6E", bold=True)
                elif str(d.value) == "ELLE":
                    d.font = Font(color="C0392B", bold=True)
        for j, ad in enumerate(basliklar, start=1):
            harf = get_column_letter(j)
            if ad in ("Açıklama", "Eski Değer", "Yeni Değer"):
                ws.column_dimensions[harf].width = 34
            elif "metin" in ad.lower() or "tanım" in ad.lower():
                ws.column_dimensions[harf].width = 32
            else:
                ws.column_dimensions[harf].width = 15
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = (f"A1:{get_column_letter(ws.max_column)}"
                                  f"{ws.max_row}")
    wb.save(yol)


# =====================================================================
# ANA AKIŞ
# =====================================================================

def dosya_al(etiket: str, varsayilan: str) -> str:
    """Dosya çalışma dizininde varsa yolunu döndürür, yoksa boş string.
    Colab'da dosya yüklemesi İSTENMEZ: panel açılır, yükleme oradan yapılır."""
    import os
    if os.path.exists(varsayilan):
        return varsayilan
    if COLAB:
        return ""          # panel üzerinden yüklenecek
    raise FileNotFoundError(
        f"'{varsayilan}' bulunamadı. Dosyayı script ile aynı klasöre koyun "
        f"veya AYARLAR bölümündeki adı düzeltin.")


def sureli(ad, fn, *args, **kw):
    import time
    t0 = time.perf_counter()
    sonuc = fn(*args, **kw)
    print(f"  {ad}: {time.perf_counter()-t0:.2f} sn")
    return sonuc


def bos_panel_uret(dosya: str):
    """Veri olmadan, yalnızca yükleme ekranı çalışan bir panel üretir."""
    bos_ozet = pd.DataFrame({
        "Kural": ["K1 - Hatalı Eşleşme", "K2 - Tanım/Ek Alan Uyuşmazlığı",
                  "K3 - Yasaklı Kelime", "K4 - Çokluk / MARM",
                  "K5 - Üst Kod Benzersizliği", "K6 - Kod Sonu / Menşei"],
        "Bulgu Sayısı": [0] * 6,
        "Etkilenen Malzeme Sayısı": [0] * 6,
    })
    bos_durum = pd.DataFrame(columns=[
        "Kaynak", "Malzeme", "Tanım", "K1 Akıllı Kod", "K2 Tanım/Ek Alan",
        "K3 Yasaklı Kelime", "K4 Çokluk/MARM", "K5 Üst Kod Benzersizliği",
        "K6 Kod Sonu/Menşei",
        "Alt Tarif (bilgi)", "Hatalı Kural Sayısı", "Genel Durum"])
    bos_basit = pd.DataFrame(columns=[
        "Malzeme Kısa Metni", "Malzeme", "Üst Kod", "Kural", "Hata Nedeni",
        "İlgili Alan", "Mevcut Değer", "Olması Gereken", "Alt Tarif Sayısı",
        "Yapılacak Düzeltme"])
    bos_pbi = pd.DataFrame(columns=[
        "Kural", "Kaynak", "Malzeme", "Sorun", "Alan",
        "Ürün Değeri", "Üst Kod Değeri", "Malzeme türü"])
    gecmis = None
    try:
        import os as _o
        if _o.path.exists(GECMIS_DOSYASI):
            gecmis = pd.read_excel(GECMIS_DOSYASI)
    except Exception:
        gecmis = None
    dashboard_uret(bos_ozet, bos_durum, pd.DataFrame(), bos_pbi, bos_basit,
                   dosya, gecmis=gecmis)


def panel_indir(dosya: str):
    """Paneli kullanıcının bilgisayarına indirir ve tek satır bilgi yazar.
    Ekranda başka hiçbir çıktı gösterilmez."""
    import sys
    try:
        colab_files.download(dosya)
    except Exception:
        pass
    # Gizlenen stdout'tan bağımsız olarak doğrudan gerçek çıktıya yaz
    sys.__stdout__.write("Dashboard successfully downloaded to your computer.\n")
    sys.__stdout__.flush()


def main():
    import os as _os0
    urun_yol = dosya_al("ÜRÜN KODLARI", URUN_DOSYASI)
    ust_yol = dosya_al("ÜST KODLAR", UST_DOSYASI)

    # Colab'da veri dosyası yoksa: analiz yapılmaz, boş panel açılır ve
    # kullanıcı verisini doğrudan panelden yükler.
    if not urun_yol and not ust_yol and not _os0.path.exists(BIRLESIK_DOSYA):
        bos_panel_uret(CIKTI_DASHBOARD)
        panel_indir(CIKTI_DASHBOARD)
        return

    urun = dosya_oku(urun_yol) if urun_yol else pd.DataFrame()
    ust = dosya_oku(ust_yol) if ust_yol else pd.DataFrame()

    if VERI_ONIZLEME:
        for ad, df in (("ÜRÜN KODLARI", urun), ("ÜST KODLAR", ust)):
            print(f"\n--- {ad}: sütun eşleşme kontrolü (ilk satır) ---")
            if df.empty:
                print("  (dosya boş)")
                continue
            ilk = df.iloc[0]
            for kolon in df.columns:
                deger = ilk[kolon]
                deger = "" if pd.isna(deger) else str(deger)
                print(f"  {kolon:<28} = {deger}")
        print()

    ayarlar_yukle()
    istisna = istisna_yukle()

    import os as _os
    if _os.path.exists(BIRLESIK_DOSYA):
        _hepsi = dosya_oku(BIRLESIK_DOSYA)
        _u, _p, _disi = birlesik_ayristir(_hepsi)
        if not _u.empty or not _p.empty:
            urun, ust = _u, _p
            urun_yol = ust_yol = BIRLESIK_DOSYA
            print(f"Bilgi: birleşik liste kullanıldı ({BIRLESIK_DOSYA}) — "
                  f"{len(urun):,} ürün kodu + {len(ust):,} üst kod"
                  + (f", {_disi} kayıt tanınmadı" if _disi else ""))
            urun_tanim = tanim_sutunu_bul(urun, urun_yol)
            ust_tanim = tanim_sutunu_bul(ust, ust_yol)

    urun, urun_disi = kapsam_filtrele(urun, GECERLI_URUN_TURLERI, "ÜRÜN KODLARI")
    ust, ust_disi = kapsam_filtrele(ust, GECERLI_UST_TURLERI, "ÜST KODLAR")
    kapsam_disi = urun_disi + ust_disi

    urun_tanim = tanim_sutunu_bul(urun, urun_yol)
    ust_tanim = tanim_sutunu_bul(ust, ust_yol)

    print("Kural 1 çalışıyor (akıllı kod eşleşme kontrolü)...")
    k1 = sureli("süre", kural1, urun, ust, haric_tanim_kolonlari=[c for c in (urun_tanim, ust_tanim) if c])

    print("Kural 2 çalışıyor (tanım <-> ek veri alanı tutarlılığı)...")
    k2 = pd.concat([
        kural2(urun, urun_tanim, "Ürün Kodu"),
        kural2(ust, ust_tanim, "Üst Kod"),
    ], ignore_index=True)

    print("Kural 3 çalışıyor (yasaklı kelimeler)...")
    k3 = pd.concat([
        kural3(urun, urun_tanim, "Ürün Kodu"),
        kural3(ust, ust_tanim, "Üst Kod"),
    ], ignore_index=True)

    print("Alt tarifler taranıyor (aynı tanım, farklı kod — bilgi)...")
    alt = pd.concat([
        alt_tarifler(urun, urun_tanim, "Ürün Kodu"),
        alt_tarifler(ust, ust_tanim, "Üst Kod"),
    ], ignore_index=True)

    print("Kural 4 çalışıyor (çokluk ↔ MARM adet sayacı)...")
    marm = marm_oku()
    _ka, _tek, _cift = _kural2_kod_haritalari()
    k4 = kural4(urun, ust, marm, urun_tanim, _tek, _cift)

    print("Kural 5 çalışıyor (üst kodların ayırt edici kombinasyonu)...")
    k5 = kural5(ust, marm, urun, ust_tanim, _tek, _cift)

    print("Kural 6 çalışıyor (kod son 2 hane ↔ Menşei)...")
    k6 = kural6(urun, urun_tanim)


    # Silinmiş malzemeleri (ÜB dzy.silme iştr. dolu) tüm bulgulardan düş
    silinmis = silinmis_malzemeler(urun, ust)
    if silinmis:
        toplam_dusen = 0
        for _ad in ("k1", "k2", "k3", "k4", "k5", "k6", "alt"):
            _df = locals()[_ad]
            _yeni, _n = silinmis_dus(_df, silinmis)
            if _ad == "k1": k1 = _yeni
            elif _ad == "k2": k2 = _yeni
            elif _ad == "k3": k3 = _yeni
            elif _ad == "k4": k4 = _yeni
            elif _ad == "k5": k5 = _yeni
            elif _ad == "k6": k6 = _yeni
            else: alt = _yeni
            toplam_dusen += _n
        print(f"Bilgi: {len(silinmis):,} malzeme silinmiş olarak işaretli "
              f"(ÜB dzy.silme iştr.) — {toplam_dusen:,} bulgu listeden düşüldü.")

    # Onaylı istisnaları düş
    atlananlar = []
    k1, a1 = istisna_uygula(k1, "K1", istisna); atlananlar.append(a1)
    k2, a2 = istisna_uygula(k2, "K2", istisna); atlananlar.append(a2)
    k3, a3 = istisna_uygula(k3, "K3", istisna); atlananlar.append(a3)
    k4, a4 = istisna_uygula(k4, "K4", istisna); atlananlar.append(a4)
    k5, a5 = istisna_uygula(k5, "K5", istisna); atlananlar.append(a5)
    k6, a6k = istisna_uygula(k6, "K6", istisna); atlananlar.append(a6k)
    alt, a6 = istisna_uygula(alt, "ALT", istisna); atlananlar.append(a6)
    atlanan_df = pd.concat([a for a in atlananlar if not a.empty],
                           ignore_index=True) if any(not a.empty for a in atlananlar) else pd.DataFrame()
    if not atlanan_df.empty:
        print(f"  Bilgi: {len(atlanan_df)} bulgu onaylı istisna olarak listeden düşüldü.")
    # ---- Power BI için tek tablo (long format) ----
    pbi_parcalar = []
    if not k1.empty:
        t = k1.copy(); t["Kural"] = "K1 - Hatalı Eşleşme"; t["Kaynak"] = "Ürün Kodu"
        pbi_parcalar.append(t[["Kural", "Kaynak", "Malzeme", "Sorun", "Alan",
                               "Ürün Değeri", "Üst Kod Değeri"]])
    if not k2.empty:
        t = k2.copy(); t["Kural"] = "K2 - Tanım/Ek Alan Uyuşmazlığı"
        t["Alan"] = t["Anahtar Kelime"]; t["Ürün Değeri"] = t["Tanım"]
        t["Üst Kod Değeri"] = t["Kontrol Edilen Alanlar"]
        pbi_parcalar.append(t[["Kural", "Kaynak", "Malzeme", "Sorun", "Alan",
                               "Ürün Değeri", "Üst Kod Değeri"]])
    if not k3.empty:
        t = k3.copy(); t["Kural"] = "K3 - Yasaklı Kelime"
        t["Alan"] = t["Yasaklı Kelimeler"]; t["Ürün Değeri"] = t["Tanım"]
        t["Üst Kod Değeri"] = ""
        pbi_parcalar.append(t[["Kural", "Kaynak", "Malzeme", "Sorun", "Alan",
                               "Ürün Değeri", "Üst Kod Değeri"]])
    if not k4.empty:
        t = k4.copy(); t["Kural"] = "K4 - Çokluk / MARM"
        t["Alan"] = "Çokluk (xN)"
        t["Ürün Değeri"] = "x" + t["Tanımdaki Çokluk"].astype(str)
        t["Üst Kod Değeri"] = t["MARM Çokluk"]
        pbi_parcalar.append(t[["Kural", "Kaynak", "Malzeme", "Sorun", "Alan",
                               "Ürün Değeri", "Üst Kod Değeri"]])
    if not k5.empty:
        t = k5.copy(); t["Kural"] = "K5 - Üst Kod Benzersizliği"
        t["Alan"] = "Ayırt edici alanlar"
        t["Ürün Değeri"] = t["Ayırt Edici Değerler"]
        t["Üst Kod Değeri"] = t["Eşleşen Kodlar"]
        pbi_parcalar.append(t[["Kural", "Kaynak", "Malzeme", "Sorun", "Alan",
                               "Ürün Değeri", "Üst Kod Değeri"]])
    if not k6.empty:
        t = k6.copy(); t["Kural"] = "K6 - Kod Sonu / Menşei"
        t["Alan"] = "Menşei"
        t["Ürün Değeri"] = t["Menşei Değeri"]
        t["Üst Kod Değeri"] = t["Beklenen Menşei"]
        pbi_parcalar.append(t[["Kural", "Kaynak", "Malzeme", "Sorun", "Alan",
                               "Ürün Değeri", "Üst Kod Değeri"]])
    if not alt.empty:
        t = alt.copy(); t["Kural"] = "Bilgi - Alt Tarifler"
        t["Alan"] = "Tanım"; t["Ürün Değeri"] = t["Tanım"]
        t["Üst Kod Değeri"] = t["Paylaşan Kodlar"]
        pbi_parcalar.append(t[["Kural", "Kaynak", "Malzeme", "Sorun", "Alan",
                               "Ürün Değeri", "Üst Kod Değeri"]])

    pbi = (pd.concat(pbi_parcalar, ignore_index=True)
           if pbi_parcalar else pd.DataFrame(
               columns=["Kural", "Kaynak", "Malzeme", "Sorun", "Alan",
                        "Ürün Değeri", "Üst Kod Değeri"]))

    # Power BI'da malzeme özellikleriyle kesişim için ana veri de eklenir
    # (Marka, Pazar, Menşei gibi alanlarla slicer/filtre yapabilmek için)
    ozellik_kolonlari = [c for c in ["Malzeme", "Malzeme türü", "Mal grubu",
                                     "Raporlama Markası", "Raporlama Alt Markası",
                                     "Pazar", "Menşei", "Ek Alan", "Ambalaj Tipi",
                                     "S&OP Kategorisi", "SKU Grup"]
                         if c in urun.columns]
    if ozellik_kolonlari:
        ozellikler = pd.concat([urun[ozellik_kolonlari],
                                ust[[c for c in ozellik_kolonlari if c in ust.columns]]],
                               ignore_index=True).drop_duplicates("Malzeme")
        pbi = pbi.merge(ozellikler, on="Malzeme", how="left")

    # ---- Özet ----
    ozet = pd.DataFrame({
        "Kural": ["K1 - Hatalı Eşleşme", "K2 - Tanım/Ek Alan Uyuşmazlığı",
                  "K3 - Yasaklı Kelime", "K4 - Çokluk / MARM",
                  "K5 - Üst Kod Benzersizliği", "K6 - Kod Sonu / Menşei"],
        "Bulgu Sayısı": [len(k1), len(k2), len(k3), len(k4), len(k5), len(k6)],
        "Etkilenen Malzeme Sayısı": [
            k1["Malzeme"].nunique() if not k1.empty else 0,
            k2["Malzeme"].nunique() if not k2.empty else 0,
            k3["Malzeme"].nunique() if not k3.empty else 0,
            k4["Malzeme"].nunique() if not k4.empty else 0,
            k5["Malzeme"].nunique() if not k5.empty else 0,
            k6["Malzeme"].nunique() if not k6.empty else 0,
        ],
    })

    # ---- K1 alan bazında döküm ----
    if not k1.empty:
        k1_alan = (k1.groupby("Alan").agg(
            Bulgu=("Malzeme", "size"),
            Malzeme_Sayısı=("Malzeme", "nunique"))
            .sort_values("Bulgu", ascending=False).reset_index())
        print("\n--- K1: Alan bazında bulgu dökümü ---")
        print(k1_alan.to_string(index=False))
        print("Not: Bir alan tek başına binlerce bulgu üretiyorsa ve iş")
        print("     kuralı gereği ürün/üst kod arasında farklı olması")
        print("     normalse, o alanı ISTISNA_ALANLAR listesine ekleyin.")
    else:
        k1_alan = pd.DataFrame()

    # ---- Genel durum tablosu: her malzeme için DOĞRU/HATALI özeti ----
    durum_parcalar = []
    for kaynak, df, tanim_kol in (("Ürün Kodu", urun, urun_tanim),
                                   ("Üst Kod", ust, ust_tanim)):
        t = pd.DataFrame({"Kaynak": kaynak, "Malzeme": df["Malzeme"]})
        t["Tanım"] = df[tanim_kol] if tanim_kol else ""
        durum_parcalar.append(t)
    durum = pd.concat(durum_parcalar, ignore_index=True)
    # Aynı malzeme kodu birden fazla satırda geçebilir (örn. iki üst koda
    # bağlı ürün); durum tablosunda malzeme başına tek satır tutulur.
    durum = durum.drop_duplicates(subset=["Malzeme"], keep="first").reset_index(drop=True)
    if silinmis:
        durum = durum[~durum["Malzeme"].astype(str).str.strip().isin(silinmis)]
        durum = durum.reset_index(drop=True)

    k1_set = set(k1["Malzeme"]) if not k1.empty else set()
    k2_set = set(k2["Malzeme"]) if not k2.empty else set()
    k3_set = set(k3["Malzeme"]) if not k3.empty else set()
    k4_set = set(k4["Malzeme"].astype(str)) if not k4.empty else set()
    k5_set = set(k5["Malzeme"].astype(str)) if not k5.empty else set()
    k6_set = set(k6["Malzeme"].astype(str)) if not k6.empty else set()
    alt_set = set(alt["Malzeme"].astype(str)) if not alt.empty else set()

    def _durum(m, hatali_set, uygulanir=True):
        if not uygulanir:
            return "-"
        return "HATALI" if m in hatali_set else "DOĞRU"

    # K1 yalnızca ürün kodlarına uygulanır; K2-K4 tanım varsa uygulanır
    durum["K1 Akıllı Kod"] = [
        _durum(m, k1_set, uygulanir=(k == "Ürün Kodu"))
        for m, k in zip(durum["Malzeme"], durum["Kaynak"])]
    tanim_var = {"Ürün Kodu": bool(urun_tanim), "Üst Kod": bool(ust_tanim)}
    for ad, s in (("K2 Tanım/Ek Alan", k2_set), ("K3 Yasaklı Kelime", k3_set)):
        durum[ad] = [_durum(m, s, uygulanir=tanim_var[k])
                     for m, k in zip(durum["Malzeme"], durum["Kaynak"])]
    # K4 (çokluk) ve K5 (üst kod tekilliği) yalnızca ürün kodlarına uygulanır
    durum["K4 Çokluk/MARM"] = [
        _durum(m, k4_set, uygulanir=(k == "Ürün Kodu"))
        for m, k in zip(durum["Malzeme"].astype(str), durum["Kaynak"])]
    durum["K5 Üst Kod Benzersizliği"] = [
        _durum(m, k5_set, uygulanir=(k == "Üst Kod"))
        for m, k in zip(durum["Malzeme"].astype(str), durum["Kaynak"])]
    # K6 (kod sonu ↔ menşei) yalnızca ürün kodlarına uygulanır
    durum["K6 Kod Sonu/Menşei"] = [
        _durum(m, k6_set, uygulanir=(k == "Ürün Kodu"))
        for m, k in zip(durum["Malzeme"].astype(str), durum["Kaynak"])]
    # Alt tarif hata değil, bilgi: sağlık hesabına dahil edilmez
    durum["Alt Tarif (bilgi)"] = [
        "ALT TARİF" if m in alt_set else "-"
        for m in durum["Malzeme"].astype(str)]

    kural_kolonlari = ["K1 Akıllı Kod", "K2 Tanım/Ek Alan",
                       "K3 Yasaklı Kelime", "K4 Çokluk/MARM",
                       "K5 Üst Kod Benzersizliği", "K6 Kod Sonu/Menşei"]
    durum["Hatalı Kural Sayısı"] = (durum[kural_kolonlari] == "HATALI").sum(axis=1)
    durum["Genel Durum"] = durum["Hatalı Kural Sayısı"].map(
        lambda n: "HATALI" if n > 0 else "DOĞRU")
    durum = durum.sort_values(["Genel Durum", "Hatalı Kural Sayısı", "Malzeme"],
                              ascending=[True, False, True])

    saglik = 100 * (durum["Genel Durum"] == "DOĞRU").mean()
    print(f"\nGenel veri sağlığı: %{saglik:.1f} "
          f"({(durum['Genel Durum']=='DOĞRU').sum()} DOĞRU / "
          f"{(durum['Genel Durum']=='HATALI').sum()} HATALI / "
          f"{len(durum)} toplam malzeme)")

    # ---- Excel çıktısı ----
    try:
        import xlsxwriter  # noqa: F401
        motor = "xlsxwriter"
    except ImportError:
        motor = "openpyxl"
    with pd.ExcelWriter(CIKTI_EXCEL, engine=motor) as w:
        ozet.to_excel(w, sheet_name="Özet", index=False)
        durum.to_excel(w, sheet_name="Genel_Durum", index=False)
        (k1 if not k1.empty else pd.DataFrame({"Bilgi": ["Bulgu yok"]})
         ).to_excel(w, sheet_name="K1_Hatalı_Eşleşme", index=False)
        if not k1_alan.empty:
            k1_alan.to_excel(w, sheet_name="K1_Alan_Özeti", index=False)
        (k2 if not k2.empty else pd.DataFrame({"Bilgi": ["Bulgu yok"]})
         ).to_excel(w, sheet_name="K2_Tanım_EkAlan", index=False)
        if not k2.empty and "Önerilen Tanım" in k2.columns:
            duzeltme = k2[k2["Önerilen Tanım"].astype(str).str.strip() != ""][
                ["Kaynak", "Malzeme", "Tanım", "Önerilen Tanım"]
            ].rename(columns={"Tanım": "Mevcut Tanım"})
            if not duzeltme.empty:
                duzeltme.to_excel(w, sheet_name="K2_Toplu_Düzeltme", index=False)
        (k3 if not k3.empty else pd.DataFrame({"Bilgi": ["Bulgu yok"]})
         ).to_excel(w, sheet_name="K3_Yasaklı_Kelime", index=False)
        (k5 if not k5.empty else pd.DataFrame({"Bilgi": ["Bulgu yok"]})
         ).to_excel(w, sheet_name="K5_Üst_Kod_Benzersizliği", index=False)
        (k6 if not k6.empty else pd.DataFrame({"Bilgi": ["Bulgu yok"]})
         ).to_excel(w, sheet_name="K6_Kod_Sonu_Menşei", index=False)
        (alt if not alt.empty else pd.DataFrame({"Bilgi": ["Kayıt yok"]})
         ).to_excel(w, sheet_name="Alt_Tarifler", index=False)
        (k4 if not k4.empty else pd.DataFrame({"Bilgi": ["Bulgu yok"]})
         ).to_excel(w, sheet_name="K4_Çokluk_MARM", index=False)
        if not atlanan_df.empty:
            atlanan_df.to_excel(w, sheet_name="Onaylı_İstisnalar", index=False)
        pbi.to_excel(w, sheet_name="PowerBI_Bulgular", index=False)

    tanim_map = {}
    for df_, tk in ((urun, urun_tanim), (ust, ust_tanim)):
        if tk:
            tanim_map.update(dict(zip(df_["Malzeme"].astype(str),
                                      df_[tk].fillna("").astype(str))))
    basit = basit_duzeltme_tablosu(k1, k2, k3, k4, k5, alt, urun, tanim_map,
                                   k6=k6)
    bicimli_bulgu_excel(basit, CIKTI_BULGU_XLSX)
    from openpyxl import load_workbook
    wb_sonuc = load_workbook(CIKTI_EXCEL)
    kategoriler = kategori_analiz_excel(urun, ust, durum, wb=wb_sonuc)
    kategori_sayfasi_duzenle(wb_sonuc, kategoriler)
    hazir_sayi = 0
    if not k2.empty and "Önerilen Tanım" in k2.columns:
        hazir_sayi = int((k2["Önerilen Tanım"].astype(str).str.strip() != "").sum())
    ozet_sayfasi_bicimle(wb_sonuc, ozet, durum, kapsam_disi=kapsam_disi,
                         istisna_n=len(atlanan_df), hazir=hazir_sayi)
    genel_durum_bicimle(wb_sonuc, durum)
    grafikler_sayfasi(wb_sonuc, ozet, durum, kategoriler)
    # Sayfa sırası: Grafikler · Özet · Genel_Durum · kural sayfaları · Kategori
    sira = ["Grafikler", "Charts", "Özet", "Summary", "Genel_Durum", "Overall"]
    hedef = 0
    for s in sira:
        if s in wb_sonuc.sheetnames:
            wb_sonuc.move_sheet(s, offset=hedef - wb_sonuc.sheetnames.index(s))
            hedef += 1
    wb_sonuc.save(CIKTI_EXCEL)
    _du, _dp, _gunluk = duzeltilmis_veri_uret(
        urun, ust, k1, k2, k3, k4, k5, urun_tanim, ust_tanim, k6=k6)
    print(f"Düzeltilmiş veri: {CIKTI_DUZELTILMIS} "
          f"({int((_gunluk['Durum'] == 'DÜZELTİLDİ').sum())} otomatik düzeltme, "
          f"{int((_gunluk['Durum'] == 'ELLE').sum())} elle karar)")

    gecmis_kaydet(ozet, durum, len(atlanan_df))
    _gecmis_df = None
    try:
        import os as _os2
        if _os2.path.exists(GECMIS_DOSYASI):
            _gecmis_df = pd.read_excel(GECMIS_DOSYASI)
    except Exception:
        _gecmis_df = None
    dashboard_uret(ozet, durum, k2, pbi, basit, CIKTI_DASHBOARD,
                   istisna_n=len(atlanan_df), kapsam_disi=kapsam_disi,
                   gecmis=_gecmis_df)

    ham_bulgular = basit_duzeltme_tablosu(k1, k2, k3, k4, k5, alt, urun, tanim_map, k6=k6,
                                          birlestir=False, k4_satirlari=True)
    powerbi_modeli_uret(ham_bulgular, durum, ozet, kategoriler, urun, ust,
                        basit=basit, kapsam_disi=kapsam_disi,
                        istisna_n=len(atlanan_df), hazir=hazir_sayi)
    powerbi_rehber_yaz()
    print(f"Power BI modeli: {CIKTI_PBI_MODEL} + {CIKTI_PBI_REHBER}")

    # Power BI şablonu (.pbit): veri, ilişki, ölçü ve görseller gömülü
    try:
        import powerbi_pbit_uret
        pbit, tema = powerbi_pbit_uret.modelden_uret(CIKTI_PBI_MODEL)
        tek = powerbi_pbit_uret.tek_tablo_uret(CIKTI_PBI_MODEL)
        print(f"Power BI şablonu: {pbit} (Desktop ile aç > Yükle)")
        print(f"  tarayıcı sürümü: {tek} · tema: {tema}")
    except Exception as e:
        print(f"Bilgi: .pbit şablonu üretilemedi ({e}); "
              f"{CIKTI_PBI_MODEL} + rehber ile elle kurulabilir.")

    print("\n===== ÖZET =====")
    print(ozet.to_string(index=False))
    print(f"\nExcel çıktısı : {CIKTI_EXCEL}")
    print(f"Bulgu Excel'i : {CIKTI_BULGU_XLSX} (biçimlendirilmiş, filtreli)")
    print(f"  (Kategori analizi ve grafikler aynı dosyanın içinde)")
    print(f"Dashboard     : {CIKTI_DASHBOARD} (tarayıcıda açın)")

    if COLAB:
        panel_indir(CIKTI_DASHBOARD)


if __name__ == "__main__":
    if COLAB:
        # Colab'da yalnızca panel görünür: analiz günlüğü ekrana basılmaz.
        import contextlib
        import io as _io
        _gunluk = _io.StringIO()
        try:
            with contextlib.redirect_stdout(_gunluk):
                main()
        except Exception:
            # Hata durumunda gizlenen günlük gösterilir ki sorun anlaşılsın
            print(_gunluk.getvalue())
            raise
    else:
        main()
